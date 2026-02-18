import copy
import itertools
import json

import math
import random
import re
from json import JSONDecodeError

import openpyxl
from openpyxl.utils import get_column_letter

from scipy.stats import norm

from rich.tree import Tree
from rich.console import Console
from rich.table import Table

from CoreEngine import Weapon, Spell, Monster, Player, Encounter, MonAction
from CoreEngine.DNDClasses import Barbarian, Bard, Cleric, Druid, Fighter, Monk, Paladin, Ranger, Rogue, Sorcerer
# from BackEndAPI import SimulationRequestsAPI

from datetime import date, datetime
import os

console = Console()

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "CoreEngine/data")

PLAYER_LIST_FILE = os.path.join(DATA_DIR, "player_list.json")
MONSTER_LIST_FILE = os.path.join(DATA_DIR, "monster_list.json")
ENCOUNTER_LIST_FILE = os.path.join(DATA_DIR, "encounter_list.json")
CONDITION_LIST_FILE = os.path.join(DATA_DIR, "condition_list.json")
SPELL_LIST_FILE = os.path.join(DATA_DIR, "spell_list.json")
STATUS_EFFECT_LIST_FILE = os.path.join(DATA_DIR, "status_effect_list.json")
WEAPONS_LIST_FILE = os.path.join(DATA_DIR, "weapons_list.json")
BASIC_ACTION_LIST_FILE = os.path.join(DATA_DIR, "basic_actions.json")

import sys

if "pytest" in sys.modules:
    TEST_MODE = True
else:
    TEST_MODE = False


#DEAD METHODS (to remove)
def characterSelectionMenu(partydata):
    print("Select desired character: ")
    i = 0
    for i in range(len(partydata)):
        print(f"{i + 1}. {partydata[i]["stats"]["name"]}")
    print(f"{i + 2}. Create a Character")
    print(f"{i + 3}. Exit")
def chooseWeapons(player):
    print("Choose weapons that", player.getName(), "uses.")
    print("Type 'Done' to continue.")
    weaponFile = open(WEAPONS_LIST_FILE)
    weaponData = json.load(weaponFile)
    for data in weaponData:
        print(data['name'])
    wChoice = input()
    while wChoice != "Done":
        found = False
        i = 0
        while not found and i < len(weaponData):
            if wChoice == weaponData[i]["name"]:
                found = True
            else:
                i += 1
        if found is False:
            wChoice = input("Invalid choice. Please select another weapon, or type 'Done' to continue: ")
        else:
            weaponName = weaponData[i]["name"]
            properties = weaponData[i]["properties"]
            damageDice = properties["damage"]
            diceNum = 0
            diceType = 0
            if len(damageDice) > 1:
                print("Versatile weapon: Choose between", damageDice[0], "and", damageDice[1])
                dChoice = input().split("d")
                diceNum = int(dChoice[0])
                diceType = int(dChoice[1])
            else:
                damageDice = damageDice[0].split("d")
                diceNum = int(damageDice[0])
                diceType = int(damageDice[1])
            damageType = properties["damageType"]
            weaponStat = properties["weaponStat"]
            if len(weaponStat) > 1:
                if player.getStat("STR") > player.getStat("DEX"):
                    weaponStat = weaponStat[0]
                    damMod = player.getMod(weaponStat)
                else:
                    weaponStat = weaponStat[1]
                    damMod = player.getMod(weaponStat)
            else:
                weaponStat = weaponStat[0]
                damMod = player.getMod(weaponStat)

            player.addWeapon(weaponName, weaponStat, diceNum, diceType, damageType, damMod)
            print("Choose weapons that", player.getName(), "uses.")
            print("Type 'Done' to continue.")
            wChoice = input()
    weaponFile.close()
def printSpells(spells):
    index = 0
    print("Cantrips: ")
    for lvl in range(0, 10):
        done = False
        while not done and index < len(spells):
            if spells[index]["level"] == lvl:
                print(spells[index]["spellname"])
                index += 1
            else:
                done = True
        if lvl != 9:
            print(f"\nLvl {lvl + 1} Spells: ")
def chooseSpells(player):
    spellFile = open(SPELL_LIST_FILE)
    spellData = json.load(spellFile)
    relevantSpellData = []
    playerCap = -1
    if player.getClass() == "cleric" or player.getClass() == "sorcerer" or player.getClass() == "wizard" or player.getClass() == "bard" or player.getClass() == "druid" or player.getClass() == "warlock":
        playerCap = math.ceil(player.getLevel() / 2) #Full casters
    elif (player.getClass() == "artificer" or player.getClass() == "paladin"
          or player.getClass() == "ranger" or player.getClass() == "fighter" or player.getClass() == "rogue"):
        playerCap = math.ceil(player.getLevel() / 3) #Half casters
    for spell in spellData:
        if spell["level"] <= playerCap:
            found = False
            i = 0
            while not found and i < len(spell["classes"]):
                if player.getClass().lower() == spell["classes"][i].lower():
                    relevantSpellData.append(spell)
                    found = True
                else:
                    i += 1

    printSpells(relevantSpellData)
    print("Type the name of the spell to add to your character")
    print("Type 'Spell Scroll' to choose a spell from other classes")
    print("Type 'Done' to continue.")
    sChoice = input()
    while sChoice != "Done":
        if sChoice == "Spell Scroll":
            printSpells(spellData)
            print("Type the name of the spell to add to your character")
            print("Type 'Done' to continue.")
            sChoice = input()
            while sChoice != "Done":
                chooseSpellsExecuteMenu(sChoice, spellData, player)

                print("Type 'Done' to continue.")
                sChoice = input()
        else:
            chooseSpellsExecuteMenu(sChoice, relevantSpellData, player)

            print("Type 'Spell Scroll' to choose a spell from other classes")
            print("Type 'Done' to continue.")
            sChoice = input()
    spellFile.close()
def choosePlayerStats():
    print("Enter Name: ")
    name = input()
    print("Enter Class: ")
    className = input()
    print("Enter level: ")
    level = int(input())
    print("Enter stat Array: ")
    stats = input("STR DEX CON INT WIS CHA")
    stats = stats.split()
    for i in range(len(stats)):
        stats[i] = int(stats[i])
    ac = int(input("What is the player's AC?"))
    hp = -1 #Schema: sets HP to max in Player object
    conImmunities = []
    activeStatusEffects = []
    activeConditions = []
    damImmunes = ""
    damResists = ""
    damVulns = ""

    player = Player(name, stats, {}, ac, hp, className, level, conImmunities,
                    damImmunes, damResists, damVulns, activeStatusEffects, activeConditions)
    return player
def chooseCharacter(partydata, filename):
    characterSelectionMenu(partydata)
    menuInput = int(input())
    while menuInput < 1 or menuInput > (len(partydata) + 2):
        print("Incorrect value! Please try again from the numbered list.")
        characterSelectionMenu(partydata)
        menuInput = int(input())
    repeat = True
    while repeat:
        found = False
        j = 0
        while not found and j < len(partydata):
            if (menuInput == (len(partydata) + 1) or menuInput == (len(partydata) + 2)
                    or partydata[(menuInput - 1)]["stats"]["name"] == partydata[j]["stats"]["name"]):
                found = True
            else:
                j += 1
        if found:
            repeat = False

    if menuInput == (len(partydata) + 1):  # Create a Character
        player = createCharacter()
        savePlayer(player, filename)
        return player
    elif menuInput != (len(partydata) + 2): #Not Exit
        player = getPlayerStats(partydata[menuInput - 1])
        getSavedWeapons(player, partydata[menuInput - 1])
        getSavedSpells(player, partydata[menuInput - 1])
        print(f"Chosen {partydata[menuInput - 1]["stats"]["name"]}!")
        return player
    else: #Chosen exit
        return None
def chooseEncounter():
    #prints all encounters saved in encounter_list.json, and returns the index of the chosen encounter
    try:
        with open(ENCOUNTER_LIST_FILE, "r") as f:
            encounters = json.load(f)
    except JSONDecodeError:
        return None
    except FileNotFoundError:
        return None
    if len(encounters) != 0:
        entries = []
        for encounter in encounters:
            if not encounter["completed"]:
                entries.append(f"{encounter['name']} - Date: {encounter['date']}")

        col_width = 32  # total width for each column, adjust as needed

        for i in range(0, len(entries), 4):
            chunk = entries[i:i + 4]
            # left-align each chunk inside the fixed width
            print("".join(f"{entry:<{col_width}}" for entry in chunk))

        encName = input("Please select the name of the encounter you'd like to load: ")
        found = False
        encIdx = 0
        while not found:
            idx = 0
            while not found and idx < len(encounters):
                if encName.lower() == encounters[idx]["name"].lower():
                    found = True
                    encIdx = idx
                else:
                    idx += 1
            if not found:
                print("Name not in encounters!")
                encName = input("Please select the name of the encounter you'd like to load: ")

        return encIdx
    else:
        return -1
def createEncounter():
    #Creates an encounter object based on user choices and saves it to encounter_list.json
    name = (input("\nPlease name the Encounter: "))
    encDate = (date.today().strftime("%Y-%m-%d"))
    encounter = Encounter(name, encDate)

    partySet = False
    playerFile = open(PLAYER_LIST_FILE)
    try:
        partydata = json.load(playerFile)
    except json.JSONDecodeError:
        partydata = []
    while not partySet and encounter.playerSize() < 4:
        if len(partydata) == 0:
            player = createCharacter()
            savePlayer(player, PLAYER_LIST_FILE)
        else:
            player = chooseCharacter(partydata, PLAYER_LIST_FILE)
        if player is not None:
            encounter.addPlayer(player)
        elif player is not None and encounter.playerSize() == 4:
            print("Ending player selection : Max party size is 4.")
            partySet = True
        else:
            partySet = True
    playerFile.close()

    printMonsterChoices()
    monChoice = input("Choose any monsters from this list to add to encounter: \nType 'Exit' to exit.")
    with open(MONSTER_LIST_FILE, "r") as f:
        monsters = json.load(f)
    numAdded = 0
    while monChoice != "Exit" or numAdded == 0:
        low = 0
        idx = 0
        high = len(monsters) - 1
        found = False

        while low <= high and not found: #quick binary search for monster at given index
            idx = (low + high) // 2
            if monsters[idx]["name"] == monChoice:
                found = True  # Target found, return its index
            elif monsters[idx]["name"] < monChoice:
                low = idx + 1  # Search in the right half
            else:
                high = idx - 1  # Search in the left half

        if found:
            monChoice += str(numAdded)
            cr = monsters[idx]["cr"]
            cType = monsters[idx]["creatureType"]
            stats = monsters[idx]["statArray"]
            hp = monsters[idx]["hit_points"]
            maxHP = hp
            ac = monsters[idx]["AC"]
            saveProfs = monsters[idx]["saveProfs"]
            lResists = monsters[idx]["lResists"]
            damResists = monsters[idx]["damResists"]
            damImmunes = monsters[idx]["damImmunes"]
            damVulns = monsters[idx]["damVulns"]
            conImmunes = monsters[idx]["conImmunes"]
            lairAction = False if monsters[idx]["lairAction"] in [False, "False", "false"] else True
            enemy = True
            actions = loadMonsterActions(monsters[idx])
            spellInfo = loadMonsterSpells(monsters[idx])
            magicResist = monsters[idx].get("magicResist", False)
            legActions = monsters[idx].get("legActions", [])

            encounter.addMonster(Monster(monChoice, cr, cType, stats, hp, maxHP,
                                         ac, saveProfs, lResists, damResists,
                                         damImmunes, damVulns, conImmunes, [],
                                        [], lairAction, magicResist,
                                         enemy, actions, spellInfo, legActions))
            print("Monster added to encounter!")
            numAdded += 1
        else:
            print("Unknown Monster! Please try again.")
        monChoice = input("Choose any monsters from this list to add to encounter: \nType 'Exit' to exit.")
        while monChoice == "Exit" and numAdded == 0:
            print("Must add atleast one monster to encounter! Please try again.")
            monChoice = input("Choose any monsters from this list to add to encounter: \nType 'Exit' to exit.")

    setInitiative(encounter)
    print("Encounter created!")
    f.close()
    return encounter
def setInitiative(encounter):
    iValues = []

    for i in range(encounter.playerSize()):
        print(encounter.getPlayer(i).getName())
        iValue = int(input("Initiative value: "))
        iValues.append({
            'name': encounter.getPlayer(i).getName(),
            'iValue': iValue,
            'turnType': 'Player',
            'currentTurn': False
        })
    for i in range(encounter.monsterSize()):
        print(encounter.getMonster(i).getName())
        iValue = int(input("Initiative value: "))
        iValues.append({
            'name': encounter.getMonster(i).getName(),
            'iValue': iValue,
            'turnType': 'Monster',
            'currentTurn': False
        })
        if encounter.getMonster(i).hasLairAction():
            iValues.append({
                'name': encounter.getMonster(i).getName() + " LAIR",
                'iValue': 20,
                'turnType': 'LairAction',
                'currentTurn': False
            })

    def sort_key(c):
        if c['turnType'] == 'Player':
            # Find the player object with this name
            for i in range(encounter.playerSize()):
                if encounter.getPlayer(i).getName() == c['name']:
                    return c['iValue'], encounter.getPlayer(i).getStat("DEX")
        else:
            # Find the monster object with this name
            for i in range(encounter.monsterSize()):
                if encounter.getMonster(i).getName() == c['name']:
                    return c['iValue'], encounter.getMonster(i).getStat("DEX")
        # Fallback in case no match (shouldn’t happen)
        return c['iValue'], 0 #Deals with initiative ties.

    initiative = sorted(iValues, key=sort_key, reverse=True)

    initiative[0]["currentTurn"] = True
    for c in initiative:
        c["actionResource"] = 1
        c["bonusActionResource"] = 1
    encounter.setInitiative(initiative)

    return initiative
def applyScaling(action, upcastLevel):
    #All possible values include:
    # XdY, XdY PER2, X, extraTarget, extraTarget2 for probability stuff
    #mult PER2, 1CR, 1CR PER2 for summons
        #NOTE: Any PER2 scales should be concluded before calling this method, not after.

    baseLevel = action.getLvl()
    scaling = action.getScaling()
    scaled_action = copy.deepcopy(action)  # safe copy

    if "d" in scaling: #XdY
        parts = scaling.split()
        base = parts[0]  # e.g. "1d6"
        num, die = map(int, base.lower().split("d"))
        delta_levels = upcastLevel - baseLevel

        extraDice = delta_levels * num
        scaled_action.setDice(action.getDiceNum() + extraDice, action.getSides(), action.getDamMod())
    elif scaling.isdigit(): #X
        scaled_action.setDamMod(action.getDamMod() + scaling)
    elif "extraTarget" in scaling:
        if scaling[-1] == "2":
            scaled_action.setNumTarget(action.getNumTarget() + 2)
        else:
            scaled_action.setNumTarget(action.getNumTarget() + 1)
    elif "1CR" in scaling:
        statusEffects = action.getStatusEffects()
        for i in range(len(statusEffects)):
            if statusEffects[i]["name"] == "Summon":
                statusEffects[i]["effect"]["crCap"] += 1
            else:
                pass
        action.setStatusEffects(statusEffects)
    elif "mult" in scaling:
        statusEffects = action.getStatusEffects()
        for i in range(len(statusEffects)):
            if statusEffects[i]["name"] == "Summon":
                statusEffects[i]["effect"]["numSummons"] *= 2
            else:
                pass
        action.setStatusEffects(statusEffects)
    elif "extraSummon" in scaling:
        amt = int(scaling[-1]) if scaling[-1].isdigit() else 1
        statusEffects = action.getStatusEffects()
        for i in range(len(statusEffects)):
            if statusEffects[i]["name"] == "Summon":
                statusEffects[i]["effect"]["numSummons"] += amt
            else:
                pass
        action.setStatusEffects(statusEffects)
    scaled_action.setLingEffects([])
    scaled_action.setExtraEffect({})
    scaled_action.setLingSaves({})
    scaled_action.setScaling("")
    return scaled_action
def getRollResults(action, selectedTargets):
    rollResult = []
    if isinstance(action, Spell):
        onlyType = ""
        immType = ""
        cap = 0
        if action.getSpecialNotes():
            specialNotes = action.getSpecialNotes()
            if any("hpcap" in note.lower() for note in specialNotes):
                for note in specialNotes:
                    if "hpcap" in note.lower():
                        try:
                            cap = int(note.split("hpCap")[1])
                        except:
                            cap = 0
                        rollResults = []
                        for target in selectedTargets:
                            if target["Statblock"].getHP() < cap:
                                if action.getRollType().lower() != 'save':
                                    rollResults.append('y')
                                else:
                                    rollResults.append('n')
                            else:
                                if action.getRollType().lower() != 'save':
                                    rollResults.append('n')
                                else:
                                    rollResults.append('y')
                        return rollResults
            if any("only" in note.lower() for note in specialNotes):
                for note in specialNotes:
                    if "only" in note.lower():
                        try:
                            onlyType = note.lower().split("only")[0]
                        except:
                            onlyType = ""
            if any("immune" in note.lower() for note in specialNotes):
                for note in specialNotes:
                    if "immune" in note.lower():
                        try:
                            immType = note.lower().split("immune")[0]
                        except:
                            immType = ""

        # NOTE: Basic actions are counted as spells here.
        if len(selectedTargets) > 1:
            if action.getRollType().lower() in ["tohit", "onhit"]:
                for i, target in enumerate(selectedTargets):
                    target = target["Statblock"] if isinstance(target, dict) else target
                    if isinstance(target, Player) and immType == "humanoid" or (onlyType and onlyType != "humanoid"):
                        print(f"{target.getName()} is not applicable to {action.getName()}!")
                        roll = 'n'
                    elif isinstance(target, Monster) and target.getCreatureType() == immType or (onlyType and target.getCreatureType != onlyType):
                        print(f"{target.getName()} is not applicable to {action.getName()}!")
                        roll = 'n'
                    else:
                        print(f"TARGET {i + 1} OF {len(selectedTargets)}")
                        roll = input(f"{target.getName()} - Did the spell hit? (y/n/crit): ").lower()
                        while roll not in ["y", "n", "crit"]:
                            print("BAD INPUT")
                            roll = input(f"{target.getName()} - Did the spell hit? (y/n/crit): ").lower()
                    rollResult.append(roll)
            elif action.getRollType().lower() in ["save"]:
                for i, target in enumerate(selectedTargets):
                    target = target["Statblock"] if isinstance(target, dict) else target
                    if isinstance(target, Player) and immType == "humanoid" or (onlyType and onlyType != "humanoid"):
                        print(f"{target.getName()} is not applicable to {action.getName()}!")
                        roll = 'y'
                    elif isinstance(target, Monster) and target.getCreatureType() == immType or (onlyType and target.getCreatureType != onlyType):
                        print(f"{target.getName()} is not applicable to {action.getName()}!")
                        roll = 'y'
                    else:
                        print(f"TARGET {i + 1} OF {len(selectedTargets)}")
                        roll = input(f"{target.getName()} - Did the creature save? (y/n): ").lower()
                        while roll not in ["y", "n"]:
                            print("BAD INPUT")
                            roll = input(f"{target.getName()} - Did the creature save? (y/n): ").lower()
                    rollResult.append(roll)
            else:  # Autohit
                for i, target in enumerate(selectedTargets):
                    target = target["Statblock"] if isinstance(target, dict) else target
                    if isinstance(target, Player) and immType == "humanoid" or (onlyType and onlyType != "humanoid"):
                        print(f"{target.getName()} is not applicable to {action.getName()}!")
                        roll = 'n'
                    elif isinstance(target, Monster) and  target.getCreatureType() == immType or (onlyType and target.getCreatureType != onlyType):
                        print(f"{target.getName()} is not applicable to {action.getName()}!")
                        roll = 'n'
                    else:
                        roll = 'y'
                    rollResult.append(roll)
        else:
            try:
                target = selectedTargets[0]
            except:
                print("DEBUG")
            target = target["Statblock"] if isinstance(target, dict) else target
            if isinstance(target, Player) and immType == "humanoid" or (onlyType and onlyType != "humanoid"):
                print(f"{target.getName()} is not applicable to {action.getName()}!")
                roll = 'n' if action.getRollType().lower() != "save" else "y"
            elif isinstance(target, Monster) and target.getCreatureType() == immType or (onlyType and target.getCreatureType != onlyType):
                print(f"{target.getName()} is not applicable to {action.getName()}!")
                roll = 'n' if action.getRollType().lower() != "save" else "y"
            else:
                if action.getRollType().lower() in ["tohit", "onhit"]:
                    roll = input(f"{target.getName()} - Did the spell hit? (y/n/crit): ").lower()
                    while roll not in ["y", "n", "crit"]:
                        print("BAD INPUT")
                        roll = input(f"{target.getName()} - Did the spell hit? (y/n/crit): ").lower()
                elif action.getRollType().lower() in ["save"]:
                    roll = input(f"{target.getName()} - Did the creature save? (y/n): ").lower()
                    while roll not in ["y", "n"]:
                        print("BAD INPUT")
                        roll = input(f"{target.getName()} - Did the creature save? (y/n): ").lower()
                else: #Autohit
                    roll = 'y'
            rollResult.append(roll)
    else:  # Assuming weapon
        try:
            debug_target = selectedTargets[0]
        except:
            print("DEBUG")
        roll = input(f"{selectedTargets[0]["Statblock"].getName()} - Did the attack hit? (y/n/crit): ").lower()
        while roll not in ["y", "n", "crit"]:
            print("BAD INPUT")
            roll = input(f"{selectedTargets[0]["Statblock"].getName()} - Did the attack hit? (y/n/crit): ").lower()
        rollResult.append(roll)
    return rollResult
def getDiceResults(action, selectedTargets, rollResults):
    """
    Returns a list of damage values matching each selected target.
    Uses user-entered actual damage values, applying halving logic only if needed.

    rollResults: list of 'y'/'n' indicating success/failure per target
    """

    # No damage at all? Return 0 for each target.
    if action.getMean() == 0:
        return [0] * len(selectedTargets)


    rollType = action.getRollType().lower() if isinstance(action, Spell) else "tohit"
    isSaveSpell = (rollType == 'save')
    isAutoHit = (rollType == 'autohit')
    hasHalfOnSave = (isSaveSpell and action.getHalfSave())

    damages = []

    if isinstance(action.getDamType(), list) and action.getDamType() not in [0, 1]:
        #NOTE: Only spells with multi-damType are save spells.
        if action.getDamType()[-1] == "AND":
            if action.getNumTarget() in [-1, 1]:
                dams = [int(input(f"Input actual {action.getDamType()[i]} damage dealt")) for i in
                        range(len(action.getDamType()) - 1)]
                multiTarget = False
            else:
                multiTarget = True
            for i, target in enumerate(selectedTargets):
                target = target["Statblock"]
                if multiTarget: #Dams is not set - so set dams
                    dams = [int(input(f"Input actual {action.getDamType()[i]} damage dealt")) for i in
                        range(len(action.getDamType()) - 1)]
                for j in range(len(action.getDamType())):
                    if target.isImmune(action.getDamType()[j]):
                        dams[j] = 0
                    elif target.isVulnerable(action.getDamType()[j]):
                        if not target.isResistant(action.getDamType()[j]):
                            dams[j] *= 2
                    elif target.isResistant(action.getDamType()[j]):
                        dams[j] = math.floor(dams[j] / 2)
                dam = sum(dams)
                if hasHalfOnSave and rollResults[i] == 'y':
                    dam /= 2
                damages.append(dam)
        else:
            [print(action.getDamType()[i]) for i in range(len(action.getDamType()) - 1)]
            damType = input("Select Damage Type: ").lower()
            while damType not in action.getDamType():
                damType = input("Bad input! Please select from listed damage types: ").lower()
            if action.getNumTarget() in [-1, 1]:
                dams = int(input(f"Input actual {damType} damage dealt"))
                multiTarget = False
            else:
                multiTarget = True
            for i, target in enumerate(selectedTargets):
                if isinstance(target, dict):
                    target = target["Statblock"]
                if multiTarget:
                    dams = int(input(f"Input actual {damType} damage dealt"))
                if target.isImmune(damType):
                    dams = 0
                elif target.isVulnerable(damType):
                    if not target.isResistant(damType):
                        dams *= 2
                elif target.isResistant(damType):
                    dams = math.floor(dams / 2)
                if hasHalfOnSave and rollResults[i] == 'y':
                    dams /= 2
                damages.append(dams)
        return damages


    # ---- CASE 1: Spell applies the same roll to all targets (AutoHit or HalfSave AOE) ----
    if isAutoHit or hasHalfOnSave:
        # One damage roll, then adjust per target
        if action.getNumTarget() > 1: #Autohit multitargets
            for i in range(len(rollResults)):
                damages.append(int(input("Input actual damage value dealt: ")))
        else: #Autohit single/AOE, halfSaves
            baseDamage = int(input("Input actual damage value dealt: "))

            for result in rollResults:
                if isSaveSpell:
                    # 'y' means target succeeded save → gets half
                    dam = baseDamage / 2 if result.lower() == 'y' else baseDamage
                else:
                    # AutoHit: success always hits full damage unless explicitly failed (rare)
                    dam = baseDamage
                damages.append(int(dam))

        return damages

    # ---- CASE 2: Save spells WITHOUT half damage ----
    if isSaveSpell and not hasHalfOnSave:
        dam = int(input("Input actual damage value dealt: "))
        for result in rollResults:
            if result.lower() == 'y':
                damages.append(0)
            else:
                damages.append(dam)
        return damages

    # ---- CASE 3: Everything else (To-Hit spells, weapon attacks, etc.) ----
    # Each target may have different hit/miss result
    for i, result in enumerate(rollResults):
        if rollType.lower() in ['tohit', 'onhit'] and rollResults[i] == 'crit':
            dam = int(input("Input actual damage value dealt: "))
            damages.append(dam * 2)
        elif result.lower() == 'n':
            damages.append(0)
        else:
            dam = int(input("Input actual damage value dealt: "))
            damages.append(dam)

    return damages
def enemySetTargets(initiative, duplicateTargets):
    affectedCreatures = []
    if not TEST_MODE:
        affInput = input("Which creatures were affected? Please input one at a time. ('Exit' to exit)").lower()
    else:
        validTargets = []
        for creature in initiative:
            if (creature["turnType"] == "Player" or
                    (creature["turnType"] == "Monster" and creature["Statblock"].isActiveStatusEffect("SwitchSides"))
                    and not creature["Statblock"].isActiveCondition("Dead")
                    and not creature["Statblock"].isActiveCondition("Out of Combat")):
                validTargets.append(creature)
        affInput = input(f"Which creatures were affected? ({0}) ({[f"{target["name"]}" for target in validTargets]})").lower()
    tgtctr = 0
    while affInput != "exit":
        found = False
        idx = 0
        while not found and idx < len(initiative):  # find creature
            if initiative[idx]["name"].lower() == affInput.lower():
                if not duplicateTargets:
                    if affInput in affectedCreatures:
                        print("Invalid selection: Creature is already targeted!")
                        break
                if initiative[idx]["Statblock"].isActiveCondition("Dead") or initiative[idx]["Statblock"].isActiveCondition("Out Of Combat"):
                    print("Invalid selection: Creature cannot be targeted!")
                    break
                else:
                    creature = initiative[idx]
                    if (creature["turnType"] == "Player"
                            or (creature["turnType"] == "Monster" and creature["Statblock"].isActiveStatusEffect(
                                "SwitchSides"))
                            and not creature["Statblock"].isActiveCondition("Dead")
                            and not creature["Statblock"].isActiveCondition("Out of Combat")):
                        validTarget = True
                    else:
                        validTarget = False
                    if not validTarget:
                        print("Invalid selection: Creature cannot be targeted!")
                        break
                    else:
                        affectedCreatures.append(initiative[idx])
                        tgtctr += 1
                        found = True
            idx += 1
        if not found and idx == len(initiative):
            print("Creature not found!")
        if not TEST_MODE:
            affInput = input("Which creatures were affected? Please input one at a time. ('Exit' to exit)").lower()
        else:
            validTargets = []
            for creature in initiative:
                if (creature["turnType"] == "Player" or
                        (creature["turnType"] == "Monster" and creature["Statblock"].isActiveStatusEffect(
                            "SwitchSides"))
                        and not creature["Statblock"].isActiveCondition("Dead")
                        and not creature["Statblock"].isActiveCondition("Out of Combat")):
                    validTargets.append(creature)
            affInput = input(
                f"Which creatures were affected? ({tgtctr}) ({[f"{target["name"]}" for target in validTargets]})").lower()
    return affectedCreatures
def enemyTurnMANUAL(initiative, encounter):
    printCreatureNames(initiative)
    affectedCreatures = enemySetTargets(initiative, True)
    for creature in affectedCreatures:
        if not creature["Statblock"].isActiveCondition("Dead") \
            and not creature["Statblock"].isActiveCondition("Out of Combat"):
            printCreatureStats(affectedCreatures)
            nextCreature = False
            print(f"{creature['name']} SELECTION: ")
            while not nextCreature:
                print("1. Change HP")
                print("2. Change Conditions")
                print("3. Change Status Effects")
                print("4. Exit")
                damChoice = int(input("Choose enemy outcome"))
                while damChoice < 1 or damChoice > 4:
                    print("Incorrect input.")
                    damChoice=int(input())
                if damChoice == 1: #Change HP
                    damage = int(input("How much damage?"))
                    if damage > 0:
                        creature["Statblock"].setHP(creature["Statblock"].getHP() - damage)
                        if damage != 0 and any("concentration" in se["name"].lower() for se in creature["Statblock"].getActiveStatusEffects()):
                            print("Active concentration detected!")
                            removeConc = input("Did the creature lose concentration? (y/n)")
                            while removeConc not in ["y", "n"]:
                                print("BAD INPUT")
                                removeConc = input(f"Did the {creature['name']} lose concentration (y/n)")
                            if removeConc == "y":
                                concentration = {}
                                for se in creature["Statblock"].getActiveStatusEffects():
                                    if se["name"].lower() == "concentration":
                                        concentration = se
                                        break
                                endConcentration(creature, concentration, initiative)
                                if encounter.monsterSize() != len(initiative):
                                    summonCleanUp(encounter, initiative)
                        if creature["Statblock"].isActiveCondition("downed") and damage > 0:
                            removeCondition("downed", creature)
                            addCondition("dead", creature, -1)
                        elif creature["Statblock"].getHP() <= 0:
                            creature["Statblock"].setHP(0)
                            if creature["turnType"].lower() == "player":
                                addCondition("downed", creature, -1)
                            else:
                                addCondition("dead", creature, -1)
                    else:
                        print("Invalid: Cannot have negative damage!")
                elif damChoice == 2: #Change Conditions
                    endConditionChange = False
                    while not endConditionChange:
                        print("1. Add Condition")
                        print("2. Remove Condition")
                        print("3. Exit")
                        condChoice = int(input())
                        while condChoice < 1 or condChoice > 3:
                            print("Incorrect input.")
                            condChoice=int(input())
                        if condChoice == 1: #Add Condition
                            printConditions()
                            conditionToAdd = input("Which condition has been added?").lower()
                            validCondition = False
                            while not validCondition:
                                if enemyCanMutate(conditionToAdd, creature):
                                    if addCondition(conditionToAdd, creature, -1):
                                        validCondition = True
                                    else:
                                        if conditionToAdd == "exit":
                                            validCondition = True
                                if not validCondition:
                                    print("INVALID CHOICE")
                                    conditionToAdd = input("Which condition has been added? 'Exit' to exit").lower()
                        elif condChoice == 2: #Remove Condition
                            if len(creature["Statblock"].getActiveConditions()) != 0:
                                for condition in creature["Statblock"].getActiveConditions():
                                    if isinstance(condition, dict):
                                        if -1 in condition["resultID"]:
                                            print(condition["cond"])
                                    else:
                                        print(condition)
                                conditionToRemove = input("Which condition has been removed? 'Exit' to exit").lower()
                                validCondition = False
                                while not validCondition:
                                    if enemyCanMutate(conditionToRemove, creature):
                                        if removeCondition(conditionToRemove, creature):
                                            validCondition = True
                                        else:
                                            if conditionToRemove == "exit":
                                                validCondition = True
                                    if not validCondition:
                                        print("Creature does not have that condition!")
                                        conditionToRemove = input("Which condition has been removed? 'Exit' to exit").lower()
                            else:
                                print("Invalid input! No conditions exist.")
                        else:
                            endConditionChange = True
                elif damChoice == 3:
                    IGNORED = {"time stop", "transform", "summon"}

                    # Your allowed attributes by effect, per your schema
                    ATTRS_BY_EFFECT = {
                        # attack roll / save modifiers
                        "advantage": ["attack rolls for", "attack rolls against", "STR save", "DEX save", "CON save",
                                      "INT save", "WIS save", "CHA save", "ALL save"],
                        "disadvantage": ["attack rolls for", "attack rolls against", "STR save", "DEX save",
                                         "CON save", "INT save", "WIS save", "CHA save", "ALL save"],
                        "buff": ["attack rolls for", "attack rolls against", "STR save", "DEX save", "CON save",
                                 "INT save", "WIS save", "CHA save", "AC", "ALL save"],
                        "debuff": ["attack rolls for", "attack rolls against", "AC", "STR save", "DEX save",
                                   "CON save", "INT save", "WIS save", "CHA save", "ALL save"],
                        "autocrit": ["attack rolls against"],
                        "autofail": ["STR save", "DEX save", "CON save", "INT save", "WIS save", "CHA save", "ALL save"],

                        # score-based defenses
                        "resistance": ["fire", "psychic", "force", "cold", "poison", "acid", "thunder", "lightning", "radiant", "necrotic", "bludgeoning", "piercing", "slashing"],
                        "immunity": ["fire", "psychic", "force", "cold", "poison", "acid", "thunder", "lightning", "radiant", "necrotic", "bludgeoning", "piercing", "slashing"],
                        "vulnerability": ["fire", "psychic", "force", "cold", "poison", "acid", "thunder", "lightning", "radiant", "necrotic", "bludgeoning", "piercing", "slashing"],

                        # others typically have no attribute (but we’ll keep template behavior)
                        "concentration": [],
                        "switchsides": []
                    }

                    def _prompt_choice(title, options):
                        print(title)
                        for i, opt in enumerate(options, 1):
                            print(f"{i}. {opt}")
                        while True:
                            try:
                                n = int(input("> ").strip())
                                if 1 <= n <= len(options):
                                    return options[n - 1]
                            except ValueError:
                                pass
                            print("Enter a number from the list.")
                    def _build_status_effect(se_name, status_defs):
                        # find template
                        match = next((s for s in status_defs if s["name"].lower() == se_name.lower()), None)
                        if not match:
                            print("Unknown status effect.")
                            return None
                        if match["name"].lower() in IGNORED:
                            print("That effect is handled elsewhere.")
                            return None

                        effect = copy.deepcopy(match["effect"])

                        if "rolls" in effect:
                            default_roll = effect["rolls"]
                            if match["name"].lower() in ("buff", "debuff"):
                                user_roll = input(f"Enter roll for {match['name']} (default {default_roll}): ").strip()
                                if user_roll:
                                    effect["rolls"] = user_roll
                            elif default_roll in ("", None):
                                user_roll = input(f"Enter roll for {match['name']}: ").strip()
                                effect["rolls"] = user_roll or default_roll

                        if "damage" in effect and (effect["damage"] in ("", None)):
                            user_damage = input(f"Enter damage modifier for {match['name']} (e.g., /2, *2): ").strip()
                            if user_damage:
                                effect["damage"] = user_damage

                        if "value" in effect and effect["value"] == "T/F":
                            user_value = input(f"{match['name']} active? (True/False): ").strip()
                            effect["value"] = user_value

                        allowed_attrs = ATTRS_BY_EFFECT.get(match["name"].lower())
                        if allowed_attrs:
                            chosen_attrs = []
                            while True:
                                remaining = [a for a in allowed_attrs if a not in chosen_attrs]
                                if not remaining:
                                    print("All valid attributes already chosen.")
                                    break
                                chosen = _prompt_choice(
                                    f"Choose an attribute for {match['name']} (or type 0 to finish):",
                                    remaining + ["Finish selection"]
                                )
                                if chosen == "Finish selection":
                                    break
                                chosen_attrs.append(chosen)

                                again = input("Add another attribute? Y/N: ").strip().lower()
                                if again != "y":
                                    break

                            effect["attribute"] = chosen_attrs or []

                        else:
                            if "attribute" in effect and effect["attribute"]:
                                if isinstance(effect["attribute"], str):
                                    effect["attribute"] = [effect["attribute"]]

                        return {"name": match["name"], "effect": effect}
                    def modifyStatusEffects(creature):
                        with open(STATUS_EFFECT_LIST_FILE, "r") as f:
                            status_defs = json.load(f)
                        available = [s["name"] for s in status_defs if s["name"].lower() not in IGNORED]

                        endSEChange = False
                        while not endSEChange:
                            print("\n--- Status Effect Menu ---")
                            print("1. Add Status Effect")
                            print("2. Remove Status Effect")
                            print("3. Exit")
                            try:
                                seChoice = int(input().strip())
                            except ValueError:
                                seChoice = 0
                            if seChoice not in [1, 2, 3]:
                                print("Invalid input.")
                                continue

                            if seChoice == 1:
                                se_name = _prompt_choice("\nWhich status effect has been added? 'Exit' to exit", available).lower()
                                seToAdd = _build_status_effect(se_name, status_defs)
                                while not enemyCanMutate(seToAdd, creature) and se_name != "exit":
                                    print("INVALID INPUT")
                                    se_name = _prompt_choice("\nWhich status effect has been added? 'Exit' to exit", available).lower()
                                    seToAdd = _build_status_effect(se_name, status_defs)
                                if se_name == "exit":
                                    continue
                                if not seToAdd:
                                    continue

                                # Always let addStatusEffect handle merging logic
                                addStatusEffect(seToAdd, creature, -1)

                            elif seChoice == 2:
                                active = creature["Statblock"].getActiveStatusEffects()
                                if not active:
                                    print("No active status effects.")
                                    continue
                                to_remove = _prompt_choice(
                                    "\nSelect a status effect to remove: 'Exit' to exit",
                                    [e["name"] for e in active]
                                ).lower()
                                while not enemyCanMutate(to_remove, creature) and to_remove != "exit":
                                    print("INVALID INPUT")
                                    to_remove = _prompt_choice(
                                        "\nSelect a status effect to remove: 'Exit' to exit",
                                        [e["name"] for e in active]
                                    ).lower()
                                if to_remove == "exit":
                                    continue
                                removeStatusEffect(to_remove, creature)
                            else:
                                endSEChange = True

                    modifyStatusEffects(creature)
                else:
                    nextCreature = True
#PRINT METHODS (dead but may use to debug)
def printEncounterState(initiative):
    """
    Prints a Rich table snapshot of the current encounter initiative order.
    Includes players and monsters, with HP, AC, and conditions.
    """
    table = Table(title="Initiative Order", show_lines=True)

    table.add_column("Current Turn", no_wrap=True)
    table.add_column("Name", no_wrap=True)
    table.add_column("Type", no_wrap=True)
    table.add_column("HP")
    table.add_column("AC")
    table.add_column("Enemy")

    # initiative list: should contain players/monsters in turn order
    for iValue, creature in enumerate(initiative, start=1):
        if 'Statblock' in creature:
            # Get name / hp / ac
            name = creature['name']
            hp_str = f"{creature["Statblock"].getHP()}/{creature["Statblock"].getMaxHP()}"
            ac_str = str(creature["Statblock"].getAC())

            if creature["turnType"] == "Player":
                enemy_str = False
            elif any("switchsides" in se["name"].lower() for se in creature["Statblock"].getActiveStatusEffects()):
                enemy_str = False
            else:
                enemy_str = True

            if creature["currentTurn"]:
                table.add_row("True", name, creature["turnType"], hp_str, ac_str, str(enemy_str), style="bold")
            else:
                table.add_row("False", name, creature["turnType"], hp_str, ac_str, str(enemy_str))
        else:
            if creature["currentTurn"]:
                table.add_row("True", creature['name'], creature["turnType"], "", "", "True", style="bold")
            else:
                table.add_row("False", creature['name'], creature["turnType"], "", "", "True")
    console.print(table)
def printCreatureNames(initiative, againstPlayers=True):
    creatureTable = Table(title="Creatures in Initiative", show_lines=True, min_width=20)
    creatureTable.add_column("Name", justify="center")
    if againstPlayers:
        for creature in initiative:
            if (not "LAIR" in creature['name']
                    and (creature["turnType"] == "Player" or
                    (creature["turnType"] == "Monster" and creature["Statblock"].isActiveStatusEffect("SwitchSides"))
                and not creature["Statblock"].isActiveCondition("Dead")
                and not creature["Statblock"].isActiveCondition("Out of Combat"))):
                creatureTable.add_row(creature['name'])
    else:
        for creature in initiative:
            if not "LAIR" in creature['name'] \
                and not creature["Statblock"].isActiveCondition("Dead") \
                and not creature["Statblock"].isActiveCondition("Out of Combat"):
                creatureTable.add_row(creature['name'])
    console.print(creatureTable)
def printCreatureStats(affectedCreatures):
    creatureStatsTable = Table(title="Creature Stats", show_lines=True, min_width=20)
    creatureStatsTable.add_column("Name")
    creatureStatsTable.add_column("HP")
    creatureStatsTable.add_column("Conditions", no_wrap=True)
    creatureStatsTable.add_column("Status Effects", no_wrap=True)

    for creature in affectedCreatures:
        hp_str = f"{creature['Statblock'].getHP()}/{creature['Statblock'].getMaxHP()}"

        conds = creature["Statblock"].getActiveConditions()
        condsTOSTR = []
        for ci, c in enumerate(conds):
            if isinstance(c, dict):
                if -1 in c["resultID"]:
                    condsTOSTR.append(c["cond"].lower())
                else:
                    condsTOSTR.append(f"{c["cond"].lower()} - PLAYERMADE")
            else:
                condsTOSTR.append(c.lower())

        conds_str = "\n".join(condsTOSTR)

        if not creature["Statblock"].getActiveStatusEffects():
            statEffects = ""
        else:
            printableEffects = all("attribute" in effect and effect["attribute"] for effect in creature["Statblock"].getActiveStatusEffects())
            statEffects = Tree("Status Effects") if printableEffects else Tree("")
            for statEffect in creature["Statblock"].getActiveStatusEffects():
                effect = statEffect["effect"]
                if "attribute" in effect and effect["attribute"]:
                    node = statEffects.add(f"[green]{statEffect['name']}[/green]")
                    attrs = effect["attribute"]
                    if isinstance(attrs, list):
                        attr_text = ", ".join(attrs)
                    else:
                        attr_text = str(attrs)
                    node.add(f"[white]{attr_text}[/white]")
                    if -1 not in effect["resultID"]:
                        node.add(f"PLAYERMADE")

        creatureStatsTable.add_row(str(creature["name"]), hp_str, conds_str, statEffects)

    console.print(creatureStatsTable)
def printCreatureStats_DEBUG(affectedCreatures):
    creatureStatsTable = Table(title="Creature Stats", show_lines=True, min_width=20)
    creatureStatsTable.add_column("Name", max_width=10)
    creatureStatsTable.add_column("HP", min_width=5)
    creatureStatsTable.add_column("Conditions", no_wrap=True)
    creatureStatsTable.add_column("Status Effects", no_wrap=True, max_width=30)

    for creature in affectedCreatures:
        hp_str = f"{creature['Statblock'].getHP()}/{creature['Statblock'].getMaxHP()}"

        conds = creature["Statblock"].getActiveConditions()
        conds = [f"{c["cond"].lower()}\n - {[cID for cID in c["resultID"]]}" if isinstance(c, dict) else c.lower() for c in conds]
        conds_str = "\n".join(conds)

        if not creature["Statblock"].getActiveStatusEffects():
            statEffects = ""
        else:
            statEffects = Tree("Status Effects")
            for statEffect in creature["Statblock"].getActiveStatusEffects():
                node = statEffects.add(f"[green]{statEffect['name']}[/green]")
                effect = statEffect["effect"]

                if "attribute" in effect and effect["attribute"]:
                    attrs = effect["attribute"]
                    if isinstance(attrs, list):
                        attr_text = ", ".join(attrs)
                    else:
                        attr_text = str(attrs)
                    node.add(f"[white]{attr_text}[/white]")
                elif "spell" in effect and effect["spell"]:
                    spells = [spell["spellname"] for spell in effect["spell"]]
                    node.add(f"[white]{spells}[/white]")
                if "resultID" in effect and effect["resultID"]:
                    ids = effect["resultID"]
                    if isinstance(ids, list):
                        id_text = ", ".join([str(id) for id in ids])
                    else:
                        id_text = str(ids)
                    node.add(f"[white]{id_text}[/white]")


        creatureStatsTable.add_row(str(creature["name"]), hp_str, conds_str, statEffects)

    console.print(creatureStatsTable)
def printConditions():
    with open(CONDITION_LIST_FILE, "r") as f:
        condData = json.load(f)

    conds = Tree("[bold]All Conditions[/bold]")

    for condition in condData:
        cond_node = conds.add(f"[yellow]{condition['name']}[/yellow]")

        if condition["statusEffects"]:
            for status in condition["statusEffects"]:
                effect_node = cond_node.add(f"[green]{status['name']}[/green]")
                effect = status["effect"]

                if "attribute" in effect:
                    attributes = effect["attribute"]

                    if isinstance(attributes, list):
                        attr_text = ", ".join(attributes)
                    else:
                        attr_text = str(attributes)


                    effect_node.add(f"[white]{attr_text}[/white]")

        if condition["conditionEffects"]:
            cond_effects = cond_node.add("[magenta]Condition Effects[/magenta]")
            for effect in condition["conditionEffects"]:
                cond_effects.add(f"{effect}")

    console.print(conds)
    f.close()
def printStatusEffects():
    with open(STATUS_EFFECT_LIST_FILE, "r") as f:
        statusData = json.load(f)

    # Root of the tree
    statuses = Tree("[bold]All Status Effects[/bold]")

    for effectData in statusData:
        # Create the individual status node
        status_node = Tree(f"[green]{effectData['name']}[/green]")

        effect = effectData["effect"]

        # Only display attribute line if it exists and is not empty
        if "attribute" in effect and effect["attribute"]:
            attrs = effect["attribute"]

            if isinstance(attrs, list):
                attr_text = ", ".join(attrs)
            else:
                attr_text = str(attrs)

            status_node.add(f"[white]{attr_text}[/white]")

        # Add the effect node to the main tree with padding
        statuses.add(status_node)  # top/bottom spacing

    console.print(statuses)



#PLAYER/MONSTER/SPELL/WEAPON CREATE/SAVE/LOAD METHODS
def getPlayerStats(data):
    #Creates a player object based on pulled JSON player data.
    #TODO: Add new player attributes to this method.
    stats = data["stats"]
    saveProfs = stats["saveProfs"]
    saveProfs = {a : int(i) for a, i in saveProfs.items()}
    playerName = stats["name"]
    playerLvl = int(stats["level"])
    playerAC = int(stats["ac"])
    if "hp" in stats:
        playerHP = int(stats["hp"])
    else:
        playerHP = -1
    class_type = stats["characterClass"]
    playerStats = stats["statArray"]
    playerStats = {a : int(i) for a, i in playerStats.items()}

    conImmunities = stats["conImmunities"]
    activeStatusEffects = stats["activeStatusEffects"]
    activeConditions = stats["activeConditions"]

    damImmunes = stats["damImmunes"]
    damResists = stats["damResists"]
    damVulns = stats["damVulns"]

    cid = stats["cid"]
    position = stats["position"]

    player = Player(playerName, playerStats,  saveProfs,playerAC, playerHP,
                    class_type,playerLvl, conImmunities, damImmunes,
        damResists, damVulns, activeStatusEffects,
                    activeConditions, cid, position)
    return player
def getSavedWeapons(player, data):
    weapons = data["weapons"]
    for weapon in weapons:
        weaponName = weapon["name"]
        properties = weapon["properties"]
        diceProperties = properties["damage"].split("d")
        diceNum = int(diceProperties[0])
        diceType = int(diceProperties[1])
        damageType = properties["damageType"]
        damMod = player.getMod(properties["weaponStat"])
        statType = properties["weaponStat"]

        player.addWeapon(weaponName, statType, diceNum, diceType, damageType, damMod)
def getSavedSpells(player, data):
    spells = data["spells"]
    for spell in spells:
        # print(spell)
        spellName = spell["spellname"]
        spellLvl = int(spell["level"])

        target = spell["targeting"]
        if isinstance(target, list):
            target = target[0]
        targetNum = int(target["number"])
        damType = target["damType"]
        if len(damType) == 1:
            damType = damType[0]
        selfTarget = False
        if "self" in target and target["self"]:
            selfTarget = True
        conditions = None
        if "conditions" in target:
            conditions = target["conditions"]
        statusEffect = None
        if "statusEffect" in target:
            statusEffect = target["statusEffect"]
        lingEffect = None
        if "lingEffect" in target:
            lingEffect = target["lingEffect"]
        extraEffect = None
        if "extraEffect" in target:
            extraEffect = target["extraEffect"]
        lingSaves = None
        if "lingSave" in target:
            lingSaves = target["lingSave"]
        scaling = None
        if "scaling" in target:
            scaling = target["scaling"]
        specialNotes = False
        if "specialNotes" in target:
            specialNotes = target["specialNotes"]
        actionCost = target["actionCost"]

        rolls = target["rolls"]
        rollType = rolls["rollType"]
        saveType = rolls["saveType"]
        halfSave = rolls["halfSave"]
        damage = rolls["damage"]
        damageMod = 0
        if "damageMod" in rolls:  # Accounts for schema error
            if rolls["damageMod"] == "spellMod":
                damageMod = player.getSpellMod()
            elif rolls["damageMod"] != "":
                damageMod = int(rolls["damageMod"])
        diceNum = 0
        diceType = 0
        if damage and damage != "none":
            damage = damage.split("d")
            diceNum = int(damage[0])
            diceType = int(damage[1])

        # if spellLvl == 0:
        #     # Cantrips scale at lvls 5, 11, and 17.
        #     if scaling and "d" in scaling:
        #         if player.getLevel() >= 5:
        #             diceNum += 1
        #             if player.getLevel() >= 11:
        #                 diceNum += 1
        #                 if player.getLevel() >= 17:
        #                     diceNum += 1
        #     elif scaling and "extraTarget" in scaling:
        #         if player.getLevel() >= 5:
        #             targetNum += 1
        #             if player.getLevel() >= 11:
        #                 targetNum += 1
        #                 if player.getLevel() >= 17:
        #                     targetNum += 1

        player.addSpell(spellName, spellLvl, selfTarget,
                        targetNum, rollType, saveType, halfSave, damageMod, diceNum, diceType,
                        damType, conditions, statusEffect, lingEffect, extraEffect, lingSaves,
                        scaling, actionCost, specialNotes)
def addChosenSpell(spell, player):
        spellName = spell["spellname"]
        spellLvl = spell["level"]

        if isinstance(spell["targeting"], list) and len(spell["targeting"]) > 1: #Multiple possible effects
            for spellTarget in spell["targeting"]:
                newSpell = {
                    "spellname": spellTarget["targetType"],
                    "level": spellLvl,
                    "targeting": spellTarget
                }
                addChosenSpell(newSpell, player) #Adds the multiple types of effects as individual spells.
        else:
            if isinstance(spell["targeting"], list):
                targeting = spell["targeting"][0]
            else:
                targeting = spell["targeting" ]
            selfTarget = targeting["self"]
            targetNum = int(targeting["number"])
            damType = targeting["damType"] #TODO: Ensure that damType is always a list when saved.
            #TODO: Ensure that action costs are saved and loaded.
            if len(damType) == 1:
                damType = damType[0]
            conditions = None
            if len(targeting["conditions"]) != 0:
                conditions = targeting["conditions"]
            statusEffect = None
            if targeting["statusEffect"]:
                statusEffect = targeting["statusEffect"]
            lingEffect = None
            if targeting["lingEffect"]:
                lingEffect = targeting["lingEffect"]
            extraEffect = None
            if targeting["extraEffect"]:
                extraEffect = targeting["extraEffect"]
            lingSaves = None
            if targeting["lingSave"]:
                lingSaves = targeting["lingSave"]
            scaling = targeting["scaling"]
            specialNotes = None
            if len(targeting["specialNotes"]) != 0:
                specialNotes = targeting["specialNotes"]
            actionCost = targeting["actionCost"]

            spellRolls = targeting["rolls"]
            rollType = spellRolls["rollType"]
            saveType = spellRolls["saveType"]
            halfSave = spellRolls["halfSave"]
            damageDice = spellRolls["damage"]
            diceNum = 0
            diceType = 0
            if damageDice != "":
                damageDice = damageDice.split("d")
                diceNum = int(damageDice[0])
                diceType = int(damageDice[1])
            damageMod = 0
            if "damageMod" in spellRolls: #Accounts for schema error
                if spellRolls["damageMod"] == "spellMod":
                    damageMod = player.getSpellMod()
                elif spellRolls["damageMod"] != "":
                    damageMod = int(spellRolls["damageMod"])

            if spellLvl == 0:
                # Cantrips scale at lvls 5, 11, and 17.
                if scaling and "d" in scaling:
                    if player.getLevel() >= 5:
                        diceNum += 1
                        if player.getLevel() >= 11:
                            diceNum += 1
                            if player.getLevel() >= 17:
                                diceNum += 1
                elif scaling and "extraTarget" in scaling:
                    if player.getLevel() >= 5:
                        targetNum += 1
                        if player.getLevel() >= 11:
                            targetNum += 1
                            if player.getLevel() >= 17:
                                targetNum += 1

            player.addSpell(spellName, spellLvl, selfTarget,
                            targetNum, rollType, saveType, halfSave, damageMod, diceNum, diceType,
                            damType, conditions, statusEffect, lingEffect, extraEffect, lingSaves,
                            scaling, actionCost, specialNotes)
def chooseSpellsExecuteMenu(sChoice, data, player):
    idx = findSpell(sChoice, data)
    if idx != -1:
        spell = data[idx]
        addChosenSpell(spell, player)
        data.remove(spell)

        printSpells(data)
        print("Type the name of the spell to add to your character")
    else:
        print("Invalid spell. Please type the name of the spell within the list to add to your character.")
def findSpell(spellName, spellData):
    found = False
    i = 0
    while not found and i < len(spellData):
        if spellData[i]["spellname"] == spellName:
            found = True
        else:
            i += 1
    if found:
        return i
    else:
        return -1
def savePlayer(player, filename):
    #Adds a serialized player to an existing player_list JSON file.
    try:
        with open(filename, "r") as f:
            players = json.load(f)
    except FileNotFoundError:
        players = []
    except json.JSONDecodeError:
        players = []

    stats_dict = {
        "name": player.getName(),
        "level": str(player.getLevel()),
        "ac" : str(player.getAC()),
        "hp" : str(player.getHP()),
        "class": player.getClass(),
        "conImmunities": player.getConImmunities(),
        "activeStatusEffects": player.getActiveStatusEffects(),
        "activeConditions": player.getActiveConditions(),
        "saveProfs" : [str(player.getMod(stat)) for stat in ["STR", "DEX", "CON", "INT", "WIS", "CHA"]],
        "damImmunes": player.getDamImmunities(),
        "damResists": player.getDamResistances(),
        "damVulns": player.getDamVulnerabilities(),
        "statArray": [str(player.getStat(stat)) for stat in ["STR", "DEX", "CON", "INT", "WIS", "CHA"]],
        "spellSlots" : player.getSpellSlots()
    }

    spells_list = []
    for i in range(player.getSpellLength()):
        spells_list.append(player.getSpell(i).toDict())

    weapons_list = []
    for i in range(player.getWeaponLength()):
        weapons_list.append(player.getWeapon(i).toDict())

    player_dict = {
        "stats": stats_dict,
        "spells": spells_list,
        "weapons": weapons_list
    }

    players.append(player_dict)

    with open(filename, "w") as f:
        json.dump(players, f, indent=4)
def createCharacter():
    player = choosePlayerStats()
    chooseWeapons(player)

    className = player.getClass()
    if className.lower() == "barbarian" or className == "rogue" or className == "fighter" or className == "monk":
        print("Does this character have any spells? Please select Y/N: ")
        hasSpells = input()
    else:
        hasSpells = "Y"
    if hasSpells == "Y":
        chooseSpells(player)

    return player
def printMonsterChoices():
    with open(MONSTER_LIST_FILE, "r") as f:
        monsters = json.load(f)

    entries = [f"{monster['name']} - CR{monster['cr']}" for monster in monsters]

    col_width = 32  # total width for each column, adjust as needed

    for i in range(0, len(entries), 4):
        chunk = entries[i:i + 4]
        # left-align each chunk inside the fixed width
        print("".join(f"{entry:<{col_width}}" for entry in chunk))
def loadMonsterActions(monsterData):
    #Loads the actions of a given monster into python objects from JSON format.
    """ Action example:
    {
                "name": "Bite",
                "desc": "Melee Weapon Attack: +11 to hit, reach 10 ft., one target. Hit: 17 (2d10 + 6) piercing damage plus 4 (1d8) acid damage.",
                "actionRange": "10",
                "numTarget": 1,
                "shape": "",
                "rolls": {
                    "rollType": "toHit",
                    "saveType": "",
                    "halfSave": false,
                    "saveDC": "",
                    "damage": "2d10",
                    "attackBonus": 11,
                    "damMod": 6
                },
                "damType": [
                    "piercing"
                ],
                "conditions": [
                    ""
                ],
                "statusEffect": [],
                "lingEffect": {},
                "extraEffect": {},
                "lingSave": {},
                "actionCost": "action",
                "recharge": "",
                "specialNotes": [],
                "extraDamage": [
                    {
                        "dice": "1d8",
                        "mod": 0,
                        "type": "acid"
                    }
                ]
    """
    actionData = monsterData["actions"]
    actions = []
    for action in actionData:
        actionName = action["name"]
        # if isinstance(action["targeting"][0], list) and len(action["targeting"][0] > 1):
        # Multiple possible effects.
        # NOTE: This is not possible right now, as all current statblocks are NOT including actions with multiple possible effects.
        # It only takes the first possible effect from the list. If this is changed later, this code will be relevant.
        #     for actionTarget in action["targeting"][0]:
        #         newaction = {
        #             "actionname": actionTarget["targetType"],
        #             "level": actionLvl,
        #             "targeting": actionTarget
        #         }
        #         addChosenaction(newaction, player)  # Adds the multiple types of effects as individual actions.
        # else:
        numTarget = action.get("numTarget", -10)
        numTarget = int(action.get("number", 1)) if numTarget == -10 else int(numTarget)
        selfTarget = True if numTarget == 0 else False
        damType = action["damType"]
        if len(damType) == 1:
            damType = damType[0]
        actionCost = "action"
        recharge = action.get("recharge", [])
        actionRange = action["actionRange"] #Unused for now, comes into play in mapping system
        actionDesc = action["desc"]
        actionShape = action["shape"] #Unused for now, comes into play in mapping system
        extraDamage = action.get("extraDamage", [])
        #TODO: Ensure extra damage is added in and calculated in expectedDamage calcs later on.

        conditions = None
        if action["conditions"] is not None and len(action["conditions"]) != 0:
            conditions = action["conditions"]
        statusEffect = None
        if action["statusEffect"]:
            statusEffect = action["statusEffect"]
        lingEffect = None
        if action["lingEffect"]:
            lingEffect = action["lingEffect"]
        extraEffect = None
        if action["extraEffect"]:
            extraEffect = action["extraEffect"]
        lingSaves = None
        if action["lingSave"]:
            lingSaves = action["lingSave"]
        specialNotes = action.get("specialNotes", [])
        if specialNotes:
            specialNotes = action["specialNotes"]

        actionRolls = action["rolls"]
        rollType = actionRolls["rollType"]
        saveType = actionRolls["saveType"]
        halfSave = actionRolls["halfSave"]
        saveDC = actionRolls["saveDC"]
        damageDice = actionRolls["damage"]
        attackBonus = int(actionRolls["attackBonus"]) if "attackBonus" in actionRolls and actionRolls["attackBonus"] else ""
        diceNum = 0
        diceType = 0
        if damageDice != "":
            damageDice = damageDice.split("d")
            diceNum = int(damageDice[0])
            diceType = int(damageDice[1])
        damageMod = 0
        if (("damMod" in actionRolls and actionRolls["damMod"] != "") or
                ("damageMod" in actionRolls and actionRolls["damageMod"] != "")):  # Accounts for schema error
            damageMod = int(actionRolls.get("damageMod", 0))
            damageMod = int(actionRolls.get("damMod", 0)) if damageMod == 0 else damageMod
        actions.append(MonAction(actionName, actionDesc, selfTarget, numTarget, actionRange, actionShape,
                 rollType, saveType, saveDC, halfSave, damageMod, diceNum, diceType,
                 attackBonus, extraDamage, damType, conditions, statusEffect, lingEffect, extraEffect,
                 lingSaves, actionCost, recharge, specialNotes))
    return actions
def loadMonsterSpells(monsterData):
    """
        {
            "type": "wisdom",
            "DC": "18",
            "attackRoll": "10",
            "spells": [
                {
                    "name": "sacred flame"
                },
                {
                    "name": "spare the dying"
                },
                {
                    "name": "thaumaturgy"
                },
                {
                    "name": "command"
                },
                {
                    "name": "detect evil and good"
                },
                {
                    "name": "detect magic"
                },
                {
                    "name": "lesser restoration"
                },
                {
                    "name": "zone of truth"
                },
                {
                    "name": "dispel magic"
                },
                {
                    "name": "tongues"
                },
                {
                    "name": "banishment"
                },
                {
                    "name": "freedom of movement"
                },
                {
                    "name": "flame strike"
                },
                {
                    "name": "greater restoration"
                },
                {
                    "name": "heroes' feast"
                }
            ],
            "spellSlots": [
                [
                    "4",
                    "4"
                ],
                [
                    "3",
                    "3"
                ],
                [
                    "3",
                    "3"
                ],
                [
                    "3",
                    "3"
                ],
                [
                    "2",
                    "2"
                ],
                [
                    "1",
                    "1"
                ]
            ]
        },
    """
    def loadSpell(spell):
        spellName = spell["spellname"]
        spellLvl = spell["level"]

        if isinstance(spell["targeting"], list) and len(
                spell["targeting"]) > 1:  # Multiple possible effects
            for spellTarget in spell["targeting"]:
                newSpell = {
                    "spellname": spellTarget["targetType"],
                    "level": spellLvl,
                    "targeting": spellTarget
                }

                loadSpell(newSpell)  # Adds the multiple types of effects as individual spells.
        else:
            if isinstance(spell["targeting"], list):
                # TODO: Change so that "targeting" is always a list when saved, even with 1 option.
                targeting = spell["targeting"][0]
            else:
                targeting = spell["targeting"]
            selfTarget = targeting["self"]
            targetNum = int(targeting["number"])
            damType = targeting["damType"]  # TODO: Ensure that damType is always a list when saved.
            # TODO: Ensure that action costs are saved and loaded.
            if len(damType) == 1:
                damType = damType[0]
            conditions = None
            if len(targeting["conditions"]) != 0:
                conditions = targeting["conditions"]
            statusEffect = None
            if targeting["statusEffect"]:
                statusEffect = targeting["statusEffect"]
            lingEffect = None
            if targeting["lingEffect"]:
                lingEffect = targeting["lingEffect"]
            extraEffect = None
            if targeting["extraEffect"]:
                extraEffect = targeting["extraEffect"]
            lingSaves = None
            if targeting["lingSave"]:
                lingSaves = targeting["lingSave"]
            scaling = targeting["scaling"]
            specialNotes = None
            if len(targeting["specialNotes"]) != 0:
                specialNotes = targeting["specialNotes"]

            spellRolls = targeting["rolls"]
            rollType = spellRolls["rollType"]
            saveType = spellRolls["saveType"]
            halfSave = spellRolls["halfSave"]
            damageDice = spellRolls["damage"]
            diceNum = 0
            diceType = 0
            if damageDice != "":
                damageDice = damageDice.split("d")
                diceNum = int(damageDice[0])
                diceType = int(damageDice[1])
            damageMod = 0
            if "damageMod" in spellRolls:  # Accounts for schema error
                if spellRolls["damageMod"] != "":
                    damageMod = int(spellRolls["damageMod"])
            return {
                "spellname": spellName,
                "level": spellLvl,
                "targeting": {
                    "self": selfTarget,
                    "number": str(targetNum),
                    "rolls": {
                        "rollType": rollType,
                        "saveType": saveType,
                        "halfSave": halfSave,
                        "damage": f"{diceNum}d{diceType}",
                        "damMod" : damageMod
                    },
                    "damType": damType,
                    "conditions": conditions,
                    "statusEffect": statusEffect,
                    "lingEffect": lingEffect,
                    "extraEffect": extraEffect,
                    "lingSave": lingSaves,
                    "scaling": scaling,
                    "specialNotes": specialNotes
                }
            }
    #Converts string names to actual spells if they exist in the spell list.
    spellJSON = monsterData["spellInfo"]
    if not spellJSON:
        return {}
    spellType = spellJSON["type"]
    spellDC = spellJSON["DC"]
    spellAttack = spellJSON["attackRoll"]
    spells = spellJSON["spells"]
    with open(SPELL_LIST_FILE, "r") as rf:
        rawSpellData =json.load(rf)
        for i, spell in enumerate(spells):
            if spell["name"].lower() in [s["spellname"].lower() for s in rawSpellData]:
                spellIdx = [s["spellname"].lower() for s in rawSpellData].index(spell["name"].lower())
                spells[i] = loadSpell(rawSpellData[spellIdx])
    if "spellSlots" in spellJSON:
        spellSlots = spellJSON["spellSlots"]
    else:
        spellSlots = []
    spellInfo = {
        "type": spellType,
        "DC": spellDC,
        "attackRoll" : spellAttack,
        "spells" : spells,
        "spellSlots" : spellSlots,
    }
    return spellInfo

#ENCOUNTER CREATE/SAVE/LOAD METHODS
def loadEncounter(encounterData):
    #REFACTORING NOTES:
    #Uses encounterData from parameter instead of pulling it here.
    if encounterData["completed"]:
        print("FATAL ERROR: cannot load a completed encounter!")
        return None
    encounter = Encounter(encounterData["name"], encounterData["date"], encounterData["eid"])

    for playerJSON in encounterData["players"]:
        #TODO: refactor getPlayerStats
        playerObj = getPlayerStats(playerJSON)
        getSavedWeapons(playerObj, playerJSON)
        getSavedSpells(playerObj, playerJSON)
        encounter.addPlayer(playerObj)

    for monsterJSON in encounterData["monsters"]:
        #TODO: Add monster actions to this method.
        #name, cr, cType, stats, hp, maxHP, ac, saveProfs, lResists, damResists,
        #damImmunes, damVulns, conImmunes, lairAction, legAction
        name = monsterJSON["name"]
        cr = monsterJSON["cr"]
        cType = monsterJSON["creatureType"]
        stats = monsterJSON["statArray"]
        stats = {a: int(i) for a, i in stats.items()}
        hp = int(monsterJSON["hp"])
        maxHP = int(monsterJSON["maxHP"])
        ac = int(monsterJSON["ac"])
        saveProfs = {a : int(i) for a, i in monsterJSON["saveProfs"].items()}
        lResists = int(monsterJSON["lResists"])
        damResists = monsterJSON["damResists"]
        damImmunes = monsterJSON["damImmunes"]
        damVulns = monsterJSON["damVulns"]
        conImmunes = monsterJSON["conImmunes"]
        if "activeCons" in monsterJSON:
            acons = "activeCons"
            activeConditions = monsterJSON[acons]
        elif "activeConditions" in monsterJSON:
            acons = "activeConditions"
            activeConditions = monsterJSON[acons]
        else:
            activeConditions = []
        activeStatusEffects = monsterJSON["activeStatusEffects"]
        lairAction = bool(monsterJSON["lairAction"])
        magicResist = monsterJSON.get("magicResist", False)
        legActions = monsterJSON.get("legActions", [])
        enemy = bool(monsterJSON["enemy"])
        actions = loadMonsterActions(monsterJSON)
        spellInfo = loadMonsterSpells(monsterJSON)
        cid = monsterJSON.get("cid", "")
        position = monsterJSON.get("position", [0, 0])
        encounter.addMonster(Monster(name, cr, cType, stats, hp, maxHP,
                                     ac, saveProfs, lResists,damResists,
                                     damImmunes, damVulns, conImmunes, activeConditions,
                                     activeStatusEffects,lairAction, magicResist,
                                     enemy, actions, spellInfo, legActions, cid, position))

    for resultJSON in encounterData["results"]:
        encounter.addResult(resultJSON)

    encounter.setInitiative(encounterData["initiative"])
    return encounter
def saveEncounter(encounter):
    # Adds a serialized encounter to an existing encounter_list JSON file.
    #REFACTOR CHANGES: Removed inputs so that encounter always overwrites.
    try:
        with open(ENCOUNTER_LIST_FILE, "r") as f:
            encounterFile = json.load(f)
    except FileNotFoundError:
        encounterFile = []
    except json.JSONDecodeError:
        # File exists but is corrupted or partially written
        encounterFile = []

    duplicate = False
    overwrite = False
    idx = 0
    while not duplicate and idx < len(encounterFile):
        if encounter.getName() == encounterFile[idx]["name"]:
            duplicate = True
            overwrite = True
            encounterFile.remove(encounterFile[idx])
        idx += 1

    monster_list = []
    for i in range(encounter.monsterSize()):
        monster_list.append(encounter.getMonster(i).toDict())

    player_list = []
    for i in range(encounter.playerSize()):
        player = encounter.getPlayer(i)
        stats_dict = {
            "name": player.getName(),
            "level": str(player.getLevel()),
            "ac": str(player.getAC()),
            "hp": str(player.getHP()),
            "maxhp" : str(player.getMaxHP()),
            "cid" : str(player.getCID()),
            "position" : player.getPosition(),
            "characterClass": player.getClass(),
            "conImmunities": player.getConImmunities(),
            "activeStatusEffects": player.getActiveStatusEffects(),
            "activeConditions": player.getActiveConditions(),
            "saveProfs": {stat : str(player.getSaveProf(stat)) for stat in ["STR", "DEX", "CON", "INT", "WIS", "CHA"]},
            "damImmunes": player.getDamImmunities(),
            "damResists": player.getDamResistances(),
            "damVulns": player.getDamVulnerabilities(),
            "statArray": {stat : str(player.getStat(stat)) for stat in ["STR", "DEX", "CON", "INT", "WIS", "CHA"]},
            "spellSlots": player.getSpellSlots()
        }

        spells_list = []
        for j in range(player.getSpellLength()):
            spells_list.append(player.getSpell(j).toDict())

        weapons_list = []
        for j in range(player.getWeaponLength()):
            weapons_list.append(player.getWeapon(j).toDict())

        player_dict = {
            "stats": stats_dict,
            "spells": spells_list,
            "weapons": weapons_list
        }
        player_list.append(player_dict)

    result_list = [encounter.getResultByIdx(i) for i in range(encounter.resultSize())]

    name = encounter.getName()
    if duplicate and not overwrite:
        lastChar = name[len(name) - 1]
        if lastChar.isdigit() and name[len(name) - 2] == "_": #Overwrite the overwrite
            lastChar = int(lastChar) + 1
            name = name[:-1]
            name += str(lastChar)
        else: #Overwrite a non-overwrite
            name += "_0"

    encounter_dict = {
        "name": name,
        "date": encounter.getDate(),
        "eid" : encounter.getEID(),
        "completed" : encounter.isComplete(),
        "monsters": monster_list,
        "players" : player_list,
        "results": result_list,
        "initiative": encounter.getInitiative()
    }

    encounterFile.append(encounter_dict)
    try:
        with open(ENCOUNTER_LIST_FILE, "w") as f:
            json.dump(encounterFile, f, indent=4)
    except TypeError as e:
        print(e)
        console.print(encounterFile)

#ENEMY TURN METHODS
def addCondition(condToAdd, creature, resultID):
    if isinstance(creature, dict):
        creature = creature["Statblock"]
    with open(CONDITION_LIST_FILE, "r") as f:
        condData = json.load(f)
    for condition in condData:
        if condToAdd.lower() == condition['name'].lower(): #Find condition
            for activeCond in creature.getActiveConditions():
                match = False
                if isinstance(condToAdd, dict):
                    if isinstance(activeCond, dict):
                        if condToAdd["cond"].lower() == activeCond["cond"].lower():
                            match = True
                    else:
                        if condToAdd["cond"].lower() == activeCond.lower():
                            match = True
                elif isinstance(activeCond, dict):
                    if condToAdd.lower() == activeCond["cond"].lower():
                        match = True
                else:
                    if condToAdd.lower() == activeCond.lower():
                        match = True
                if match:
                    if isinstance(activeCond, dict):
                        if any(resultID == rID for rID in activeCond["resultID"]):
                            return False
                        activeCond["resultID"].append(resultID)
                        return True
                    return False

            condToAdd = {"cond" : condToAdd, "resultID" : [resultID]}
            creature.addCondition(condToAdd)
            return True
    print("Invalid input: Condition does not exist!")
    return False
def removeCondition(condToRemove, creature):
    if isinstance(creature, dict):
        creature = creature["Statblock"]
    if condToRemove.lower() == "dead":
        return False
    with open(CONDITION_LIST_FILE, "r") as f:
        condData = json.load(f)
    for condition in condData:
        if condToRemove.lower() == condition['name'].lower():
            return creature.removeCondition(condToRemove)
    return False
def addStatusEffect(effect, creature, resultID):
    effect = copy.deepcopy(effect)
    if "attribute" in effect["effect"]:
        effect["effect"]["attribute"] = ensureList(effect["effect"]["attribute"])
    if isinstance(creature, dict):
        creature = creature["Statblock"]
    activeStatusEffects = creature.getActiveStatusEffects()
    for activeStatus in activeStatusEffects:
        if activeStatus["name"].lower() == effect["name"].lower():
            if "attribute" in activeStatus["effect"]:
                activeStatus["effect"]["attribute"] = ensureList(activeStatus["effect"]["attribute"])
            if all(item in activeStatus["effect"]["attribute"] for item in effect["effect"]["attribute"]):
                #All attributes are the exact same.
                if not any(resultID == rID for rID in activeStatus["effect"]["resultID"]):
                    activeStatus["effect"]["attribute"].extend(effect["effect"]["attribute"])
                    for i in range(len(effect["effect"]["attribute"])):
                        activeStatus["effect"]["resultID"].append(resultID)
                    return True
                else:
                    return False
            elif any(item in activeStatus["effect"]["attribute"] for item in effect["effect"]["attribute"]):
                #Any of the attributes match.
                nonActiveAttr = [item if item not in activeStatus["effect"]["attribute"] else None for item in effect["effect"]["attribute"]]
                for attr in nonActiveAttr:
                    if attr is not None:
                        activeStatus["effect"]["attribute"].append(attr)
                for i in range(len(nonActiveAttr)):
                    activeStatus["effect"]["resultID"].append(resultID) if nonActiveAttr[i] is not None and resultID != -1 else False
                return True
            if effect["name"].lower() in ["lingeffect", "lingsave"]:
                #If there is a match AND the effect is lingEffect/lingSave
                if any(resultID == rID for rID in activeStatus["effect"]["resultID"]):
                    return False
                if effect["name"].lower() == "lingeffect":
                    ling = creature.getActiveStatusEffect("lingEffect")
                else:
                    ling = creature.getActiveStatusEffect("lingsave")
                ling["effect"]["spell"].extend(effect["spell"])
                ling["effect"]["resultID"].extend(effect["resultID"])
    effect["effect"]["resultID"] = [resultID]
    creature.addStatusEffect(effect)
    return True
def removeStatusEffect(name, creature):
    creature = creature["Statblock"] if isinstance(creature, dict) else creature
    for effect in creature.getActiveStatusEffects():
        if name.lower() == effect["name"].lower():
            return creature.removeStatusEffect(name)
def enemyCanMutate(newEffect, creature):
    if isinstance(creature, dict):
        creature = creature["Statblock"]
    effect = {}
    if "name" in newEffect:
        for e in creature.getActiveStatusEffects():
            if newEffect["name"].lower() == e["name"].lower():
                effect = e
    else:
        for c in creature.getActiveConditions():
            if isinstance(c, dict):
                if c["cond"].lower() == c["cond"].lower():
                        effect = c
            else:
                if c.lower() == newEffect.lower():
                    effect = c
    if not effect:
        return True

    if not isinstance(effect, dict):
        return True
    if "cond" in effect:
        if isinstance(effect["resultID"], list):
            if -1 in effect["resultID"]:
                return True
            return False
        else:
            if effect["resultID"] == -1:
                return True
        return False
    else:
        if not "effect" in effect or not "resultID" in effect["effect"]:
            return True
        if isinstance(effect["effect"]["resultID"], list):
            if -1 in effect["effect"]["resultID"]:
                return True
            return False
        else:
            if effect["effect"]["resultID"] == -1:
                return True
            return False
def endOfEncounter(initiative):
    allPlayersDead = True
    for playerTurns in initiative:
        if playerTurns["turnType"] == "Player":
            if not playerTurns["Statblock"].isActiveCondition("Dead") and not playerTurns["Statblock"].isActiveCondition("Out of Combat"):
                allPlayersDead = False
                break
    allMonstersDead = True
    for monsterTurn in initiative:
        if monsterTurn["turnType"] == "Monster":
            if not monsterTurn["Statblock"].isActiveCondition("Dead") and not monsterTurn["Statblock"].isActiveCondition("Out of Combat"):
                allMonstersDead = False
                break
    return allPlayersDead or allMonstersDead

#GENERAL HELPER METHODS
def translateBasicAction(creature, action):
        #Should work for both monsters and players
        actionName = action["spellname"]
        targeting = action["targeting"][0]
        selfTarget = targeting["self"]
        targetNum = int(targeting["number"])
        damType = targeting["damType"]
        if len(damType) == 1:
            damType = damType[0]
        conditions = None
        if len(targeting["conditions"]) != 0:
            conditions = targeting["conditions"]
        statusEffect = None
        if targeting["statusEffect"]:
            statusEffect = targeting["statusEffect"]
        lingEffect = None
        if targeting["lingEffect"]:
            lingEffect = targeting["lingEffect"]
        extraEffect = None
        if targeting["extraEffect"]:
            extraEffect = targeting["extraEffect"]
        lingSaves = None
        if targeting["lingSave"]:
            lingSaves = targeting["lingSave"]

        actionRolls = targeting["rolls"]
        rollType = actionRolls["rollType"]
        saveType = actionRolls["saveType"]
        halfSave = actionRolls["halfSave"]
        damageDice = actionRolls["damage"]
        diceNum = 0
        diceType = 0
        if damageDice != "":
            damageDice = damageDice.split("d")
            diceNum = int(damageDice[0])
            diceType = int(damageDice[1])
        damageMod = 0
        if "damageMod" in targeting:  # Accounts for schema error
            if actionRolls["damageMod"] == "spellmod" and isinstance(creature, Player):
                damageMod = creature.getSpellMod()
            elif actionRolls["damageMod"] != "":
                damageMod = int(actionRolls["damageMod"])

        if actionName.lower() == "grapple":
            saveDC = 8 + creature.getProfBonus() + creature.getMod("STR")
        elif actionName.lower() == "shove" or actionName.lower() == "hide":
            saveDC = 8 + creature.getProfBonus() + creature.getMod("DEX")
        else:
            saveDC = 0

        return MonAction(actionName, "", selfTarget,
                     targetNum, 5, "", rollType, saveType,
                         saveDC, halfSave, damageMod, diceNum, diceType,
                     "", [], damType, conditions,
                         statusEffect, lingEffect, extraEffect, lingSaves,
                         "action","", [])
def defineBasicActions(actionNames, actionTypes, actionProbs, actionEDams, actionImpacts, player, initiative):
    #Load basic actions
    try:
        with open(BASIC_ACTION_LIST_FILE, "r") as f: #Basic actions are hardcoded into basic_actions file.
            actions = json.load(f)
    except FileNotFoundError:
        actions = []
    except json.JSONDecodeError:
        actions = []

    grapple, shove, dodge = actions[0], actions[1], actions[2]
    grapple = translateBasicAction(player, grapple)
    shove = translateBasicAction(player, shove)
    dodge = translateBasicAction(player, dodge)

    grappleProb = calcTotalSaveProbability(player, grapple, initiative)  # Calculate probability of save
    grappleProb["probSuccess"] = 0 if grappleProb["probSuccess"] < 0 else grappleProb["probSuccess"]
    grappleProb["probSuccess"] = 1 if grappleProb["probSuccess"] > 1 else grappleProb["probSuccess"]
    probToStr = f"{grappleProb["probSuccess"]}"
    probToStr += f" - {grappleProb["probLingEffect"]}LE" if grappleProb["probLingEffect"] else ""
    probToStr += f" - {grappleProb['probExtraEffect']}EE" if grappleProb["probExtraEffect"] else ""
    probToStr += f" - {grappleProb['probLingSaves']}LS" if grappleProb["probLingSaves"] else ""
    grappleProb = probToStr

    shoveProb = calcTotalSaveProbability(player, shove, initiative)
    shoveProb["probSuccess"] = 0 if shoveProb["probSuccess"] < 0 else shoveProb["probSuccess"]
    shoveProb["probSuccess"] = 1 if shoveProb["probSuccess"] > 1 else shoveProb["probSuccess"]
    probToStr = f"{shoveProb["probSuccess"]}"
    probToStr += f" - {shoveProb["probLingEffect"]}LE" if shoveProb["probLingEffect"] else ""
    probToStr += f" - {shoveProb['probExtraEffect']}EE" if shoveProb["probExtraEffect"] else ""
    probToStr += f" - {shoveProb['probLingSaves']}LS" if shoveProb["probLingSaves"] else ""
    shoveProb = probToStr

    dodgeProb = 1.0

    grappleImpact = calcImpact(player, grapple, grappleProb, 0, initiative)  # Calculates the impact of these actions to the encounter.
    shoveImpact = calcImpact(player, shove, shoveProb, 0, initiative)
    dodgeImpact = calcImpact(player, dodge, dodgeProb, 0, initiative)

    actionNames.append(grapple.getName())
    actionNames.append(shove.getName())
    actionNames.append(dodge.getName())
    actionTypes.append("Basic")
    actionTypes.append("Basic")
    actionTypes.append("Basic")
    actionProbs.append(grappleProb)
    actionProbs.append(shoveProb)
    # dodgeProb is autoProc
    actionProbs.append(dodgeProb)
    # grapple, shove, dodgeProb has no expected damage.
    actionEDams.append(0)
    actionEDams.append(0)
    actionEDams.append(0)
    # actionImpacts.append(0)
    # actionImpacts.append(0)
    # actionImpacts.append(0)
    actionImpacts.append(grappleImpact)
    actionImpacts.append(shoveImpact)
    actionImpacts.append(dodgeImpact)
def scaledProbabilities(player, action, initiative):
    if action.getScaling():
        scaling_values = []
        step = 1
        if "per2" in action.getScaling().lower():
            step = 2
        for lvl in range(action.getLvl() + 1, 10, step):  # levels N+1 to 9
            scaled_spell = applyScaling(action, lvl)
            roll_type = scaled_spell.getRollType().lower()
            if roll_type == "tohit":
                prob = calcTotalToHitProbability(player, scaled_spell, initiative)
            elif roll_type == "save":
                prob = calcTotalSaveProbability(player, scaled_spell, initiative)
            elif roll_type == "autohit":
                prob = calcTotalAutoHitProbability(player, scaled_spell, initiative)
            elif roll_type == "onhit":
                prob = calcOnHitProbability(scaled_spell, [player.getWeapon(i) for i in range(player.getWeaponLength())], player, initiative)
            else:
                prob = -1
            scaling_values.append(prob["probSuccess"])

        scalarTree = Tree("Upcasted probabilities of Success: ")
        for i in range(len(scaling_values)):
            scalarTree.add(f"{i + 1} - ").add(scaling_values[i])
        console.print(scalarTree)

        avg_scaled_prob = sum(scaling_values) / len(scaling_values)
        return avg_scaled_prob
def calcDamProbs(creatureStats, action, modifier, threshold):
    def prob_damage_at_least_normal(thresh, diceNum, sides, flatMod=0):
        #Approximates P(sum of dice + flatMod >= threshold)
        mu = diceNum * (sides + 1) / 2
        sigma = math.sqrt(diceNum * (sides ** 2 - 1) / 12)
        z = (thresh - flatMod - 0.5 - mu) / sigma
        return 1 - norm.cdf(z)
    def prob_damage_at_least(thresh, dice, sides, flat):
        if dice >= 6:
            return prob_damage_at_least_normal(thresh, dice, sides, flat)

        total = 0
        favorable = 0
        for roll in itertools.product(range(1, sides + 1), repeat=dice):
            dmg = sum(roll) + flat
            total += 1
            if dmg >= thresh:
                favorable += 1
        return favorable / total
    diceNum = action.getDiceNum()
    dieType = action.getSides()
    if threshold == "NORM":
        threshold = ((diceNum * dieType) + modifier) / 2
    if threshold == "MULT":
        threshold = (diceNum * dieType) + modifier

    normDamProb = 0
    critDamProb = 0
    if isinstance(creatureStats, Player) and not isinstance(action.getDamType(), list) and action.getDamType().lower() == "healing":
        if creatureStats.isActiveCondition("Downed") or creatureStats.isActiveCondition("Stabilized"):
            normDamProb = 1.0
        elif action.getMean() != 0:
            normDamProb = (action.getMean() / creatureStats.getHP())
        else:
            normDamProb = 0
        return normDamProb, 0
    if isinstance(action.getDamType(), list) and len(action.getDamType()) != 1 and len(action.getDamType()) != 0:
        if action.getDamType()[-1] == "AND":
            numDamTypes = len(action.getDamType()) - 1
            # Assume a clean divide when having multiple damage types. 10d6 should only have 2 damTypes, never 3.
            #MEANING, there should never be a prime number of diceNums with multiple damageTypes.
                #Looking at you, Ice Storm. You suck.
            diceNums = list(itertools.repeat(diceNum / numDamTypes, numDamTypes))
            diceNums = [int(diceNum) for diceNum in diceNums]
            for i in range(len(diceNums)):
                if creatureStats.isImmune(action.getDamType()[i]):
                    diceNums[i] = 0
                elif creatureStats.isVulnerable(action.getDamType()[i]):
                    if not creatureStats.isResistant(action.getDamType()[i]):
                        diceNums[i] *= 2
                elif creatureStats.isResistant(action.getDamType()[i]):
                    diceNums[i] = math.floor(diceNums[i] / 2)
            diceNum = sum(diceNums)
            normDamProb += prob_damage_at_least(threshold, diceNum, dieType, modifier)
            critDamProb += prob_damage_at_least(threshold, 2 * diceNum, dieType, modifier)
        else:  # OR
            if all(True if creatureStats.isImmune(damType) and damType != "OR" else False for damType in action.getDamType()):
                normDamProb = 0
                critDamProb = 0
            elif all(creatureStats.isResistant(damType) for damType in action.getDamType()):
                if not any(creatureStats.isVulnerable(damType) for damType in action.getDamType()):
                    normDamProb = prob_damage_at_least(threshold * 2, diceNum, dieType, modifier)
                    critDamProb = prob_damage_at_least(threshold * 2, 2 * diceNum, dieType, modifier)
                else:
                    normDamProb = prob_damage_at_least(threshold, diceNum, dieType, modifier)
                    critDamProb = prob_damage_at_least(threshold, 2 * diceNum, dieType, modifier)
            elif any(creatureStats.isVulnerable(damType) for damType in action.getDamType()):
                diceNum *= 2
                normDamProb = prob_damage_at_least(threshold, diceNum, dieType, modifier)
                critDamProb = prob_damage_at_least(threshold, 2 * diceNum, dieType, modifier)
            else:
                normDamProb = prob_damage_at_least(threshold, diceNum, dieType, modifier)
                critDamProb = prob_damage_at_least(threshold, 2 * diceNum, dieType, modifier)
    else:
        if creatureStats.isImmune(action.getDamType()):
            normDamProb = 0
            critDamProb = 0
        elif creatureStats.isVulnerable(action.getDamType()):
            if not creatureStats.isResistant(action.getDamType()):
                diceNum *= 2
                normDamProb = prob_damage_at_least(threshold, diceNum, dieType, modifier)
                critDamProb = prob_damage_at_least(threshold, 2 * diceNum, dieType, modifier)
            else:
                normDamProb = prob_damage_at_least(threshold, diceNum, dieType, modifier)
                critDamProb = prob_damage_at_least(threshold, 2 * diceNum, dieType, modifier)
        elif creatureStats.isResistant(action.getDamType()):
            # Halving final damage means threshold effectively doubles
            normDamProb = prob_damage_at_least(threshold * 2, diceNum, dieType, modifier)
            critDamProb = prob_damage_at_least(threshold * 2, 2 * diceNum, dieType, modifier)
        else:
            normDamProb = prob_damage_at_least(threshold, diceNum, dieType, modifier)
            critDamProb = prob_damage_at_least(threshold, 2 * diceNum, dieType, modifier)
    return normDamProb, critDamProb
def translateLingEffect(action, lingEffect, spellMod):
    if (isinstance(lingEffect, dict) and "repeat" in lingEffect) or lingEffect == "repeat":
        action = copy.deepcopy(action)
        action.setLingEffects({})
        action.setLingSaves({})
        return action

    # 2. Parse damage info
    try:
        lingDieNum, lingDieType = map(int, lingEffect["rolls"]["damage"].split("d"))
    except Exception:
        lingDieNum, lingDieType = 0, 0

    # 3. Handle damage modifier
    damMod = 0
    if "damageMod" in lingEffect["rolls"]:
        dm = lingEffect["rolls"]["damageMod"]
        if dm == "spellMod":
            damMod = spellMod
        elif isinstance(dm, str) and dm.isdigit():
            damMod = int(dm)

    # 4. Build the lingering Spell object
    if "number" in lingEffect:
        numTarget = int(lingEffect["number"])
    else:
        numTarget = action.getNumTarget()

    if "specialNotes" in lingEffect:
        specialNotes = lingEffect["specialNotes"]
    else:
        specialNotes = []

    if "conditions" in lingEffect:
        conditions = lingEffect["conditions"]
    else:
        conditions = []

    if "statusEffect" in lingEffect:
        statusEffect = lingEffect["statusEffect"]
    else:
        statusEffect = []

    lingDamType = lingEffect["damType"]
    if isinstance(lingDamType, list) and len(lingDamType) == 1:
        lingDamType = lingDamType[0]
    lingSpell = Spell(
        action.getName(), action.getLvl(), action.getSelfTarget(), numTarget,
        lingEffect["rolls"]["rollType"], lingEffect["rolls"]["saveType"], lingEffect["rolls"]["halfSave"],
        damMod, lingDieNum, lingDieType, lingDamType,
        conditions, statusEffect, {}, {}, {}, "", "",specialNotes
    )
    return lingSpell
def calcLingeringEffectProbability(player, action, lingEffect, initiative, successProb):
    # 1. Repeat check
    if "repeat" in lingEffect and lingEffect["repeat"] == True:
        return successProb

    lingSpell = translateLingEffect(action, lingEffect, player.getSpellMod())

    # 5. Route to correct probability function
    roll_type = lingSpell.getRollType().lower()
    if roll_type == "tohit":
        lingEffectProb = 0
        numMonsters = 0
        for creature in initiative:
            if (creature["turnType"] == "Monster"
                    and not creature["Statblock"].isActiveStatusEffect("SwitchSides")
                    and not creature["Statblock"].isActiveCondition("Dead")
                    and not creature["Statblock"].isActiveCondition("Out of Combat")):
                lingEffectProb += calcIndividualToHitProbability(player, lingSpell, creature)
                numMonsters += 1
        lingEffectProb = lingEffectProb / numMonsters if numMonsters > 0 else 0
    elif roll_type == "save":
        lingEffectProb = calcTotalSaveProbability(player, lingSpell, initiative)["probSuccess"]
    elif roll_type == "autohit":
        lingEffectProb = calcTotalAutoHitProbability(player, lingSpell, initiative)["probSuccess"]
    else:
        lingEffectProb = 0
    #No spells have onHit lingering effects
    return lingEffectProb
def calcLingeringSavesProbability(player, spell, initiative):
    saveProb = 0
    numMonsters = 0
    for creature in initiative:
        if (creature["turnType"] == "Monster"
                and not creature["Statblock"].isActiveStatusEffect("SwitchSides")
                and not creature["Statblock"].isActiveCondition("Dead")
                and not creature["Statblock"].isActiveCondition("Out of Combat")):
            saveProb += ((21 - player.getDC()) + creature["Statblock"].getSaveProf(spell.getLingSaves()["saveType"])) / 20
            numMonsters += 1
    return saveProb / numMonsters if numMonsters != 0 else 0
def getMultiTargetWeights(player, action, initiative):
    modifier = player.getSpellMod()
    weights = []
    if isinstance(player, Player):
        isPlayerTurn=True
    else:
        isPlayerTurn=False
    for creature in initiative:
        if isValidTarget(action, creature, isPlayerTurn):
            eDam = calcIndividualExpectedDamage(player, action, creature)
            hp = creature["Statblock"].getHP()
            if action.getRollType().lower() == "tohit":
                probNormalHit, critChance = defProbHit(player, creature["Statblock"], modifier)
                killDamProb, killCritDamProb = calcDamProbs(creature["Statblock"], action, modifier, hp)
                killProb = probNormalHit * killDamProb + critChance * killCritDamProb
            elif action.getRollType().lower() == "save":
                specImm, specRes, specVuln = saveSpecialNotesCheck(action, creature["Statblock"])
                probFail = 1 - defSave(action, player.getDC(), creature["Statblock"])
                modifier = action.getDamMod()
                killDamProb = calcDamProbs(creature["Statblock"], action, modifier, hp)[0]
                if action.getHalfSave():
                    probSave = 1 - probFail
                    saveKillDamProb = calcDamProbs(creature["Statblock"], action, modifier, hp * 2)[0]
                    killProb = probSave * saveKillDamProb + probFail * killDamProb
                else:
                    killProb = probFail * killDamProb
                resetSaveSpecialNotesCheck(specImm, specRes, specVuln, creature["Statblock"])
            elif action.getRollType().lower() == "autohit":
                killProb = calcDamProbs(creature["Statblock"], action, modifier, hp)[0]
            else:
                killProb = 0

            weights.append({"Weight": (eDam / max(hp, 1)) + 1.5 * killProb, "Creature": copy.deepcopy(creature)})
    weights = sorted(weights, key=lambda x: x['Weight'], reverse=True)
    weights = [weight["Creature"] for weight in weights]
    weights = weights[0:min(action.getNumTarget(), len(weights))]
    return weights
def isValidTarget(action, creature, isPlayerTurn=True):
    if isPlayerTurn:
        if isinstance(action, Weapon) or (isinstance(action, Spell) and action.getDamType() != "healing"):
            if (creature["turnType"] == "Monster"
                and not creature["Statblock"].isActiveStatusEffect("SwitchSides")
                and not creature["Statblock"].isActiveCondition("Dead")
                and not creature["Statblock"].isActiveCondition("Out of Combat")):
                validTarget = True
            else:
                validTarget = False
        elif action.getDamType() == "healing":
            if (creature["turnType"] == "Player"
                     or (creature["turnType"] == "Monster" and creature["Statblock"].isActiveStatusEffect("SwitchSides"))
                     and not creature["Statblock"].isActiveCondition("Dead")
                     and not creature["Statblock"].isActiveCondition("Out of Combat")):
                validTarget = True
            else:
                validTarget = False
        else:
            validTarget = False
        return validTarget
    else:
        if isinstance(action, MonAction) or isinstance(action, Spell) and action.getDamType() != "healing":
            if (creature["turnType"] == "Player"
                and not creature["Statblock"].isActiveStatusEffect("SwitchSides")
                and not creature["Statblock"].isActiveCondition("Dead")
                and not creature["Statblock"].isActiveCondition("Out of Combat")):
                validTarget = True
            else:
                validTarget = False
        elif action.getDamType() == "healing":
            if (creature["turnType"] == "Monster"
                     or (creature["turnType"] == "Player" and creature["Statblock"].isActiveStatusEffect("SwitchSides"))
                     and not creature["Statblock"].isActiveCondition("Dead")
                     and not creature["Statblock"].isActiveCondition("Out of Combat")):
                validTarget = True
            else:
                validTarget = False
        else:
            validTarget = False
        return validTarget
def ensureList(x):
    return x if isinstance(x, list) else [x]
def _cr_to_float(cr_str: str) -> float:
    """'1/4' -> 0.25, '2' -> 2.0"""
    s = str(cr_str).strip()
    if "/" in s:
        num, den = s.split("/")
        return float(num) / float(den)
    return float(s)

#EXPECTED DAMAGE METHODS
def calcIndividualExpectedDamage(player, action, creature):
    #Only for weapon attacks and single-target spells.
    if isinstance(player, Player):
        isPlayerTurn=True
    else:
        isPlayerTurn=False
    if isValidTarget(action, creature, isPlayerTurn):
        creatureStats = creature["Statblock"]
        probNormalHit = -1
        critChance = -1
        modifier = 0
        damModifier = 0
        if isinstance(action, Weapon):
            modifier = player.getMod(action.getStatType())
            probNormalHit, critChance = defProbHit(player, creatureStats, modifier)
            damModifier = modifier
        else:
            if action.getRollType().lower() == "tohit":
                if isinstance(player, Player):
                    modifier = player.getSpellMod()
                else:
                    modifier = action.getAttackBonus()
                probNormalHit, critChance = defProbHit(player, creatureStats, modifier)
                damModifier = action.getDamMod()
            elif action.getRollType().lower() == "save":
                probNormalHit = 1 - defSave(action, player.getDC(), creatureStats)
                critChance = 0
                damModifier = action.getDamMod()
            elif action.getRollType().lower() == "autohit" or action.getRollType().lower() == "onhit":
                probNormalHit = 1.0
                critChance = 0
                damModifier = action.getDamMod()

        diceNum = copy.deepcopy(action.getDiceNum())
        dieType = copy.deepcopy(action.getSides())

        expectedNormalDamage = ((diceNum * (dieType + 1)) / 2) + damModifier
        expectedCritDamage = ((expectedNormalDamage - modifier) * 2) + damModifier

        if isinstance(action.getDamType(), list) and len(action.getDamType()) != 1 and len(action.getDamType()) != 0:
            if action.getDamType()[-1] == "AND":
                numDamTypes = len(action.getDamType()) - 1
                # Assume a clean divide when having multiple damage types. 10d6 should only have 2 damTypes, never 3.
                diceNums = list(itertools.repeat(diceNum / numDamTypes, numDamTypes))
                diceNums = [int(diceNum) for diceNum in diceNums]
                for i in range(len(diceNums)):
                    if creatureStats.isImmune(action.getDamType()[i]):
                        diceNums[i] = 0
                    elif creatureStats.isVulnerable(action.getDamType()[i]):
                        if not creatureStats.isResistant(action.getDamType()[i]):
                            diceNums[i] *= 2
                    elif creatureStats.isResistant(action.getDamType()[i]):
                        diceNums[i] = math.floor(diceNums[i] / 2)
                diceNum = sum(diceNums)
                expectedNormalDamage = ((diceNum * (dieType + 1)) / 2) + modifier
                expectedCritDamage = ((expectedNormalDamage - modifier) * 2) + modifier
            else:  # OR
                if all(creatureStats.isImmune(damType) for damType in action.getDamType()):
                    expectedNormalDamage = 0
                    expectedCritDamage = 0
                elif all(creatureStats.isResistant(damType) for damType in action.getDamType()):
                    if not any(creatureStats.isVulnerable(damType) for damType in action.getDamType()):
                        expectedNormalDamage /= 2
                        expectedCritDamage /= 2
                elif any(creatureStats.isVulnerable(damType) for damType in action.getDamType()):
                    expectedNormalDamage *= 2
                    expectedCritDamage *= 2
        else:
            damType = action.getDamType()
            if creatureStats.isImmune(damType):
                expectedNormalDamage *= 0
                expectedCritDamage *= 0
            elif creatureStats.isVulnerable(damType):
                if not creatureStats.isResistant(damType):
                    expectedNormalDamage *= 2
                    expectedCritDamage *= 2
            elif creatureStats.isResistant(damType):
                expectedNormalDamage /= 2
                expectedCritDamage /= 2

        expectedDamage = probNormalHit * expectedNormalDamage + critChance * expectedCritDamage
        return round(expectedDamage, 3)
    else:
        return None
def calcTotalExpectedDamage(player, action, initiative):
    eDamages = 0
    numMonsters = 0
    if isinstance(player, Player):
        isPlayerTurn=True
    else:
        isPlayerTurn=False
    if isinstance(action, Spell) and action.getRollType().lower() == "onhit":
        weaponMean = 0
    for creature in initiative:
        if isValidTarget(action, creature, isPlayerTurn):
            if isinstance(action, Weapon):
                eDamages += calcIndividualExpectedDamage(player, action, creature)
            elif action.getRollType().lower() == "tohit":
                if action.getNumTarget() == 1:
                    eDamages += calcIndividualExpectedDamage(player, action, creature)
                else: #Multitarget
                    weights = getMultiTargetWeights(player, action, initiative)
                    if weights:
                        tempTargetStore = action.getNumTarget()
                        action.setNumTarget(1)
                        for weight in weights:
                            eDamages += calcIndividualExpectedDamage(player, action, weight)
                            numMonsters += 1
                        action.setNumTarget(tempTargetStore)
                    break
            elif action.getRollType().lower() == "save":
                if action.getNumTarget() == 1:
                    eDamages += calcIndividualExpectedDamage(player, action, creature)
                elif action.getNumTarget() > 1:
                    weights = getMultiTargetWeights(player, action, initiative)
                    if weights:
                        tempTargetStore = action.getNumTarget()
                        action.setNumTarget(1)
                        for weight in weights:
                            eDamages += calcIndividualExpectedDamage(player, action, weight)
                            numMonsters += 1
                        action.setNumTarget(tempTargetStore)
                    break
                else:
                    targets = [
                        creature for creature in initiative
                        if isValidTarget(action, creature, isPlayerTurn)
                    ]

                    # Cache individual probabilities once
                    e_cache = {creature["name"]: calcIndividualExpectedDamage(player, action, creature)
                               for creature in targets}

                    eDam = avgOverAOETargets(e_cache, [creature["Statblock"] for creature in targets])
                    return round(eDam, 2)
            elif action.getRollType().lower() == "autohit":
                eDamages += calcIndividualExpectedDamage(player, action, creature)
            elif action.getRollType().lower() == "onhit":
                #Get the expected damage for the highest probToHit weapon
                if player.getWeaponLength() == 0:
                    return 0
                weaponDamages = []
                for i in range(player.getWeaponLength()):
                    weaponDamages.append(calcIndividualExpectedDamage(player, player.getWeapon(i), creature))
                weaponDam = max(weaponDamages)
                weaponMean = player.getWeapon(weaponDamages.index(weaponDam)).getMean()
                eDamages += calcIndividualExpectedDamage(player, action, creature) + weaponDam
            numMonsters += 1
    if isinstance(action, Spell) and action.getRollType().lower() == "onhit":
        spells = [player.getSpell(i) for i in range(player.getSpellLength())]
        spell = player.getSpell(spells.index(action))
        spell.setMean(spell.getMean() + weaponMean)
    if numMonsters != 0:
        return round(eDamages / numMonsters, 2)
    else:
        return 0

#PROB OF SUCCESS METHODS
def defProbHit(player, creatureStats, mod):
    toHitMod = player.getProfBonus() + mod
    critChance = 0.05  # 1 in 20
    toHitMod, critChance = influenceToHit(player, creatureStats, toHitMod, critChance)
    probHit = min(max((21 - creatureStats.getAC() + toHitMod) / 20, 0.05), 0.95)

    autoCritConditions = ["Paralyzed", "Unconscious"]
    # checking for players being able to autocrit
    autoCrit = False
    if creatureStats.isActiveStatusEffect("autocrit"):
        if "attack rolls against" in creatureStats.getActiveStatusEffect("autocrit")["effect"]["attribute"]:
            probHit = 1.0
            autoCrit = True
    elif any(condition in creatureStats.getActiveConditions() for condition in autoCritConditions):
        probHit = 1.0
        autoCrit = True
    if not autoCrit:
        probNormalHit = probHit - critChance
    else:
        probNormalHit = 0
        critChance = 1.0
    return probNormalHit, critChance
def influenceToHit(player, creatureStats, toHitMod, critChance):
    activeStatEffects = []  # Ensure no duplicate stat effects are applied

    # Check for player status effects
    if player.isActiveStatusEffect("Advantage"):
        advEffect = player.getActiveStatusEffect("Advantage")
        if "attack rolls for" in advEffect["effect"]["attribute"]:
            critChance = .0975  # See documentation
            toHitMod += int(player.getActiveStatusEffect("Advantage")["effect"]["rolls"])
            activeStatEffects.append("Advantage")
    elif player.isActiveStatusEffect("Disadvantage"):
        disadvEffect = player.getActiveStatusEffect("Disadvantage")
        if "attack rolls for" in disadvEffect["effect"]["attribute"]:
            if "Advantage" not in activeStatEffects:
                critChance = .025  # See documentation
            else:
                critChance = .05
            toHitMod += int(player.getActiveStatusEffect("Disadvantage")["effect"]["rolls"])
            activeStatEffects.append("Disadvantage")
    elif player.isActiveStatusEffect("Buff"):
        buffEffect = player.getActiveStatusEffect("Buff")
        if "attack rolls for" in buffEffect["effect"]["attribute"]:
            buffDieNum, buffDieType = buffEffect["effect"]["rolls"].split("d")
            buffDieNum, buffDieType = int(buffDieNum), int(buffDieType)
            toHitMod += (sum([int(i) for i in range(1, buffDieType + 1)]) / buffDieType)
            activeStatEffects.append("Buff")
        elif "AC" in buffEffect["effect"]["attribute"]:
            buffNum = buffEffect["effect"]["rolls"]
            buffNum = int(buffNum)
            toHitMod -= buffNum
            activeStatEffects.append("Buff")
    elif player.isActiveStatusEffect("Debuff"):
        buffEffect = player.getActiveStatusEffect("Debuff")
        if "attack rolls for" in buffEffect["effect"]["attribute"]:
            buffDieNum, buffDieType = buffEffect["effect"]["rolls"].split("d")
            buffDieNum, buffDieType = int(buffDieNum), int(buffDieType)
            toHitMod -= (sum([int(i) for i in range(1, buffDieType + 1)]) / buffDieNum)
            activeStatEffects.append("Debuff")

    # Check for player conditions (Hardcoded)
    disadvForConditions = ["Blinded", "Frightened", "Poisoned", "Prone", "Restrained"]
    advForConditions = ["Invisible", "GreaterInvisible"]
    if any(condition in player.getActiveConditions() for condition in
           disadvForConditions) and "Disadvantage" not in activeStatEffects:
        toHitMod += -4
        if "Advantage" not in activeStatEffects:
            critChance = .025  # See documentation
        else:
            critChance = .05
        activeStatEffects.append("Disadvantage")
    elif any(condition in player.getActiveConditions() for condition in
             advForConditions) and "Advantage" not in activeStatEffects:
        toHitMod += 4
        if "Disadvantage" not in activeStatEffects:
            critChance = .0975  # See documentation
        else:
            critChance = .05
        activeStatEffects.append("Advantage")

    # Checking for monster status effects
    if creatureStats.isActiveStatusEffect("Advantage") and "Advantage" not in activeStatEffects:
        advEffect = creatureStats.getActiveStatusEffect("Advantage")
        if "attack rolls against" in advEffect["effect"]["attribute"]:
            if "Disadvantage" not in activeStatEffects:
                critChance = .0975  # See documentation
            else:
                critChance = .05
            toHitMod += int(creatureStats.getActiveStatusEffect("Advantage")["effect"]["rolls"])
            activeStatEffects.append("Advantage")
    elif creatureStats.isActiveStatusEffect("Disadvantage") and "Disadvantage" not in activeStatEffects:
        if "Advantage" not in activeStatEffects:
            critChance = .025  # See documentation
        else:
            critChance = .05
        disadvEffect = creatureStats.getActiveStatusEffect("Disadvantage")
        if "attack rolls against" in disadvEffect["effect"]["attribute"]:
            toHitMod += int(creatureStats.getActiveStatusEffect("Disadvantage")["effect"]["rolls"])
            activeStatEffects.append("Disadvantage")
    elif creatureStats.isActiveStatusEffect("Buff"):
        buffEffect = creatureStats.getActiveStatusEffect("Buff")
        if "attack rolls against" in buffEffect["effect"]["attribute"]:
            buffDieNum, buffDieType = buffEffect["effect"]["rolls"].split("d")
            buffDieNum, buffDieType = int(buffDieNum), int(buffDieType)
            toHitMod += (sum([int(i) for i in range(1, buffDieType + 1)]) / buffDieType)
    elif creatureStats.isActiveStatusEffect("Debuff"):
        buffEffect = creatureStats.getActiveStatusEffect("Debuff")
        if "attack rolls against" in buffEffect["effect"]["attribute"]:
            buffDieNum, buffDieType = buffEffect["effect"]["rolls"].split("d")
            buffDieNum, buffDieType = int(buffDieNum), int(buffDieType)
            toHitMod -= (sum([int(i) for i in range(1, buffDieType + 1)]) / buffDieNum)

    # Checking for monster conditions
    advAgainstConditions = ["Blinded", "Prone", "Restrained", "Stunned"]
    if any(condition in creatureStats.getActiveConditions() for condition in
           advForConditions) and "Disadvantage" not in activeStatEffects:
        toHitMod += -4
        if "Advantage" not in activeStatEffects:
            critChance = .025  # See documentation
        else:
            critChance = .05
        activeStatEffects.append("Disadvantage")
    elif any(condition in creatureStats.getActiveConditions() for condition in
             advAgainstConditions) and "Advantage" not in activeStatEffects:
        toHitMod += 4
        if "Disadvantage" not in activeStatEffects:
            critChance = .0975  # See documentation
        else:
            critChance = .05
        activeStatEffects.append("Advantage")

    return toHitMod, critChance
def calcIndividualToHitProbability(player, action, creature):
    creatureStats = creature["Statblock"]
    modifier = 0
    if isinstance(player, Player):
        isPlayerTurn=True
    else:
        isPlayerTurn=False
    if isValidTarget(action, creature, isPlayerTurn):
        if isinstance(action, Weapon):
            modifier = player.getMod(action.getStatType())
        elif isinstance(action, MonAction):
            modifier = action.getAttackBonus()
        elif isinstance(action, Spell):
            modifier = player.getSpellMod()
        else:
            return None
    else:
        return None

    probNormalHit, critChance = defProbHit(player, creatureStats, modifier)
    if isinstance(action, MonAction):
        modifier = action.getDamMod()
    normDamProb, critDamProb = calcDamProbs(creatureStats, action, modifier, "NORM")

    #NOTE: No specialNotes relevant to probability of success in all toHit spells
    successProb = probNormalHit * normDamProb + critChance * critDamProb
    return successProb
def calcTotalToHitProbability(player, action, initiative):
    #Only 1 or >1 targets for spells; also covers weapons.
    if isinstance(action, Weapon) or action.getNumTarget() == 1:
        if isinstance(player, Player):
            isPlayerTurn = True
        else:
            isPlayerTurn = False

        successProbs = 0
        numMonsters = 0

        lingEffectProb = 0
        checkLingEffects = True if isinstance(action, Spell) and action.getLingEffects() else False
        # None of the toHit spells have lingering saves to account for.

        extraEffectProb = 0
        checkExtraEffects = True if isinstance(action, Spell) and action.getExtraEffect() else False

        lingSavesProb = 0
        checkLingSaves = True if isinstance(action, Spell) and action.getLingSaves() else False

        for creature in initiative:
            if isValidTarget(action, creature, isPlayerTurn):
                successProb = calcIndividualToHitProbability(player, action, creature)
                successProbs += successProb
                numMonsters += 1
        if numMonsters != 0:
            probSuccess = round(successProbs / numMonsters, 3)
        else:
            probSuccess = 0
        if checkLingEffects:
            lingEffectProb = calcLingeringEffectProbability(player, action, action.getLingEffects(), initiative, probSuccess)
        if checkExtraEffects:
            # In terms of probability of success, lingEffects and extraEffects are the same.
            extraEffectProb = calcLingeringEffectProbability(player, action, action.getExtraEffect(), initiative, probSuccess)
        if checkLingSaves:
            lingSavesProb = calcLingeringSavesProbability(player, action, initiative)
        if numMonsters != 0:
            probSuccess = round(probSuccess, 2)
            lingEffectProb = round(lingEffectProb, 2)
            extraEffectProb = round(extraEffectProb, 2)
            lingSavesProb = round(lingSavesProb, 2)
            return {
                "probSuccess": probSuccess,
                "probLingEffect" : lingEffectProb,
                "probExtraEffect": extraEffectProb,
                "probLingSaves": lingSavesProb
            }
        return 0
    elif action.getNumTarget() > 1:
        weights = getMultiTargetWeights(player, action, initiative)
        if weights:
            tempTargetStore = action.getNumTarget()
            action.setNumTarget(1)
            totalProbSuccess = calcTotalToHitProbability(player, action, weights)
            action.setNumTarget(tempTargetStore)
            return totalProbSuccess
        else:
            return 0
    else:
        raise ValueError("Bad NumTargets!")
def defSave(spell, dc, creatureStats):
    saveMod = creatureStats.getSaveProf(spell.getSaveType())
    activeStatEffects = []  # Ensure no duplicate stat effects are applied

    # Check for creatureStats status effects
    if creatureStats.isActiveStatusEffect("Advantage"):
        advEffect = creatureStats.getActiveStatusEffect("Advantage")
        if f"{spell.getSaveType()} save" in advEffect["effect"]["attribute"] or "ALL save" in advEffect["effect"]["attribute"]:
            saveMod += int(creatureStats.getActiveStatusEffect("Advantage")["effect"]["rolls"])
            activeStatEffects.append("Advantage")
    elif creatureStats.isActiveStatusEffect("Disadvantage"):
        disadvEffect = creatureStats.getActiveStatusEffect("Disadvantage")
        if (f"{spell.getSaveType()} save" in disadvEffect["effect"]["attribute"]
                or "ALL save" in disadvEffect["effect"]["attribute"]):
            saveMod -= int(creatureStats.getActiveStatusEffect("Disadvantage")["effect"]["rolls"])
            activeStatEffects.append("Disadvantage")
    elif creatureStats.isActiveStatusEffect("Buff"):
        buffEffect = creatureStats.getActiveStatusEffect("Buff")
        if f"{spell.getSaveType()} save" in buffEffect["effect"]["attribute"] or "ALL save" in buffEffect["effect"]["attribute"]:
            try:
                buffDieNum, buffDieType = buffEffect["effect"]["rolls"].split("d")
                buffDieNum, buffDieType = int(buffDieNum), int(buffDieType)
                saveMod += (sum([i for i in range(1, buffDieType + 1)]) / buffDieType)
                activeStatEffects.append("Buff")
            except:
                print("ERROR: BAD DATA FOR BUFF ATTRIBUTE")
    elif creatureStats.isActiveStatusEffect("Debuff"):
        buffEffect = creatureStats.getActiveStatusEffect("Debuff")
        if (f"{spell.getSaveType()} save" in buffEffect["effect"]["attribute"]
                or "ALL save" in buffEffect["effect"]["attribute"]):
            try:
                buffDieNum, buffDieType = buffEffect["effect"]["rolls"].split("d")
                buffDieNum, buffDieType = int(buffDieNum), int(buffDieType)
                saveMod -= (sum([i for i in range(1, buffDieType + 1)]) / buffDieType)
                activeStatEffects.append("Debuff")
            except:
                print("ERROR: BAD DATA FOR BEBUFF ATTRIBUTE")
    if creatureStats.isActiveCondition("Restrained"):
        if "Disadvantage" not in activeStatEffects and not creatureStats.isActiveStatusEffect("Disadvantage"):
            saveMod -= 4
    if creatureStats.isActiveStatusEffect("autofail") or creatureStats.isActiveCondition("Paralyzed") \
        or creatureStats.isActiveCondition("Petrified") or creatureStats.isActiveCondition("Stunned") \
        or creatureStats.isActiveCondition("Unconscious"):
        saveProb = 0
    else:
        saveProb = ((21 - dc) + saveMod) / 20
    return saveProb
def saveSpecialNotesCheck(action, creature):
    immunities, resistances, vulnerabilities = [], [], []
    if not isinstance(action.getDamType(), list) and action.getDamType().lower() == "healing":
        return immunities, resistances, vulnerabilities
    specialNotes = action.getSpecialNotes()
    if specialNotes:
        for note in specialNotes:
            if "only" in note.lower() and creature.getCreatureType() not in specialNotes():
                return None
            elif "immune" in note.lower() and creature.getCreatureType() in note:
                immunities = copy.deepcopy(creature.getDamImmunities())
                damType = action.getDamType()
                if isinstance(damType, list):
                    if damType[-1] == "AND" or damType[-1] == "OR":
                        for i in range(len(damType) - 1):
                            if not creature.isImmune(damType[i]):
                                creature.addDamImmunity(damType[i])
                    else:
                        raise ValueError("Listed damTypes MUST end in AND or OR!")
                else:
                    if not creature.isImmune(damType):
                        creature.addDamImmunity(damType)
                break
            elif "resist" in note.lower() and creature.getCreatureType() in note:
                resistances = copy.deepcopy(creature.getDamResistances())
                damType = action.getDamType()
                if isinstance(damType, list):
                    if damType[-1] == "AND" or damType[-1] == "OR":
                        for i in range(len(damType) - 1):
                            if not creature.isResistant(damType[i]):
                                creature.addDamResist(damType[i])
                    else:
                        raise ValueError("Listed damTypes MUST end in AND or OR!")
                else:
                    if not creature.isResistant(damType):
                        creature.addDamResist(damType)
                break
            elif "vulnerable" in note.lower() and creature.getCreatureType() in note:
                vulnerabilities = copy.deepcopy(creature.getDamVulnerabilities())
                damType = action.getDamType()
                if isinstance(damType, list):
                    if damType[-1] == "AND" or damType[-1] == "OR":
                        for i in range(len(damType) - 1):
                            if not creature.isVulnerable(damType[i]):
                                creature.addDamVuln(damType[i])
                    else:
                        raise ValueError("Listed damTypes MUST end in AND or OR!")
                else:
                    if not creature.isVulnerable(damType):
                        creature.addDamVulnerability(damType)
                break
    return immunities, resistances, vulnerabilities
def resetSaveSpecialNotesCheck(specImm, specRes, specVuln, creature):
    if specImm:
        for imm in creature.getDamImmunities():
            if imm not in specImm:
                creature.removeDamImmunity(imm)
    elif specRes:
        for res in creature.getDamResistances():
            if res not in specRes:
                creature.removeDamResist(res)
    elif specVuln:
        for vuln in creature.getDamVulnerabilities():
            if vuln not in specVuln:
                creature.removeDamVulnerability(vuln)
def avgOverAOETargets(cache, targets):
    numTargets = len(targets)
    if numTargets == 0:
        return 0

    max_trials = 1000
    numTrials = 0
    while 2 ** numTrials < max_trials: numTrials += 1
    if numTargets <= (numTrials - 1):
        subset_results = []
        # Enumerate all subsets (skip empty)
        from itertools import combinations
        for k in range(1, numTargets + 1):
            for combo in combinations(targets, k):
                avg_p = sum(cache[creature.getName()] for creature in combo) / k
                subset_results.append(avg_p)
        result = sum(subset_results) / len(subset_results)
    else:
        # Monte Carlo random subsets
        import random
        mean = 0.0
        meanSumSquares = 0.0
        variance = 999.0
        for t in range(1, max_trials + 1):
            subset_size = random.randint(1, numTargets)
            combo = random.sample(targets, subset_size)
            avg_p = sum(cache[creature.getName()] for creature in combo) / subset_size

            # incremental variance update
            delta = avg_p - mean
            mean += delta / t
            meanSumSquares += delta * (avg_p - mean)
            if t > 1:
                variance = meanSumSquares / (t - 1)
            if t >= 100 and variance < 0.0005:
                break
        result = mean
    return result
def calcIndividualSaveProbability(action, dc, creature):
    try:
        specImm, specRes, specVuln = saveSpecialNotesCheck(action, creature)
    except TypeError:
        return 0
    #Defines chance the creature fails the save.
    probFail = 1 - defSave(action, dc, creature)
    modifier = action.getDamMod()
    failDamProb = calcDamProbs(creature, action, modifier, "NORM")[0]
    if action.getHalfSave():
        probSave = 1 - probFail
        saveDamProb = calcDamProbs(creature, action, modifier, "MULT")[0]
        probSuccess = (probSave * saveDamProb) + (probFail * failDamProb)
    elif action.getMean() != 0: #Not all save spells deal damage
        probSuccess = probFail * failDamProb
    else:
        probSuccess = probFail

    resetSaveSpecialNotesCheck(specImm, specRes, specVuln, creature)
    return probSuccess
def calcTotalSaveProbability(player, action, initiative):
    #No 0 targets - everything else, fair game.
    lingEffectProb = 0
    checkLingEffects = True if action.getLingEffects() else False
    extraEffectProb = 0
    checkExtraEffects = True if action.getExtraEffect() else False
    lingSavesProb = 0
    checkLingSaves = True if action.getLingSaves() else False
    if isinstance(player, Player):
        isPlayerTurn=True
    else:
        isPlayerTurn=False
    if action.getNumTarget() == 1:
        successProbs = 0
        numMonsters = 0
        for creature in initiative:
            if isValidTarget(action, creature, isPlayerTurn):
                if isinstance(player, Player) or isinstance(action, Spell):
                    successProb = calcIndividualSaveProbability(action, player.getDC(), creature["Statblock"])
                else:
                    successProb = calcIndividualSaveProbability(action, action.getDC(), creature["Statblock"])
                successProbs += successProb
                numMonsters += 1
        if numMonsters != 0:
            probSuccess = round(successProbs / numMonsters, 2)
        else:
            probSuccess = 0
    elif action.getNumTarget() > 1:
        weights = getMultiTargetWeights(player, action, initiative)
        if weights:
            tempTargetStore = action.getNumTarget()
            action.setNumTarget(1)
            totalProbSuccess = calcTotalToHitProbability(player, action, weights)
            action.setNumTarget(tempTargetStore)
            return totalProbSuccess
        else:
            return 0
    elif action.getNumTarget() == -1:
        # Filter eligible targets
        targets = [creature["Statblock"] for creature in initiative if isValidTarget(action, creature, isPlayerTurn)]

        # Cache individual probabilities once
        p_cache = {creature.getName(): calcIndividualSaveProbability(action, player.getDC(), creature)
                   for creature in targets}

        probSuccess = avgOverAOETargets(p_cache, targets)
    else:
        raise ValueError("BAD VALUE FOR NUMTARGET")
    if checkLingEffects:
        lingEffectProb = calcLingeringEffectProbability(player, action, action.getLingEffects(), initiative,                                               probSuccess)
    if checkExtraEffects:
        # In terms of probability of success, lingEffects and extraEffects are the same.
        extraEffectProb = calcLingeringEffectProbability(player, action, action.getExtraEffect(), initiative,                                           probSuccess)
    if checkLingSaves:
        lingSavesProb = calcLingeringSavesProbability(player, action, initiative)
    probSuccess = round(probSuccess, 2)
    lingEffectProb = round(lingEffectProb, 2)
    extraEffectProb = round(extraEffectProb, 2)
    lingSavesProb = round(lingSavesProb, 2)
    return {
            "probSuccess": probSuccess,
            "probLingEffect": lingEffectProb,
            "probExtraEffect": extraEffectProb,
            "probLingSaves": lingSavesProb
    }
def calcOnHitProbability(action, weapons, player, initiative):
    #Only 1 target per spell.
    if isinstance(player, Player):
        isPlayerTurn=True
    else:
        isPlayerTurn=False
    probWeaponSuccess = []
    for weapon in weapons:
        probWeaponSuccess.append(calcTotalToHitProbability(player, weapon, initiative))
    if len(probWeaponSuccess) != 0:
        probWeaponSuccess = [prob["probSuccess"] for prob in probWeaponSuccess]
        probWeaponSuccess = max(probWeaponSuccess) #Using the weapon with the highest chance of success...
    else:
        return 0
    if action.getMean() != 0:
        probInitDam = 0
        numMonsters = 0
        for creature in initiative:
            if isValidTarget(action, creature, isPlayerTurn):
                probNormDam, probCritDam = calcDamProbs(creature["Statblock"], action, action.getDamMod(), "NORM")
                probInitDam += (probNormDam + probCritDam)
                numMonsters += 1
        if numMonsters != 0:
            probInitDam /= numMonsters
    else:
        probInitDam = 1.0 #No initial damage, so init damage would be useless. Pass

    checkLingEffects = True if isinstance(action, Spell) and action.getLingEffects() else False
    checkExtraEffects = True if isinstance(action, Spell) and action.getExtraEffect() else False
    checkLingSaves = True if isinstance(action, Spell) and action.getLingSaves() else False
    if checkLingEffects:
        #LingEffects here would only repeat the initial effect, not the weapon's success prob.
        lingEffectProb = calcLingeringEffectProbability(player, action, action.getLingEffects(), initiative,
                                                        probInitDam)
    else:
        lingEffectProb = 0
    if checkExtraEffects:
        extraEffectProb = calcLingeringEffectProbability(player, action, action.getExtraEffect(), initiative,
                                                         probInitDam)
    else:
        extraEffectProb = 0
    if checkLingSaves:
        lingSavesProb = calcLingeringSavesProbability(player, action, initiative)
    else:
        lingSavesProb = 0

    probSuccess = probWeaponSuccess * probInitDam
    return {
        "probSuccess" : round(probSuccess, 2),
        "probLingEffect": round(lingEffectProb, 2),
        "probExtraEffect": round(extraEffectProb, 2),
        "probLingSaves": round(lingSavesProb, 2)
    }
def calcIndividualAutoHitProbability(action, creature):
    try:
        specImm, specRes, specVuln = saveSpecialNotesCheck(action, creature)
    except TypeError:
        return 0
    if action.getMean() != 0:
        damProb = calcDamProbs(creature, action, action.getDamMod(), "NORM")[0]
    else:
        if specImm:
            damProb = 0
        else:
            damProb = 1.0
    resetSaveSpecialNotesCheck(specImm, specRes, specVuln, creature)

    return damProb
def calcTotalAutoHitProbability(player, action, initiative):
    #All possible targets.
    if isinstance(player, Player):
        isPlayerTurn=True
    else:
        isPlayerTurn=False
    targets = [creature for creature in initiative if isValidTarget(action, creature, isPlayerTurn)]
    # if not isinstance(action.getDamType(), list) and action.getDamType().lower() == "healing":
    #     targets = [
    #         creature for creature in initiative
    #         if (creature["turnType"] == "Player" or (creature["turnType"] == "Monster" and creature["Statblock"].isActiveStatusEffect("SwitchSides")))
    #            and not creature["Statblock"].isActiveCondition("Dead")
    #            and not creature["Statblock"].isActiveCondition("Out of Combat")
    #     ]
    # else:
    #     targets = [
    #         creature for creature in initiative
    #         if creature["turnType"] == "Monster"
    #            and not creature["Statblock"].isActiveStatusEffect("SwitchSides")
    #            and not creature["Statblock"].isActiveCondition("Dead")
    #            and not creature["Statblock"].isActiveCondition("Out of Combat")
    #     ]

    lingEffectProb = 0
    checkLingEffects = True if action.getLingEffects() else False
    extraEffectProb = 0
    checkExtraEffects = True if action.getExtraEffect() else False
    lingSavesProb = 0
    checkLingSaves = True if action.getLingSaves() else False

    if action.getNumTarget() == 1:
        probSuccess = 0
        if action.getSpecialNotes() and "HPCap" in action.getSpecialNotes():
            hpCap = 0
            specialNotes = action.getSpecialNotes()
            for note in specialNotes():
                if "HPCap" in note:
                    hpCap = int(note.split("HPCap")[1])
            for creature in targets:
                if creature["Statblock"].getHP() < hpCap:
                    probSuccess += 1
        else:
            for creature in targets:
                probSuccess += calcIndividualAutoHitProbability(action, creature["Statblock"])
        if len(targets) != 0:
            probSuccess /= len(targets)
        else:
            probSuccess = 0
    elif action.getNumTarget() > 1:
        weights = getMultiTargetWeights(player, action, targets)
        if weights:
            tempTargetStore = action.getNumTarget()
            action.setNumTarget(1)
            totalProbSuccess = calcTotalAutoHitProbability(player, action, weights)["probSuccess"]
            action.setNumTarget(tempTargetStore)
            return totalProbSuccess
        else:
            return 0
    else: #AOE targets
        # Cache individual probabilities once
        p_cache = {creature["Statblock"].getName(): calcIndividualAutoHitProbability(action, creature["Statblock"])
                   for creature in targets}

        probSuccess = avgOverAOETargets(p_cache, [target["Statblock"] for target in targets])
    # Not accounting for 0 target, since self spells are finished in playerTurn().

    if checkLingEffects:
        lingEffectProb = calcLingeringEffectProbability(player, action, action.getLingEffects(), initiative,                                               probSuccess)
    if checkExtraEffects:
        # In terms of probability of success, lingEffects and extraEffects are the same.
        extraEffectProb = calcLingeringEffectProbability(player, action, action.getExtraEffect(), initiative,                                           probSuccess)
    if checkLingSaves:
        lingSavesProb = calcLingeringSavesProbability(player, action, initiative)

    probSuccess = round(probSuccess, 2)
    lingEffectProb = round(lingEffectProb, 2)
    extraEffectProb = round(extraEffectProb, 2)
    lingSavesProb = round(lingSavesProb, 2)
    return {
            "probSuccess": probSuccess,
            "probLingEffect": lingEffectProb,
            "probExtraEffect": extraEffectProb,
            "probLingSaves": lingSavesProb
    }

#IMPACT METHODS
def computePerTargetConditionImpact(action, probSuccess, creature, useProbSuccess=True):
    positiveConditionSeverities = {
        "invisible": 2,
        "greaterinvisible": 3
    }
    negativeConditionSeverities = {
        "blinded": 3,
        "charmed": 1,
        "frightened": 2,
        "incapacitated": 3,
        "paralyzed": 4,
        "petrified": 4,
        "prone": 1,
        "restrained": 2,
        "stunned": 4,
        "unconscious": 4,
        "out of combat": 6,
        "dead": 8
    }

    # Check conditions creature currently has
    # Check conditions the action will apply
    # Average their severities and multiply by probSuccess
    severity = 0
    if not isinstance(action.getDamType(), list) and action.getDamType().lower() != "healing":
        if action.getConditions():
            for condition in action.getConditions():
                if condition.lower() not in [c["cond"].lower() for c in
                                             creature.getActiveConditions()] and not creature.isActiveConImmunity(
                        condition.lower()):
                    severity += negativeConditionSeverities.get(condition.lower(), 0)
                    severity -= positiveConditionSeverities.get(condition.lower(), 0)
    else:
        if action.getConditions():
            for condition in action.getConditions():
                if "downed" in [c["cond"].lower() for c in creature.getActiveConditions()] or "stabilized" in [
                    c["cond"].lower() for c in creature.getActiveConditions()]:
                    severity += 5
                    continue
                if condition.lower() not in [c["cond"].lower() for c in creature.getActiveConditions()]:
                    severity -= negativeConditionSeverities.get(condition.lower(), 0)
                    severity += positiveConditionSeverities.get(condition.lower(), 0)
    if useProbSuccess:
        if isinstance(probSuccess, list):
            return severity * probSuccess[0]
        return severity * probSuccess
    else:
        return severity
def computePerTargetSEImpact(action, probSuccess, creature, useProbSuccess=True):
    positiveEffectSeverities = {
        "advantage": 2,
        "resistance": 2,
        "vulnerability": 3,
        "immunity": 3,
        "buff": 1,
        "autocrit": 4,
    }

    negativeEffectSeverities = {
        "debuff": 1,
        "autofail": 4,
        "disadvantage": 2,
        "lingeffect": 1,
        "lingsave": 2,
        "summon": 4
    }
    # Check status effects creature currently has
    # Check status effects the action will apply
    # Average their severities and multiply by probSuccess
    severity = 0
    activeEffects = []
    if not isinstance(action.getDamType(),
                      list) and action.getDamType().lower() != "healing" and not action.getSelfTarget():
        if action.getStatusEffects():
            for effect in action.getStatusEffects():
                if (effect["name"].lower() not in activeEffects
                        and effect["name"].lower() not in [c["name"].lower() for c in
                                                           creature.getActiveStatusEffects()]):
                    severity += negativeEffectSeverities.get(effect["name"].lower(), 0)
                    severity -= positiveEffectSeverities.get(effect["name"].lower(), 0)
                    activeEffects.append(effect["name"].lower())
    else:
        if action.getStatusEffects():
            for effect in action.getStatusEffects():
                if (effect["name"].lower() not in activeEffects
                        and effect["name"].lower() not in [c["name"].lower() for c in
                                                           creature.getActiveStatusEffects()]):
                    # TODO: Change these checks to account for "negative" SEs that account for positive attributes
                    if effect["name"].lower() == "disadvantage" and "attack rolls against" in effect["effect"][
                        "attribute"]:
                        severity += negativeEffectSeverities.get(effect["name"].lower(), 0)
                    else:
                        severity -= negativeEffectSeverities.get(effect["name"].lower(), 0)
                        severity += positiveEffectSeverities.get(effect["name"].lower(), 0)
                    activeEffects.append(effect["name"].lower())
    if useProbSuccess:
        if isinstance(probSuccess, list):
            return severity * probSuccess[0]
        return severity * probSuccess
    else:
        return severity
def calcImpact(player, action, probSuccess, expectedDamage, initiative, leRecursion=False, layeredRecursion=False):
    def computePerTargetDamageImpact(action, creature):
        base = min(expectedDamage, creature.getMaxHP() - creature.getHP())
        frac_restored = base / max(creature.getMaxHP(), 1)

        if not isinstance(action.getDamType(), list) and action.getDamType().lower() == "healing":
            # Stronger scaling by how close to death they are
            missing_pct = (creature.getMaxHP() - creature.getHP()) / max(creature.getMaxHP(), 1)
            if creature.getHP() == creature.getMaxHP():
                missing_pct = 0
            urgency = 1 + (10 * missing_pct) ** 1.5  # 0 → 1, 50% → 1 + 2.5^4, 90% -> 1 + 4.25^4
            #Urgency dramatically scales with low HP

            damImpact = (frac_restored + urgency) * probSuccess
        else:
            try:
                damImpact = (min(int(expectedDamage) / creature.getMaxHP(), 1) * 10 ) * probSuccess
            except:
                if isinstance(probSuccess, list):
                    try:
                        damImpact = (min(int(expectedDamage) / creature.getMaxHP(), 1) * 10) * probSuccess[0]
                    except:
                        damImpact = 0
                else:
                    damImpact = 0
            killProb = calcDamProbs(creature, action, action.getDamMod(), creature.getHP())[0]
            if killProb > 0:
                damImpact += 3 * killProb

        return damImpact
    def computeCRWeight(creature):
        level = _cr_to_float(creature.getLevel())
        crW = math.log(1 + level)
        return crW
    def round_to_first_nonzero_decimal(n):
        #AI code
        if n == 0:
            return 0
        # Determine the number of decimal places for the first non-zero digit
        decimal_places = -math.floor(math.log10(abs(n)))

        # If the number is > 1 or < -1, adjust the rounding to 1 decimal place (or as desired)
        # The user request implies handling small decimals, so we ensure precision is positive.
        if decimal_places < 0:
            decimal_places = 1  # or 0, depending on desired behavior for n >= 1
        decimal_places = max(decimal_places, 2)
        return round(n, decimal_places)
    def avgOverAOETargetsForImpact(perTarget, numCreatures):
        """
        Compute the effective impact rating for an AOE spell,
        given a list of per-target impact values.

        perTarget   : list of floats (impact of hitting creature i)
        numCreatures: integer number of valid creatures in encounter

        Rules:
          - 1 creature       → just return that single impact
          - 2–10 creatures   → evaluate all 2^n - 1 subsets
          - >10 creatures    → Monte Carlo sampling
        """
        # --- CASE 1: trivial ---
        if numCreatures <= 1:
            return perTarget[0] if perTarget else 0

        # --- CASE 2: exact subset enumeration (2–10 creatures) ---
        if 2 <= numCreatures <= 10:
            import itertools
            total = 0
            count = 0
            # All non-empty subsets
            for r in range(1, numCreatures + 1):
                for subset in itertools.combinations(perTarget, r):
                    total += sum(subset)
                    count += 1
            return total / count if count > 0 else 0

        # --- CASE 3: Monte-Carlo sampling for 11+ creatures ---
        import random
        TRIALS = 1000
        impacts = []

        # adaptive variance control
        running_mean = 0
        running_sq = 0

        for i in range(1, TRIALS + 1):
            subset_impact = 0
            # random subset — each creature included with p = 0.5
            for val in perTarget:
                if random.random() < 0.5:
                    subset_impact += val
            impacts.append(subset_impact)

            # update mean/variance online
            delta = subset_impact - running_mean
            running_mean += delta / i
            running_sq += delta * (subset_impact - running_mean)

            if i >= 100:
                variance = running_sq / (i - 1)
                # If variance low enough, break early
                if variance < 0.01 * (running_mean ** 2 + 1e-9):
                    break

        return sum(impacts) / len(impacts)
    if isinstance(player, Player):
        isPlayerTurn=True
    else:
        isPlayerTurn=False
    targets = [creature for creature in initiative if isValidTarget(action, creature, isPlayerTurn)]
    perTarget = []
    checkExtraEffects = True if isinstance(action, Spell) and action.getExtraEffect() else False
    checkLingEffects = True if isinstance(action, Spell) and action.getLingEffects() else False
    checkLingSaves = True if isinstance(action, Spell) and action.getLingSaves() else False
    if isinstance(probSuccess, str):
        if checkExtraEffects:
            extraEffect = translateLingEffect(action, action.getExtraEffect(), player.getSpellMod())
            extraProb = probSuccess.split(' - ')
            if len(extraProb) == 1:
                extraProb = extraProb[0]
            else:
                for prob in extraProb:
                    if 'EE' in prob:
                        extraProb = prob.split('EE')[0]
                        extraProb = float(extraProb)
                        break
        if checkLingEffects:
            lingEffect = translateLingEffect(action, action.getLingEffects(), player.getSpellMod())
            lingEffProb = probSuccess.split(' - ')
            if len(lingEffProb) == 1 or 'LE' not in lingEffProb:
                lingEffProb = lingEffProb[0]
            else:
                for prob in lingEffProb:
                    if 'LE' in prob:
                        lingEffProb = prob.split('LE')[0]
                    if 'EE' in prob:
                        lingEffProb += f" - {prob}"
        if checkLingSaves:
            lingSProb = probSuccess.split(' - ')
            if len(lingSProb) == 1:
                lingSProb = lingSProb[0]
                lingSProb = float(lingSProb)
            else:
                for prob in lingSProb:
                    if 'LS' in prob:
                        lingSProb = prob.split('LS')[0]
                        lingSProb = float(lingSProb)
                        break
        probSuccess = probSuccess.split(' - ')[0] if ' - ' in probSuccess else probSuccess
        probSuccess = float(probSuccess)
        if probSuccess == 0 and not checkLingEffects and not checkExtraEffects:
            return 0
    else:
        checkExtraEffects = False
        checkLingEffects = False
        checkLingSaves = False
    if isinstance(expectedDamage, str):
        expectedDamage = expectedDamage.split(' - ')[1] if ' - ' in expectedDamage else expectedDamage
        expectedDamage = expectedDamage.split('W')[0] if 'W' in expectedDamage else expectedDamage
        expectedDamage = float(expectedDamage)
    if isinstance(action, Spell) and action.getNumTarget() == 0:
        targets = [initiative[idx] for idx in range(len(initiative)) if initiative[idx]["Statblock"].getName().lower() == player.getName().lower()]

    extraImpact = 0
    lingEffImpact = 0
    lingSImpact = 0
    if checkExtraEffects:
        if leRecursion:
            extraImpact = calcImpact(player, extraEffect, extraProb, extraEffect.getMean(), initiative, False, True)
        else:
            extraImpact = calcImpact(player, extraEffect, extraProb, extraEffect.getMean(), initiative)
    if checkLingEffects:
        lingEffImpact = calcImpact(player, lingEffect, lingEffProb, lingEffect.getMean(), initiative, True, False)
    if checkLingSaves:
        lingSImpact = lingSProb
    for creature in targets:
        creature = creature["Statblock"]
        turnCount = 0
        if isinstance(action, Spell):
            specialNotes = action.getSpecialNotes()
            specialCase = False
            if specialNotes:
                for note in specialNotes:
                    if 'hpcap' in note.lower():
                        cap = note.split('hpCap')[1]
                        cap = int(cap)
                        if creature.getHP() > cap:
                            perTarget.append(0)
                            specialCase = True
                            break
                    elif "only" in note.lower():
                        if isinstance(creature, Player):
                                if "humanoid" not in note.lower():
                                    perTarget.append(0)
                                    specialCase = True
                                    break
                        elif creature.getCreatureType().lower() not in note.lower():
                            perTarget.append(0)
                            specialCase = True
                            break
                    elif "immune" in note.lower():
                        if isinstance(creature, Player):
                            if "humanoid" not in note.lower():
                                perTarget.append(0)
                                specialCase = True
                                break
                        elif creature.getCreatureType().lower() in note.lower():
                            perTarget.append(0)
                            specialCase = True
                            break
                    elif "turn" in note.lower():
                        turnCount = int(note.lower().split("turn")[0])
            if specialCase:
                continue

        if action.getMean() != 0:
            dmg = computePerTargetDamageImpact(action, creature)
        else:
            dmg = 0
        cond = 0
        statEff = 0
        lvlW = 0
        if not isinstance(action, Weapon) and not layeredRecursion and not leRecursion:
            cond = computePerTargetConditionImpact(action, probSuccess, creature)
            statEff = computePerTargetSEImpact(action, probSuccess, creature)
            if turnCount != 0:
                cond *= min((turnCount * .25), 1)
                statEff *= min((turnCount * .25), 1)
        if isinstance(action, Spell) and action.getLvl() >= 1:
            lvlW = math.log(int(action.getLvl()) + 1)

        crW = computeCRWeight(creature)
        impact_i = dmg
        try:
            impact_i += (cond + statEff) * crW
        except:
            try:
                impact_i += (cond + statEff) * math.floor(crW)
            except:
                impact_i += math.floor(crW)
        impact_i += lvlW  #influence from spell level
        perTarget.append(impact_i)

    for i in range(len(perTarget)): #Normalize
        perTarget[i] = round_to_first_nonzero_decimal(perTarget[i])
        perTarget[i] = max(perTarget[i], 0)

    if isinstance(action, Weapon) or action.getNumTarget() == 1:
        if len(perTarget) > 0:
            impact = max(perTarget)
        else:
            impact = 0
    elif action.getNumTarget() == -1:
        impact = avgOverAOETargetsForImpact(perTarget, len(perTarget))
    elif action.getNumTarget() > 1:
        perTarget.sort(reverse=True)
        idx = 0
        bestTargets = []
        while idx < len(perTarget) and idx < action.getNumTarget():
            bestTargets.append(perTarget[idx])
            idx += 1
        if len(bestTargets) != 0:
            impact = sum(bestTargets) / len(bestTargets)
        else:
            impact = 0
    else:
        impact = perTarget[0] if perTarget else 0

    impact += extraImpact
    impact += (impact * lingSImpact)
    impact += lingEffImpact
    # impact = min(impact, 20) #Removing impact cap for data analysis on accuracy
    # console.print(f"{impact} - {action.getName()}")
    return round_to_first_nonzero_decimal(impact)

#RESULT CREATE/SAVE METHODS
def logActionResult(player, action, actionStats, targets, result, extraResult):
    actionName, actionProb, actionEDam, actionImpact = actionStats[0], actionStats[1], actionStats[2], actionStats[3]
    if isinstance(action, Spell):
        actionType = f"{action.getLvl()} Spell" if action.getLvl() != -1 else "Basic Action"
        conditions = ""
        if action.getConditions():
            conditions = [c["cond"].lower() if isinstance(c, dict) else c.lower() for c in action.getConditions()]
        statuseffects = ""
        if action.getStatusEffects():
            statuseffects = [se["name"].lower() for se in action.getStatusEffects()]
    else:
        actionType = "Weapon"
        conditions = ""
        statuseffects = ""

    entry = {
        "resultID" : random.randint(1, 9999999999),
        "actor": player.getName(),
        "action": actionName,
        "actionType": actionType,
        "actionProb" : actionProb,
        "actionEDam" : actionEDam,
        "actionImpact" : actionImpact,
        "targets": [t["Statblock"].getName() if isinstance(t, dict) else t.getName() for t in targets],
        "targetCRs": [t["Statblock"].getLevel() if isinstance(t, dict) else t.getLevel() for t in targets],
        "conditions": conditions,
        "statuseffects": statuseffects,
        "outcome": result,
        "extraOutcome" : extraResult,
        "timestamp": datetime.now().strftime("%H:%M:%S")
    }
    return entry
def logLingeringResult(resultID, creatureName, lingType, result):
    entry = {
        "resultID": resultID,
        "creature": creatureName,
        "lingType": lingType,
        "outcome": result,
        "timestamp": datetime.now().strftime("%H:%M:%S")
    }
    return entry
def exportEncounterResultsToExcel(encounter, filename="Encounter_Master.xlsx"):
    """
    Consolidates all results from an encounter (turn actions + lingering effects)
    and appends them to a single Excel workbook for later analysis.
    """

    # --- STEP 1: Gather & flatten results ---
    results = [encounter.getResultByIdx(i) for i in range(encounter.resultSize())]
    if not results:
        print("No results to export.")
        return

    flattened = []
    for entry in results:
        if isinstance(entry, dict):
            flattened.append(entry)
        elif isinstance(entry, list):
            flattened.extend(entry)

    # --- STEP 2: Parse into row dictionaries ---
    rows = []
    for r in flattened:
        rid = r.get("resultID")
        actor = r.get("actor", "Unknown")
        actionName = r.get("action", "Unknown")
        actionType = r.get("actionType", "Unknown")
        if actionType == "Unknown":
            actionType = r.get("type", "Unknown")
        outcome = r.get("outcome", {})

        rollResults = outcome.get("rollResults", [])
        diceResults = outcome.get("diceResults", [])
        targets = r.get("targets", [])
        if not targets and r.get("lingType", ""):
            targets = [r.get("creature", "")]
        targetCRs = r.get("targetCRs", [None]*len(targets))

        rawExpected = r.get("actionEDam", "0.0")
        predictedExpected = 0
        if " - " in rawExpected:
            predictedExpected = rawExpected.split(' - ')[1]
            predictedExpected = predictedExpected.split("W")[0]
            rawExpected = rawExpected.split(' - ')[0]
        rawExpected = float(rawExpected)
        predictedExpected = float(predictedExpected)

        probSuccess = r.get("actionProb", "")
        if ' - ' in probSuccess:
            probSuccess = probSuccess.split(' - ')[0]
            probSuccess = float(probSuccess)

        player = encounter.getPlayerByName(actor)
        action = {}
        if player and not r.get("lingType", ""):
            action = player.getSpellByName(actionName)

        for i, target in enumerate(targets):
            condSev = 0
            seSev = 0
            if action:
                monster = encounter.getMonsterByName(target)
                if not monster:
                    condSev = computePerTargetConditionImpact(action, probSuccess, player, False)
                    seSev = computePerTargetSEImpact(action, probSuccess, player, False)
                else:
                    condSev = computePerTargetConditionImpact(action, probSuccess, monster, False)
                    seSev = computePerTargetSEImpact(action, probSuccess, monster, False)
            if isinstance(targetCRs, list):
                if i < len(targetCRs):
                    targetCR = targetCRs[i]
                else:
                    targetCR = None
            else:
                targetCR = targetCRs
            if isinstance(rollResults, list):
                if i < len(rollResults):
                    rollResult = rollResults[i]
                else:
                    rollResult = None
            else:
                rollResult = rollResults
            if isinstance(diceResults, list):
                if i < len(diceResults):
                    diceResult = diceResults[i]
                else:
                    diceResult = None
            else:
                diceResult = diceResults
            row = {
                "ResultID": rid,
                "Actor": actor,
                "Action Name": actionName,
                "Action Type": actionType,
                "Target": target,
                "Target CR": targetCR,
                "Roll Result": rollResult,
                "Dice Result": diceResult,
                "probSuccess": r.get("actionProb"),
                "RAW Expected Damage": rawExpected,
                "Predicted Expected Damage": predictedExpected,
                "Impact Rating": r.get("actionImpact"),
                "Condition Severity": condSev,
                "Status Effect Severity": seSev,
                "Conditions Applied": ", ".join(r.get("conditions", [])) if r.get("conditions") else None,
                "Status Effects Applied": ", ".join(
                    [s["name"] for s in r.get("statuseffects", [])]
                ) if r.get("statusEffects") else None,
                "lingType": r.get("lingType", ""),
                "Encounter Name": encounter.getName()
            }
            rows.append(row)

    # --- STEP 3: Write or append to Excel ---
    import os
    file_exists = os.path.exists(filename)
    if not file_exists:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "All Encounters"
        headers = list(rows[0].keys())
        ws.append(headers)
        wb.save(filename)  # <-- FIRST SAVE
        print(f"[INIT] Created new workbook '{filename}'")

    wb = openpyxl.load_workbook(filename)
    ws = wb["All Encounters"] if "All Encounters" in wb.sheetnames else wb.create_sheet("All Encounters")

    # for row in rows:
        # ws.append([row.get(col) for col in ws[1]])
    for rowDict in rows:
        ws.append(list(rowDict.values()))

    for col_idx in range(1, ws.max_column + 1):
        max_len = max(len(str(ws.cell(row=r, column=col_idx).value)) for r in range(1, ws.max_row + 1))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(max_len + 2, 10)

    wb.save(filename)
    print(f"✅ Encounter '{encounter.getName()}' results appended to {filename}")
    wb.close()

#ENCOUNTER RUNTIME METHODS
def resolve_summon_spell(encounter, player, spell, initiative_list):
    def _load_monsters_indexed():
        with open(MONSTER_LIST_FILE, "r") as f:
            mons = json.load(f)
        # index by name for quick lookup
        by_name = {m["name"]: m for m in mons}
        return mons, by_name
    def _build_summon_menu_tree(spell):
        tree = Tree(f"[bold]{spell.getName()}[/bold] – Summon options")
        # collect all Summon SEs
        se_list = [se for se in (spell.getStatusEffects() or []) if se["name"].lower() == "summon"]
        # group by crCap for quick presentation
        grouped = {}
        for se in se_list:
            eff = se["effect"]
            cap = eff.get("crCap", "").strip()
            n = int(eff.get("numSummons", "1"))
            mname = eff.get("monsterName", "").strip()
            mtype = eff.get("type", "").strip()
            key = cap or "ANY"
            entry = grouped.setdefault(key, {"total": 0, "fixed": set(), "types": set()})
            entry["total"] += n
            if mname:
                entry["fixed"].add(mname)
            if mtype:
                entry["types"].add(mtype)
        for cr_cap, info in grouped.items():
            branch = tree.add(f"CR ≤ {cr_cap}  • total summons: {info['total']}")
            if info["fixed"]:
                branch.add(f"Fixed names only: {sorted(info['fixed'])}")
            if info["types"]:
                branch.add(f"Allowed types: {sorted(info['types'])}")
        return tree
    """
    Creates monsters for a 'Summon' spell and inserts them into the Encounter and the initiative.

    Parameters
    ----------
    encounter : Encounter
    caster    : Player (or creature casting)
    spell     : Spell (must contain one or more 'Summon' status effects)
    resultID  : int (the result id for this cast, used on SwitchSides effect)
    initiative_list : list of dicts (your runtime initiative array; will be mutated)

    Returns
    -------
    list[Monster] : the list of created allies
    """
    resultID = random.randint(1, 9999999999)
    # 1) Collect all Summon effects from the spell
    summon_effects = [se for se in (spell.getStatusEffects() or []) if se["name"].lower() == "summon"]
    if not summon_effects:
        console.print("[yellow]No 'Summon' effects found on this spell.[/yellow]")
        return [], initiative_list, []

    # 2) Present a summary tree to the user
    console.print(_build_summon_menu_tree(spell))

    # 3) Let the user choose a CR track to follow (if multiple caps exist)
    cr_caps = sorted({se["effect"].get("crCap", "").strip() for se in summon_effects if se.get("effect")})
    cr_caps = [c for c in cr_caps if c] or ["ANY"]
    if len(cr_caps) > 1:
        console.print("\nChoose a CR cap to use for this cast:")
        for i, cap in enumerate(cr_caps, 1):
            console.print(f"{i}. CR ≤ {cap}")
        choice_idx = -1
        while choice_idx < 1 or choice_idx > len(cr_caps):
            try:
                choice_idx = int(input(f"CR Cap(0-{len(cr_caps)})> ").strip())
            except ValueError:
                choice_idx = -1
        chosen_cap = cr_caps[choice_idx - 1]
    else:
        chosen_cap = cr_caps[0]

    # 4) Determine total summons & constraints from chosen track
    total_to_summon = 0
    fixed_names = set()
    allowed_types = set()

    for se in summon_effects:
        eff = se["effect"]
        cap = (eff.get("crCap", "") or "ANY").strip()
        if cap == chosen_cap:
            total_to_summon = max(total_to_summon, int(eff.get("numSummons", "1")))
            if eff.get("monsterName"):
                fixed_names.add(eff["monsterName"].strip())
            if eff.get("type"):
                allowed_types.add(eff["type"].strip())

    if total_to_summon <= 0:
        console.print("[yellow]No summons required for the chosen CR track.[/yellow]")
        return [], initiative_list, []

    # 5) Load monsters and build the candidate pool
    all_monsters, by_name = _load_monsters_indexed()

    def _candidate(mon):
        # CR gate
        if chosen_cap != "ANY":
            try:
                if _cr_to_float(mon["cr"]) > _cr_to_float(chosen_cap):
                    return False
            except Exception:
                return False
        # type gate
        if allowed_types and mon.get("creatureType") not in allowed_types:
            return False
        # fixed-names gate (if present, restrict to those only)
        if fixed_names and mon["name"] not in fixed_names:
            return False
        return True

    candidates = [m for m in all_monsters if _candidate(m)]
    if not candidates:
        console.print("[red]No valid monster options for the chosen CR cap / constraints.[/red]")
        return [], initiative_list, []

    # 6) Ask for a shared initiative value (all summons share the same)
    try:
        shared_i = int(input("Initiative value for all summoned creatures: ").strip())
    except ValueError:
        shared_i = random.randint(1, 20)

    # 7) Interactive picking with a 'Repeat' option
    console.print(f"\nPick [bold]{total_to_summon}[/bold] monster(s). Type the exact name, or 'Repeat' to duplicate the last choice.")

    # quick visual list (name - CR)
    chunk = 4
    rows = [f"{m['name']} (CR {m['cr']})" for m in candidates]
    for i in range(0, len(rows), chunk):
        console.print(" " + " | ".join(rows[i:i + chunk]))

    created = []
    last_choice = None
    made = 0
    repeatOnly = False
    while made < total_to_summon:
        if not repeatOnly:
            pick = input(f"Summon {made+1} of {total_to_summon}: ").strip()
        if pick.lower() == "repeat":
            if not last_choice:
                console.print("[yellow]No previous choice to repeat.[/yellow]")
                continue
            chosen = by_name.get(last_choice)
            if not chosen or not _candidate(chosen):
                console.print("[red]Previous choice no longer valid under constraints.[/red]")
                repeatOnly = False
                continue
            repeatOnly = True
        else:
            chosen = by_name.get(pick)
            if not chosen or not _candidate(chosen):
                console.print("[red]Unknown/invalid monster for this summon. Try again.[/red]")
                continue
            last_choice = pick

        # 8) Build the Monster object (mirrors your encounter creation flow)
        name = chosen["name"]
        # ensure uniqueness in encounter by suffixing an index if needed
        suffix = 1
        existing_names = {encounter.getMonster(i).getName() for i in range(encounter.monsterSize())}
        unique_name = name
        while unique_name in existing_names:
            suffix += 1
            unique_name = f"{name}S{suffix}"

        cr = chosen["cr"]
        cType = chosen["creatureType"]
        stats = chosen["statArray"]
        hp = chosen["hit_points"]
        maxHP = hp
        ac = chosen["AC"]
        saves = chosen["saveProfs"]
        lResists = chosen["lResists"]
        damResists = chosen["damResists"]
        damImmunes = chosen["damImmunes"]
        damVulns = chosen["damVulns"]
        conImmunes = chosen["conImmunes"]
        lairActions = False if chosen["lairAction"] == "False" else True
        enemy = False  # summoned to your side
        actions = loadMonsterActions(chosen)
        spellInfo = loadMonsterSpells(chosen)
        magicResist = chosen.get("magicResist", False)
        legActions = chosen.get("legActions", [])

        mon_obj = Monster(unique_name, cr, cType, stats, hp, maxHP, ac, saves,
                          lResists, damResists, damImmunes, damVulns,
                          conImmunes, [], [], lairActions,
                          magicResist, enemy, actions, spellInfo, legActions)

        # 9) Add SwitchSides status effect with resultID
        switch_effect = {
            "name": "SwitchSides",
            "effect": {
                "attribute": ["T"],
                "resultID": [resultID]
            }
        }
        # Uses your existing status-effect plumbing on the creature object
        mon_obj.addStatusEffect(switch_effect)  # method provided by your Stats/Monster class

        # 10) Insert into Encounter and initiative
        encounter.addMonster(mon_obj)
        initiative_list.append({
            "name": mon_obj.getName(),
            "iValue": shared_i,
            "turnType": "Monster",
            "currentTurn": False,
            "Statblock": mon_obj
        })
        if TEST_MODE:
            input(f"ADD SUMMON ({mon_obj.getName()}) of CR ({mon_obj.getLevel()}) TO iVALUE ({shared_i})")

        created.append(mon_obj)
        made += 1

    # 11) (Optional) keep initiative order stable: sort by iValue then by a DEX tiebreak like your setInitiative does
    def sort_key(c):
        name = c["name"]
        i_value = c["iValue"]

        if c["turnType"] == "Player":
            for i in range(encounter.playerSize()):
                player = encounter.getPlayer(i)
                if player.getName() == name:
                    return i_value, player.getStat("DEX")
        elif c["turnType"] == "Monster":
            for i in range(encounter.monsterSize()):
                monster = encounter.getMonster(i)
                if monster.getName() == name:
                    return i_value, monster.getStat("DEX")

        # fallback (shouldn't happen)
        return i_value, 0

    initiative_list.sort(key=sort_key, reverse=True)

    console.print(f"[green]{len(created)} creature(s) summoned and added to initiative.[/green]")
    createdNames = [c.getName() for c in created]
    # All summon spells take concentration
    concEffect = {
        "name": "Concentration",
        "effect": {
            "resultID": resultID,
            "concentrationTargets": createdNames,
            "summonConc" : True
            # Whatever we end up summoning - keep a list.
        }
    }
    for se in player.getActiveStatusEffects():
        if se["name"].lower() == "concentration":
            endConcentration(player, se, initiative_list)
            if encounter.monsterSize() != len(initiative_list):
                summonCleanUp(encounter, initiative_list)
            break
    player.addStatusEffect(concEffect)

    summonResult = {
        "resultID" : resultID,
        "actor": player.getName(),
        "targets": createdNames
    }
    return summonResult, initiative_list, createdNames
def endConcentration(player, concentration, initiative):
    if isinstance(player, dict):
        player = player["Statblock"]
    concTargets = concentration["effect"]["concentrationTargets"]
    summon = False
    if "summonConc" in concentration["effect"]:
        summon = True
    player.removeStatusEffect("concentration")
    cIdx = 0
    while cIdx < len(initiative):
        creature = initiative[cIdx]["Statblock"]
        summonedCreature = False
        if creature.getName() in concTargets:
            if creature.getActiveStatusEffects():
                if any(
                        isinstance(statusEffect["effect"]["resultID"], list) and
                        concentration["effect"]["resultID"] in statusEffect["effect"]["resultID"]
                        for statusEffect in creature.getActiveStatusEffects()
                ):
                    #lingEffect has a list of resultIDs. Want to match with the ONE concentrationID, and remove it.
                    statEffects = creature.getActiveStatusEffects()
                    for i in range(len(statEffects)):
                        #Removes creature if it is a summoned creature
                        if summon:
                            if concentration["effect"]["resultID"] in statEffects[i]["effect"]["resultID"]:
                                if TEST_MODE:
                                    input(f"SUMMON - DELETE CREATURE ({initiative[cIdx]["Statblock"].getName()})")
                                del initiative[cIdx]
                                summonedCreature = True
                                break

                        #Skip over any concentration effects the creature currently has
                        if not isinstance(statEffects[i]["effect"]["resultID"], list):
                            continue
                        # Remove any attributes associated with the ID
                        for j, resID in enumerate(statEffects[i]["effect"]["resultID"]):
                            if concentration["effect"]["resultID"] == resID:
                                del statEffects[i]["effect"]["resultID"][j]
                                if statEffects[i]["name"].lower() not in ["lingeffect", "lingsave"]:
                                    del statEffects[i]["effect"]["attribute"][j]
                                else:
                                    if TEST_MODE:
                                        input(f"REMOVING ACTIVE_PRE_EFFECT ({statEffects[i]["name"]}) WITH SPELL ({statEffects[i]["effect"]["spell"][j]["spellname"]})")
                                    del statEffects[i]["effect"]["spell"][j]
                    if summonedCreature:
                        continue
                    seIdx = 0
                    while seIdx < len(statEffects):
                        statEffect = statEffects[seIdx]
                        if isinstance(statEffect["effect"]["resultID"], list) and len(statEffect["effect"]["resultID"]) == 0:
                            #If no more resultID's left, then statusEffect is no longer active.
                            creature.removeStatusEffect(statEffect["name"])
                        else:
                            seIdx += 1
            if creature.getActiveConditions():
                if any(isinstance(condition, dict) and concentration["effect"]["resultID"] in condition["resultID"]
                        for condition in creature.getActiveConditions()
                ):
                    conditions = creature.getActiveConditions()
                    cidx = 0
                    while cidx < len(conditions):
                            for j in range(len(conditions[cidx]["resultID"])):
                                if concentration["effect"]["resultID"] == conditions[cidx]["resultID"][j]:
                                    del conditions[cidx]["resultID"][j]
                                    break
                            if len(conditions[cidx]["resultID"]) == 0:
                                creature.removeCondition(conditions[cidx]["cond"])
                                continue
                            cidx += 1

        if not summonedCreature:
            cIdx += 1
def summonCleanUp(encounter, initiative):
    #Iterate through initiative - if the monster is IN encounter but NOT in initiative, then delete from encounter.
    eIdx = 0
    while eIdx < encounter.monsterSize():
        foundMonster = False
        for cIdx, creature in enumerate(initiative):
            creature = creature["Statblock"]
            if encounter.getMonster(eIdx).getName().lower() == creature.getName().lower():
                foundMonster = True
                break
        if not foundMonster:
            encounter.removeMonster(encounter.getMonster(eIdx).getName())
        else:
            eIdx += 1
    iIdx = 0
    storedInitiative = encounter.getInitiative()
    while iIdx < len(storedInitiative):
        foundMonster = False
        foundPlayer = False
        pIdx = 0
        mIdx = 0
        while pIdx < encounter.playerSize() and not foundPlayer:
            creature = encounter.getPlayer(pIdx)
            if creature.getName() == storedInitiative[iIdx]["name"]:
                foundPlayer = True
            else:
                pIdx += 1
        if not foundPlayer:
            while mIdx < encounter.monsterSize() and not foundMonster:
                creature = encounter.getMonster(mIdx)
                if creature.getName() == storedInitiative[iIdx]["name"]:
                    foundMonster = True
                else:
                    mIdx += 1
        if not foundPlayer and not foundMonster:
            del storedInitiative[iIdx]
        else:
            iIdx += 1

def executePlayerAction(player, action, selectedTargets, actionResult, initiative):
    def applyEffectToTarget(creature, succeeded, damage, action, resultID):
        """Applies HP change, downed/dead logic, conditions, and status effects."""
        rollType = action.getRollType() if isinstance(action, Spell) else "tohit"

        downed_before = creature.isActiveCondition("Downed")
        stable_before = creature.isActiveCondition("Stabilized")

        if isinstance(action, Spell) and not isinstance(action.getDamType(),
                                                        list) and action.getDamType().lower() == "healing":
            creature.setHP(min(creature.getMaxHP(), creature.getHP() + damage))
        else:
            creature.setHP(creature.getHP() - damage)
        creature.setHP(math.floor(creature.getHP()))

        if creature.isActiveCondition("downed") and damage > 0:
            removeCondition("downed", creature)
            addCondition("dead", creature, -1)
        if creature.getHP() <= 0:
            creature.setHP(0)
            if isinstance(creature, Player):
                addCondition("Downed" if damage < (creature.getMaxHP() + creature.getHP()) else "Dead", creature, resultID)
            else:
                addCondition("Dead", creature, resultID)

        if creature.getHP() > 0:
            if downed_before:
                creature.removeCondition("Downed")
            if stable_before:
                creature.removeCondition("Stabilized")

        if (rollType == 'save' and not succeeded) or (rollType != 'save' and succeeded):
            if isinstance(action, Spell) and action.getConditions():
                for cond in action.getConditions():
                    addCondition(cond, creature, resultID)
            if isinstance(action, Spell) and action.getStatusEffects():
                for effect in action.getStatusEffects():
                    if effect["name"].lower() != "concentration":
                        addStatusEffect(effect, creature, resultID)
            if isinstance(action, Spell) and action.getLingSaves():
                if creature.isActiveStatusEffect("lingsave"):
                    lingSaves = creature.getActiveStatusEffect("lingsave")
                    if not any(resultID == rID for rID in lingSaves["effect"]["resultID"]):
                        lingSaves["effect"]["spell"].append(action.toDict())
                        lingSaves["effect"]["resultID"].append(actionResult["resultID"])
                else:
                    newLingSave = {
                        "name": "lingSave",
                        "effect": {
                            "spell": [action.toDict()],
                            "resultID": [
                                actionResult["resultID"]
                            ]
                        }
                    }
                    addStatusEffect(newLingSave, creature, actionResult["resultID"])
        return creature

    outcomes = actionResult['outcome']["rollResults"]
    damages = actionResult["outcome"]["diceResults"]

    if len(damages) == 1 and len(selectedTargets) != 1:
        damages = [damages[0]] * len(selectedTargets)

    if isinstance(action, Spell) and action.getStatusEffects() and \
            "concentration" in [se["name"].lower() for se in action.getStatusEffects()]:
        concEffect = {
            "name": "Concentration",
            "effect": {
                "resultID": actionResult["resultID"],
                "concentrationTargets": [t["Statblock"].getName() if isinstance(t, dict) else t.getName() for t in selectedTargets]
            }
        }
        for se in player.getActiveStatusEffects():
            if se["name"].lower() == "concentration":
                if any(effect in se["effect"]["concentrationTargets"] for effect in concEffect["effect"]["concentrationTargets"]):
                    #Edge case for summoned concentrationTargets
                    oldTargets = se["effect"]["concentrationTargets"]
                    endConcentration(player, se, initiative)
                    cetidx = 0
                    while cetidx < len(concEffect["effect"]["concentrationTargets"]):
                        cet = concEffect["effect"]["concentrationTargets"][cetidx]
                        if cet in oldTargets and cet not in [c["name"] for c in initiative]:
                            del concEffect["effect"]["concentrationTargets"][cetidx]
                            del actionResult["targets"][cetidx]
                            del actionResult["outcome"]["diceResults"][cetidx]
                            del actionResult["outcome"]["rollResults"][cetidx]
                            del selectedTargets[cetidx]
                            continue
                        cetidx += 1
                else:
                    endConcentration(player, se, initiative)
                break
        player.addStatusEffect(concEffect)

    for idx, target in enumerate(selectedTargets):
        creature = target["Statblock"] if isinstance(target, dict) else target
        succeeded = outcomes[idx].lower() in ("y", "crit")
        damage = damages[idx]
        creature = applyEffectToTarget(creature, succeeded, damage, action, actionResult["resultID"])
        if isinstance(action, Spell) and action.getLingEffects():
            transLingEffect = translateLingEffect(action, action.getLingEffects(), player.getSpellMod())
            if creature.isActiveStatusEffect("lingEffect"):
                lingEffect = creature.getActiveStatusEffect("lingEffect")
                lingEffect["effect"]["spell"].append(transLingEffect.toDict())
                lingEffect["effect"]["resultID"].append(actionResult["resultID"])
            else:
                newLingEffect = {
                    "name": "lingEffect",
                    "effect": {
                        "spell": [transLingEffect.toDict()],
                        "resultID": [actionResult["resultID"]]
                    }
                }
                addStatusEffect(newLingEffect, creature, actionResult["resultID"])

    # ---- EXTRA EFFECT PASS (if any) ----
    extra = actionResult.get("extraOutcome", None)
    if extra:
        extraOutcomes = extra["extraRollResults"]
        extraDamages = extra["extraDiceResults"]

        if len(extraDamages) == 1 and len(selectedTargets) != 1:
            extraDamages = [extraDamages[0]] * len(selectedTargets)

        for idx, target in enumerate(selectedTargets):
            creature = target["Statblock"]
            succeededExtra = extraOutcomes[idx].lower() in ("y", "crit")
            damageExtra = extraDamages[idx]
            extraEffect = action.getExtraEffect()
            extraRollType = extraEffect["rolls"]["rollType"]

            downed_before = creature.isActiveCondition("Downed")
            stable_before = creature.isActiveCondition("Stabilized")

            if isinstance(action, Spell) and not isinstance(action.getDamType(),
                                                            list) and action.getDamType().lower() == "healing":
                creature.setHP(min(creature.getMaxHP(), creature.getHP() + damageExtra))
            else:
                creature.setHP(creature.getHP() - damageExtra)
            if creature.getHP() <= 0:
                creature.setHP(0)
                if isinstance(creature, Player):
                    addCondition("Downed" if damageExtra < (creature.getMaxHP() + creature.getHP()) else "Dead", creature,
                                 actionResult["resultID"])
                else:
                    addCondition("Dead", creature, actionResult["resultID"])

            if creature.getHP() > 0:
                if downed_before:
                    creature.removeCondition("Downed")
                if stable_before:
                    creature.removeCondition("Stabilized")

            if (extraRollType == 'save' and not succeededExtra) or (extraRollType != 'save' and succeededExtra):
                if "conditions" in extraEffect and extraEffect["conditions"]:
                    for cond in extraEffect["conditions"]:
                        addCondition(cond, creature, actionResult["resultID"])
                if "statusEffect" in extraEffect and extraEffect["statusEffect"]:
                    for effect in extraEffect["statusEffect"]:
                        if effect["name"].lower() != "concentration":
                            addStatusEffect(effect, creature, actionResult["resultID"])
    if isinstance(action, Spell) and action.getSpecialNotes():
        specialNotes = action.getSpecialNotes()
        for note in specialNotes:
            if "turn" in note.lower():
                actionResult["turnCount"] = 0
                actionResult["turnCap"] = int(note.lower().split("turn")[0])
                break
def playerActionResolution(player, action, actionStats, initiative):
    def selectTargets():
        indices = []
        if numTargets == 0:
            return [player]
        for i in range(numTargets):
            goodInput = False
            AOETargeting = False if not isinstance(action, Spell) or action.getNumTarget() != -1 else True
            earlyExit = False
            while not goodInput:
                try:
                    idx = input(f"TARGET {i + 1} OF {numTargets} - Select target by index: ('Exit' to exit)")
                    if idx == 'Exit' or idx == 'exit':
                        if len(indices) == 0:
                            print("Need atleast ONE target!")
                        else:
                            earlyExit = True
                            break
                    else:
                        idx = int(idx)
                        if idx < 0 or idx > len(validTargets):
                            raise ValueError("Bad input! Must select by monster index.")
                        else:
                            if AOETargeting and idx in indices:
                                raise ValueError("Bad input! Cannot target same creature twice in an AOE.")
                            indices.append(idx)
                            goodInput = True
                except ValueError as e:
                    print(e)
            if earlyExit:
                break
        if len(indices) > 0:
            return [validTargets[i] for i in indices]
        else:
            return validTargets[0]

    """
    Resolves a player's selected action by:
        1. Asking the user which target(s) were affected.
        2. Asking if the action succeeded or failed.
        3. Logging the outcome to the encounter record.
        4. Updating monsters' HP and active status effects.

    Returns:
        Updated initiative list
    """
    if isinstance(player, Player):
        isPlayerTurn=True
    else:
        isPlayerTurn=False
    print(f"\n{player.getName()} used {action.getName()}!")

    if isinstance(action, Spell) and action.getRollType().lower() == "onhit" and player.getWeaponLength() == 0:
        print("Cannot use onHit action with no weapon to hit with!")
        return {}, initiative

    if isinstance(action, Spell) and action.getSelfTarget():
        selectedTargets = [player]
        result = ['y']
    else:
        # Step 1: List valid targets
        if not isinstance(action.getDamType(), list) and action.getDamType().lower() == "healing":
            validTargets = [
                creature for creature in initiative
                if isValidTarget(action, creature, isPlayerTurn)
            ]
        else:
            validTargets = [
                creature for creature in initiative
                if isValidTarget(action, creature, isPlayerTurn)
            ]

        # Step 2: Prompt for target(s)
        for i, target in enumerate(validTargets):
            print(f"[{i}] {target['Statblock'].getName()} (HP: {target['Statblock'].getHP()})")

        numTargets = action.getNumTarget() if isinstance(action, Spell) else 1
        numTargets = min(numTargets, len(validTargets)) if numTargets != -1 else len(validTargets)

        selectedTargets = selectTargets()

        # Step 3: Ask for success or failure
        result = getRollResults(action, selectedTargets)

    # Get dice results
    dams = getDiceResults(action, selectedTargets, result)

    # Step 4: Log results
    result = {
        "rollResults": result,
        "diceResults": dams
    }

    checkExtraEffects = True if isinstance(action, Spell) and action.getExtraEffect() else False
    if checkExtraEffects:
        extraEffect = translateLingEffect(action, action.getExtraEffect(), player.getSpellMod())
        print("EXTRA EFFECT")
        extraResult = getRollResults(extraEffect, selectedTargets)
        extraDams = getDiceResults(extraEffect, selectedTargets, extraResult)
        extraResult = {
            "extraRollResults": extraResult,
            "extraDiceResults": extraDams
        }
    else:
        extraResult = {}
    actionResult = logActionResult(player, action, actionStats, selectedTargets, result, extraResult)
    console.print(actionResult)
    # Step 5: Apply effects
    executePlayerAction(player, action, selectedTargets, actionResult, initiative)

    return actionResult, initiative
def endSpellEffect(effect, idx, creature, initiative):
    #Ends any long-lasting effect that a creature has from a given spell
    # - and ends concentration if nobody else is under that spell.
    if "effect" in effect:
        effectID = effect["effect"]["resultID"][idx]
        if idx == 0:
            removeStatusEffect(effect["name"], creature)
        else:
            del effect["effect"]["resultID"][idx]
            if effect["name"].lower() in ["lingsave", "lingeffect"]:
                del effect["effect"]["spell"][idx]
            else:
                del effect["effect"]["attribute"][idx]
    elif "cond" in effect:
        effectID = effect["resultID"][idx]
        if idx == 0:
            removeCondition(effect["cond"], creature)
        else:
            del effect["resultID"][idx]
    else:
        removeCondition(effect, creature)
        return
    #Removes associated statEffects and conditions from creature.
    for condition in creature.getActiveConditions():
        if isinstance(condition, dict):
            for ri, resultID in enumerate(condition["resultID"]):
                if effectID == resultID:
                    del condition["resultID"][ri]
                    break
            if len(condition["resultID"]) == 0:
                removeCondition(condition["cond"], creature)
    for se in creature.getActiveStatusEffects():
        if se["name"].lower() != "concentration" and effectID in se["effect"]["resultID"]:
            for ri, resultID in enumerate(se["effect"]["resultID"]):
                if effectID == resultID:
                    endSpellEffect(se, ri, creature, initiative)
                    break

    for target in initiative:
        target = target["Statblock"]
        if target.isActiveStatusEffect("concentration"):
            concEffect = target.getActiveStatusEffect("concentration")
            if concEffect["effect"]["resultID"] == effectID:
                remainingTargets = False
                for effectTarget in initiative:
                    effectTarget = effectTarget["Statblock"]
                    if effectTarget.isActiveStatusEffect(effect["name"]) and \
                            effectID in effectTarget.getActiveStatusEffect(effect["name"])["effect"]["resultID"]:
                        remainingTargets = True
                        break
                if not remainingTargets:
                    endConcentration(target, concEffect, initiative)
                break
def preTurnCheck(creature, encounter, initiative):
    def procPreEffect(lingSpell):
        # HP logic
        if lingSpell["targeting"][0]["rolls"]["damage"]:
            damage = int(input(f"{lingSpell["spellname"]} - How much damage was applied?"))
        else:
            damage = 0
        if lingSpell["targeting"][0]["extraEffect"]:
            if lingSpell["targeting"][0]["extraEffect"]["rolls"]["damage"]:
                damage += int(input("EXTRA EFFECT\nHow much damage was applied?"))
        damType = lingSpell["targeting"][0]["damType"] if "damType" in lingSpell["targeting"][0] and \
                                                       lingSpell["targeting"][0]["damType"] else ""
        if isinstance(damType, list) and len(damType) == 1:
            damType = damType[0]

        downed_before = creature.isActiveCondition("Downed")
        stable_before = creature.isActiveCondition("Stabilized")
        if not isinstance(damType, list) and damType.lower() == "healing":
            creature.setHP(min(creature.getMaxHP(), creature.getHP() + damage))
        else:
            creature.setHP(creature.getHP() - damage)

        if damage > 0 and creature.isActiveStatusEffect("concentration"):
            endConcentration(creature, creature.getActiveStatusEffect("concentration"), initiative)
            if encounter.monsterSize() != len(initiative):
                summonCleanUp(encounter, initiative)

        if creature.isActiveCondition("downed") and damage > 0:
            removeCondition("downed", creature)
            addCondition("dead", creature, -1)

        if creature.getHP() <= 0:
            creature.setHP(0)
            if isinstance(creature, Player):
                addCondition("Downed" if damage < (creature.getMaxHP() + creature.getHP()) else "Dead", creature, resultID)
            else:
                addCondition("Dead", creature, resultID)

        if creature.getHP() > 0:
            if downed_before:
                creature.removeCondition("Downed")
            if stable_before:
                creature.removeCondition("Stabilized")
        # Dont need to do statEffects/conditions, since they will linger until removed.
        return damage
    def executeLingSaves(lingSaveEffect, idx):
        idx = 0 if idx == -1 else idx
        ensureList(lingSaveEffect["effect"]["resultID"])
        lingResultInput = input(f"Lingering Save - {lingSaveEffect["effect"]["spell"][idx]["spellname"]}:\nDid the creature save? (y/n)").lower()
        while lingResultInput not in ['y', 'n'] and lingResultInput is not None:
            print("BAD INPUT")
            lingResultInput = input(f"Lingering Save - {lingSaveEffect["effect"]["spell"][idx]["spellname"]}:\nDid the creature save? (y/n)").lower()
        if lingResultInput == 'y':
            endSpellEffect(lingSaveEffect, idx, creature, initiative)
            if encounter.monsterSize() != len(initiative):
                summonCleanUp(encounter, initiative)
            return lingResultInput, 0
        else:
            lingSaveSpell = lingSaveEffect["effect"]["spell"][idx]
            lingDamInput = procPreEffect(lingSaveSpell)
        return lingResultInput, lingDamInput

    creature = creature["Statblock"] if isinstance(creature, dict) else creature
    preEffects = []
    appendTurnCountResID = []
    for effect in creature.getActiveStatusEffects():
        if effect["name"].lower() in ["lingsave", "lingeffect"]:
            preEffects.append(effect)

        #Deals with 1Turn shenanigans
        #NOTE: Potentially have this work with durations?
        # Like mirror image would be a 10Turn specialNotes, and then check it here.
            # Out of scope for project - only saying here for future reference.
        resultIDs = effect["effect"]["resultID"]
        resultIDs = ensureList(resultIDs)
        for i, resultID in enumerate(resultIDs):
            if resultID != -1:
                result = encounter.getResultByID(resultID)
                if "turnCount" in result and "turnCap" in result:
                    if int(result["turnCount"]) >= int(result["turnCap"]):
                        endSpellEffect(effect, i, creature, initiative)
                        if encounter.monsterSize() != len(initiative):
                            summonCleanUp(encounter, initiative)
                    else:
                        result["turnCount"] += 1
                        appendTurnCountResID.append(resultID)
    for cond in creature.getActiveConditions(): #Downed logic
        if isinstance(cond, dict):
            resultIDs = cond["resultID"]
            resultIDs = ensureList(resultIDs)
            for i, resultID in enumerate(resultIDs):
                if resultID != -1:
                    result = encounter.getResultByID(resultID)
                    if "turnCount" in result and "turnCap" in result and resultID not in appendTurnCountResID:
                        if int(result["turnCount"]) >= int(result["turnCap"]):
                            endSpellEffect(cond, i, creature, initiative)
                            if encounter.monsterSize() != len(initiative):
                                summonCleanUp(encounter, initiative)
                        else:
                            result["turnCount"] += 1
                            appendTurnCountResID.append(resultID)
        if (not isinstance(cond, dict) and cond.lower() == "downed") or isinstance(cond, dict) and cond["cond"].lower() == "downed":
            stillAlive = input("Is the creature still alive? (y/n/stable)").lower()
            while stillAlive not in ["y", "n", "stable"]:
                print("BAD INPUT")
                stillAlive = input("Is the creature still alive? (y/n)").lower()
            if stillAlive == "n":
                removeCondition("downed", creature)
                if isinstance(cond, dict):
                    addCondition("dead", creature, cond["resultID"][0])
                else:
                    addCondition("dead", creature, -1)
            elif stillAlive == "stable":
                removeCondition("downed", creature)
                if isinstance(cond, dict):
                    addCondition("stabilized", creature, cond["resultID"][0])
                else:
                    addCondition("stabilized", creature, -1)
    preTurnLogs = []
    if preEffects:
        for effect in preEffects:
            resultInput = 'y'
            damInput = 0
            if effect["name"].lower() == "lingsave":
                if len(effect["effect"]["resultID"]) > 1:
                    addedToLog = False
                    for i in range(len(effect["effect"]["resultID"])):
                        resultInput, damInput = executeLingSaves(effect, i)
                        resultInput = {
                            "rollResults": resultInput,
                            "diceResults": damInput
                        }
                        if len(effect["effect"]["resultID"]) > i > 0:
                            preTurnLogs.append(logLingeringResult(effect["effect"]["resultID"][i], creature.getName(), effect["name"].lower(), resultInput))
                            addedToLog = True
                    if not addedToLog:
                        preTurnLogs.append({})
                else:
                    resultInput, damInput = executeLingSaves(effect, 0)
                    resultInput = {
                        "rollResults": resultInput,
                        "diceResults": damInput
                    }
                    preTurnLogs.append(logLingeringResult(effect["effect"]["resultID"][0], creature.getName(), effect["name"].lower(), resultInput))
            elif effect["name"].lower() == "lingeffect":
                if len(effect["effect"]["resultID"]) > 1:
                    for i in range(len(effect["effect"]["resultID"])):
                        name = encounter.getResultByID(effect["effect"]["resultID"][i])["action"]
                        try:
                            effectRolls = effect["effect"]["spell"][i]["targeting"]["rolls"]
                        except:
                            effectRolls = {}
                        resultInput = input(f"Lingering Effect - {name}:\nDid the creature save? (y/n)") if effectRolls and effectRolls[
                                                                                        "rollType"].lower() == "save" else "y"
                        while resultInput not in ['y', 'n']:
                            print("BAD INPUT")
                            resultInput = input(f"Lingering Effect - {name}:\nDid the creature save? (y/n)")
                        if not effectRolls or effectRolls["halfSave"] and resultInput == 'n':
                            resultInput = {
                                "rollResults": resultInput,
                                "diceResults": 0
                            }
                            preTurnLogs.append(
                                logLingeringResult(effect["effect"]["resultID"][i], creature.getName(), effect["name"].lower(), resultInput))
                        else:
                            # lingEffect here will never be "repeat", so warning is invalid.
                            damInput = procPreEffect(effect["effect"]["spell"][i])
                            resultInput = {
                                "rollResults": resultInput,
                                "diceResults": damInput
                            }
                            preTurnLogs.append(
                                logLingeringResult(effect["effect"]["resultID"][i], creature.getName(), effect["name"].lower(), resultInput))
                else:
                    name = encounter.getResultByID(effect["effect"]["resultID"][0])["action"]
                    # print(f"Lingering Effect: {name}")
                    effectRolls = effect["effect"]["spell"][0]["targeting"]["rolls"]
                    resultInput = input(f"Lingering Effect - {name}:\nDid the creature save? (y/n)") if effectRolls["rollType"].lower() == "save" else "y"
                    while resultInput not in ['y', 'n']:
                        print("BAD INPUT")
                        resultInput = input(f"Lingering Effect - {name}:\nDid the creature save? (y/n)")
                    if not effect["effect"]["spell"][0]["targeting"]["rolls"]["halfSave"] and resultInput == 'n':
                        resultInput = {
                            "rollResults": resultInput,
                            "diceResults": 0
                        }
                        preTurnLogs.append(
                            logLingeringResult(effect["effect"]["resultID"][0], creature.getName(), effect["name"].lower(), resultInput))
                    else:
                        damInput = procPreEffect(effect["effect"]["spell"][0])
                        resultInput = {
                            "rollResults": resultInput,
                            "diceResults": damInput
                        }
                        try:
                            preTurnLogs.append(
                                logLingeringResult(effect["effect"]["resultID"][0], creature.getName(), effect["name"].lower(), resultInput))
                        except:
                            preTurnLogs.append({})
            else:
                raise ValueError("Invalid effect type in preTurnCheck")
            for result in preTurnLogs: #Result log all preTurnEffects.
                encounter.addResult(result)

def merge_sort_spells(spell_list):
    if len(spell_list) <= 1:
        return spell_list

    mid = len(spell_list) // 2
    left_half = merge_sort_spells(spell_list[:mid])
    right_half = merge_sort_spells(spell_list[mid:])

    return merge_spells(left_half, right_half)
def merge_spells(left, right):
    result = []
    i = j = 0

    while i < len(left) and j < len(right):
        left_spell = left[i]
        right_spell = right[j]

        # Primary key: spell level
        if left_spell.getLvl() < right_spell.getLvl():
            result.append(left_spell)
            i += 1
        elif left_spell.getLvl() > right_spell.getLvl():
            result.append(right_spell)
            j += 1
        else:
            # Secondary key: alphabetical by name (case-insensitive)
            if left_spell.getName().lower() <= right_spell.getName().lower():
                result.append(left_spell)
                i += 1
            else:
                result.append(right_spell)
                j += 1

    while i < len(left):
        result.append(left[i])
        i += 1

    while j < len(right):
        result.append(right[j])
        j += 1

    return result
def processSpellAnalytics(spellList, creature, initiative):
    actionNames = []
    actionTypes = []
    actionProbs = []
    actionEDams = []
    actionImpacts = []
    for i in range(len(spellList)):
        spellName = spellList[i].getName()
        spellProb = 0
        spellEDam = -1
        spellImpact = -1
        if spellList[i].getSelfTarget():
            spellProb = 1.0
            spellEDam = 0
        else:
            if spellList[i].getRollType().lower() == "tohit":
                spellProb = calcTotalToHitProbability(creature, spellList[i], initiative)
            elif spellList[i].getRollType().lower() == "save":
                spellProb = calcTotalSaveProbability(creature, spellList[i], initiative)
            elif spellList[i].getRollType().lower() == "autohit":
                if spellList[i].getDiceNum() == 0 and not spellList[i].getLingSaves() \
                        and not spellList[i].getLingEffects() and not spellList[i].getExtraEffect() \
                        and not (spellList[i].getSpecialNotes() and "HPCap" in spellList[i].getSpecialNotes()):
                    spellProb = 1.0
                    spellEDam = 0
                else:
                    if spellList[i].getStatusEffects() and any(
                            [se["name"] == "Summon" for se in spellList[i].getStatusEffects()]):
                        spellProb = 1.0
                        spellEDam = 0
                    else:
                        spellProb = calcTotalAutoHitProbability(creature, spellList[i], initiative)
            elif spellList[i].getRollType().lower() == "onhit":
                spellProb = calcOnHitProbability(spellList[i],
                                                 [creature.getWeapon(i) for i in range(creature.getWeaponLength())],
                                                 creature, initiative)
            if isinstance(spellProb, dict):
                spellProb["probSuccess"] = 0 if spellProb["probSuccess"] < 0 else spellProb["probSuccess"]
                spellProb["probSuccess"] = 1 if spellProb["probSuccess"] > 1 else spellProb["probSuccess"]
                probToStr = f"{spellProb["probSuccess"]}" if spellProb["probSuccess"] else f"0.0"
                probToStr += f" - {spellProb["probLingEffect"]}LE" if spellProb["probLingEffect"] else ""
                probToStr += f" - {spellProb['probExtraEffect']}EE" if spellProb["probExtraEffect"] else ""
                probToStr += f" - {spellProb['probLingSaves']}LS" if spellProb["probLingSaves"] else ""
            else:
                spellProb = 0 if spellProb < 0 else spellProb
                spellProb = 1 if spellProb > 1 else spellProb
                probToStr = spellProb
            spellProb = probToStr
            spellEDam = calcTotalExpectedDamage(creature, spellList[i], initiative) if spellEDam == -1 else spellEDam

        spellImpact = calcImpact(creature, spellList[i], spellProb, spellEDam, initiative)

        actionNames.append(spellName)
        actionTypes.append(f"Lvl {spellList[i].getLvl()} Spell" if spellList[i].getLvl() > 0 else "Cantrip")
        actionProbs.append(spellProb)
        actionEDams.append(spellEDam)
        actionImpacts.append(spellImpact)

    actions = [{"name" : actionNames[i], "prob" : actionProbs[i], "eDam" : actionEDams[i], "impact" : actionImpacts[i]} for i in range(len(actionNames))]
    return actions
def rankActions(actions):
    KEYS = ("prob", "eDam", "impact")

    SEG_RE = re.compile(
        r"^\s*(?P<a>\d*\.?\d+)\s*(?:-\s*(?P<b>\d*\.?\d+))?\s*(?P<tag>LS|LE|EE)?\s*$",
        re.IGNORECASE
    )

    def _mid(a, b):
        return (a + b) / 2.0 if b is not None else a
    def parse_prob_segments(prob_str_or_num):
        """
        Returns:
          initial: float
          parts: dict tag -> float  (tags in {"LS","LE","EE"})
        Accepts:
          - numeric: 0.75
          - "0.40 - 0.60LS - 0.20LE - 0.10EE"
          - "0.75 - 0.30LE"
        """
        if isinstance(prob_str_or_num, (int, float)):
            return float(prob_str_or_num), {}

        if not isinstance(prob_str_or_num, str):
            raise TypeError(f"Unsupported prob type: {type(prob_str_or_num)}")

        s = prob_str_or_num.strip()
        if s[0] == "-":
            s = s[1:]
        chunks = [c.strip() for c in s.split(" - ")]

        if not chunks:
            raise ValueError(f"Empty prob string: {prob_str_or_num!r}")

        # First chunk: initial (no tag)
        m0 = SEG_RE.match(chunks[0])
        if not m0:
            raise ValueError(f"Could not parse initial prob chunk: {chunks[0]!r} of {s}")

        a0 = float(m0.group("a"))
        b0 = float(m0.group("b")) if m0.group("b") is not None else None
        initial = _mid(a0, b0)

        parts = {}
        for chunk in chunks[1:]:
            m = SEG_RE.match(chunk)
            if not m:
                raise ValueError(f"Could not parse prob chunk: {chunk!r} from {prob_str_or_num!r}")

            a = float(m.group("a"))
            b = float(m.group("b")) if m.group("b") is not None else None
            tag = (m.group("tag") or "").upper()

            if tag not in {"LS", "LE", "EE"}:
                # In your normalization, extras always have tags; if not, skip or raise.
                raise ValueError(f"Missing/invalid tag in chunk: {chunk!r}")

            parts[tag] = _mid(a, b)

        return initial, parts
    def prob_score_weighted(initial, parts, weights=None):
        """
        Weighted average score (recommended starter).
        """
        if weights is None:
            weights = {"INIT": 0.70, "LS": 0.10, "LE": 0.10, "EE": 0.10}

        score = weights["INIT"] * initial
        for tag in ("LS", "LE", "EE"):
            if tag in parts:
                score += weights.get(tag, 0.0) * parts[tag]
        return score
    def prob_score_multiplicative(initial, parts):
        """
        Strict 'all must land' score (more punishing).
        """
        score = initial
        for tag in ("LS", "LE", "EE"):
            if tag in parts:
                score *= parts[tag]
        return score
    def prepare_actions_for_ranking(actions, score_mode="weighted"):
        out = []
        for a in actions:
            x = dict(a)
            x["probDisplay"] = a["prob"]

            init, parts = parse_prob_segments(a["prob"])
            x["probInit"] = init
            x["probParts"] = parts  # dict like {"LE": 0.25, "EE": 0.10}

            if score_mode == "weighted":
                x["prob"] = prob_score_weighted(init, parts)
            else:
                x["prob"] = prob_score_multiplicative(init, parts)

            if float(x["prob"]) < 0:
                x["probDisplay"] = 0
            elif float(x["prob"]) > 1.0:
                x["probDisplay"] = 1

            x["eDam"] = float(x["eDam"])
            x["impact"] = float(x["impact"])
            out.append(x)

        return out
    def pareto_front_set(actions, keys=KEYS):
        """
        Returns a set of ids() for non-dominated actions.
        a dominates b if it is >= on all keys and > on at least one key.
        """
        front_ids = set()
        for a in actions:
            dominated = False
            for b in actions:
                if a is b:
                    continue
                ge_all = all(b[k] >= a[k] for k in keys)
                gt_any = any(b[k] > a[k] for k in keys)
                if ge_all and gt_any:
                    dominated = True
                    break
            if not dominated:
                front_ids.add(id(a))
        return front_ids
    def topsis_scores_minmax(actions, keys=KEYS, weights=None, eps=1e-12):
        """
        TOPSIS scores for all actions using min-max normalization per metric.
        All metrics assumed 'higher is better'.
        Returns dict: id(action) -> score in [0, 1] (higher better).
        """
        if not actions:
            return {}

        if weights is None:
            weights = {k: 1.0 for k in keys}

        mins = {k: min(a[k] for a in actions) for k in keys}
        maxs = {k: max(a[k] for a in actions) for k in keys}

        # Normalize to [0,1], apply weights
        norm_rows = []
        for a in actions:
            row = {}
            for k in keys:
                rng = (maxs[k] - mins[k])
                if abs(rng) < eps:
                    v = 0.0  # all same -> doesn't matter
                else:
                    v = (a[k] - mins[k]) / (rng + eps)
                row[k] = v * weights[k]
            norm_rows.append((a, row))

        ideal_best = {k: max(r[k] for _, r in norm_rows) for k in keys}  # usually == weights[k]
        ideal_worst = {k: min(r[k] for _, r in norm_rows) for k in keys}  # usually == 0

        scores = {}
        for a, r in norm_rows:
            d_pos = math.sqrt(sum((r[k] - ideal_best[k]) ** 2 for k in keys))
            d_neg = math.sqrt(sum((r[k] - ideal_worst[k]) ** 2 for k in keys))
            score = d_neg / (d_pos + d_neg + eps)
            scores[id(a)] = score

        return scores
    def rank_all_actions(actions, weights=None):
        """
        Returns a full ranked list (best -> worst) of *every* action dict,
        adding:
          - pareto: bool
          - topsis: float
          - overallRank: int
        Sorting: pareto actions first, then TOPSIS score desc.
        """
        front_ids = pareto_front_set(actions)
        scores = topsis_scores_minmax(actions, weights=weights)

        enriched = []
        for a in actions:
            x = dict(a)  # don't mutate originals
            x["pareto"] = (id(a) in front_ids)
            x["topsis"] = scores.get(id(a), 0.0)
            enriched.append(x)

        enriched.sort(key=lambda x: (x["pareto"], x["topsis"]), reverse=True)

        for i, x in enumerate(enriched, start=1):
            x["overallRank"] = i

        return enriched

    # TODO: Double check accuracy of ranking logic
    overallRankings = prepare_actions_for_ranking(actions)
    overallRankings = rank_all_actions(overallRankings, weights={"prob": 1.0, "eDam": 1.0, "impact": 1.25})
    print("POST PROBS")
    console.print(overallRankings)
    return overallRankings
def monsterTurn(creature, initiative):
    def loadSpells(spells):
        with open("CoreEngine/data/spell_list.json", "r") as sf:
            spell_list = json.load(sf)
            spell_names = [spell["spellname"].lower() for spell in spell_list]
        for si, spell in enumerate(spells):
            if spell["name"].lower() in spell_names:
                spells[si]["spellData"] = spell_list[spell_names.index(spell["name"].lower())]
        return spells
    if endOfEncounter(initiative):
        return {}
    actionNames = []
    actionTypes = []
    actionProbs = []
    actionEDams = []
    actionImpacts = []
    actions = []

    defineBasicActions(actionNames, actionTypes, actionProbs,
                       actionEDams, actionImpacts, creature, initiative)
    if creature.isCaster():
        monSpells = [creature.getSpell(i) for i in range(creature.getSpellLength())]
        monSpells = loadSpells(monSpells)
        for spell in monSpells:
            if "spellData" not in spell:
                continue
            spellList = merge_sort_spells(monSpells)
            spellData = processSpellAnalytics(spellList, creature, initiative)
            actions.append(spellData)
    monActions = [creature.getAction(i) for i in range(creature.getActionLength())]
    for i, monAction in enumerate(monActions):
        if monAction.isBadObj():
            actions.append({"name" : monAction.getName(), "prob" : 0, "eDam": 0, "impact" : 0})
            continue

        actionName = monAction.getName()
        actionProb = 0
        actionEDam = -1
        actionImpact = -1
        if monAction.getSelfTarget():
            actionProb = 1.0
            actionEDam = 0
        else:
            if monAction.getRollType().lower() == "tohit":
                actionProb = calcTotalToHitProbability(creature, monAction, initiative)
            elif monAction.getRollType().lower() == "save":
                actionProb = calcTotalSaveProbability(creature, monAction, initiative)
            if isinstance(actionProb, dict):
                actionProb["probSuccess"] = 0 if actionProb["probSuccess"] < 0 else actionProb["probSuccess"]
                actionProb["probSuccess"] = 1 if actionProb["probSuccess"] > 1 else actionProb["probSuccess"]
                probToStr = f"{actionProb["probSuccess"]}" if actionProb["probSuccess"] else f"0.0"
                probToStr += f" - {actionProb["probLingEffect"]}LE" if actionProb["probLingEffect"] else ""
                probToStr += f" - {actionProb['probExtraEffect']}EE" if actionProb["probExtraEffect"] else ""
                probToStr += f" - {actionProb['probLingSaves']}LS" if actionProb["probLingSaves"] else ""
            else:
                actionProb = 0 if actionProb < 0 else actionProb
                actionProb = 1 if actionProb > 1 else actionProb
                probToStr = actionProb
            actionProb = probToStr
            actionEDam = calcTotalExpectedDamage(creature,
                monAction, initiative) if actionEDam == -1 else actionEDam
        actionImpact = calcImpact(creature, monAction, actionProb,
                    actionEDam, initiative)

        actionNames.append(actionName)
        actionProbs.append(actionProb)
        actionEDams.append(actionEDam)
        actionImpacts.append(actionImpact)

    actions.extend([{"name": actionNames[i], "prob": actionProbs[i], "eDam": actionEDams[i], "impact": actionImpacts[i]} for
               i in range(len(actionNames))])
    print("PRE PROBS")
    console.print(actions)
    return rankActions(actions)
def playerTurn(player, initiative):
    if endOfEncounter(initiative):
        return {}
    actionNames = []
    actionTypes = []
    actionProbs = []
    actionEDams = []
    actionImpacts = []

    defineBasicActions(actionNames, actionTypes, actionProbs,
                       actionEDams, actionImpacts,  player,initiative)
    if player.getWeaponLength() > 0:
        for i in range(player.getWeaponLength()):
            try:
                weaponProb = calcTotalToHitProbability(player, player.getWeapon(i), initiative)["probSuccess"]
            except:
                try:
                    weaponProb = int(calcTotalToHitProbability(player, player.getWeapon(i), initiative))
                except:
                    weaponProb = 0
            weaponEDam = calcTotalExpectedDamage(player, player.getWeapon(i), initiative)
            weaponImpact = calcImpact(player, player.getWeapon(i), weaponProb, weaponEDam, initiative)
            actionNames.append(player.getWeapon(i).getName())
            actionTypes.append("Weapon")
            actionProbs.append(weaponProb)
            actionEDams.append(weaponEDam)
            actionImpacts.append(weaponImpact)
    actions = [{"name": actionNames[i], "prob": actionProbs[i], "eDam": actionEDams[i], "impact": actionImpacts[i]} for
               i in range(len(actionNames))]
    if player.getSpellLength() > 0:
        spellList = [player.getSpell(i) for i in range(player.getSpellLength())]
        spellList = merge_sort_spells(spellList)
        spellActions = processSpellAnalytics(spellList, player, initiative)
        actions.extend(spellActions)
    console.print(actions)
    return rankActions(actions)
def setActiveInitiative(encounter):
    initiative = copy.deepcopy(encounter.getInitiative())
    for creature in initiative:
        # Add creature statblock to their associated turn
        # SHALLOW COPY OF MONSTER/PLAYER OBJECTS - Changes to creature["Statblock"] affect associated object in encounter
        if creature["turnType"] == "Player":
            for i in range(encounter.playerSize()):
                if creature["name"].lower() == encounter.getPlayer(i).getName().lower():
                    creature["Statblock"] = encounter.getPlayer(i)
                    break
        elif creature["turnType"] == "Monster":
            for i in range(encounter.monsterSize()):
                if creature["name"].lower() == encounter.getMonster(i).getName().lower():
                    creature["Statblock"] = encounter.getMonster(i)
                    break
    return initiative
def runEncounter(encounter):
    if len(encounter.getInitiative()) == 0:
        setInitiative(encounter)
    initiative = setActiveInitiative(encounter)

    hasLegActions = False
    # for creature in initiative:
    #     #Population monsters that have legendary actions.
    #     if creature["turnType"] == "Monster":
    #         if creature["Statblock"].hasLegAction:
    #             hasLegActions = True
    #             break

    printEncounterState(initiative)
    continueEncounter = True

    idx = 0
    found = False
    while idx < len(initiative) and not found:  # Sets the start to the current turn.
        if initiative[idx]["currentTurn"]:
            found = True
        else:
            idx += 1
    while not encounter.isComplete() and continueEncounter:
        print(f"{initiative[idx]["name"]}'s turn!")
        if initiative[idx]["turnType"] != "LairAction":
            preTurnCheck(initiative[idx], encounter, initiative)
        if initiative[idx]["turnType"] == "Player":
            if any(condition.lower() in [c["cond"].lower() for c in initiative[idx]["Statblock"].getActiveConditions()] for condition in
                   ["downed", "stabilized", "dead", "incapacitated", "paralyzed", "petrified", "stunned", "unconscious"]):
                pass
            elif initiative[idx]["Statblock"].isActiveCondition("Charmed") or initiative[idx]["Statblock"].isActiveStatusEffect("SwitchSides"):
                enemyTurnMANUAL(initiative, encounter)
            else:
                encounter.addResult(playerTurn(initiative[idx]["Statblock"], initiative, encounter))
        elif initiative[idx]["turnType"] == "Monster":
            procTurn = input("Any creatures affected? Y/N").lower()
            while procTurn != "y" and procTurn != "n":
                print("Please input only Y/N")
                procTurn = input("Any creatures affected? Y/N").lower()
            if procTurn == "y":
                enemyTurnMANUAL(initiative, encounter)
        elif initiative[idx]["turnType"] == "LairAction":
            procTurn = input("Any creatures affected? Y/N").lower()
            while procTurn != "y" and procTurn != "n":
                print("Please input only Y/N")
                procTurn = input("Any creatures affected? Y/N").lower()
            if procTurn == "y":
                enemyTurnMANUAL(initiative, encounter)
        else:
            raise ValueError("Turn type is not recognized.")
        printCreatureStats_DEBUG(initiative)
        if not endOfEncounter(initiative): #Neither creature types are dead
            initiative[idx]["currentTurn"] = False
            if idx + 1 == len(initiative):  # End of the list
                idx = 0
            else:
                idx += 1
            initiative[idx]["currentTurn"] = True

            # TODO: reimplement legendary actions here, but with actual legActions from enemy statblock.
            # if hasLegActions:
            #     executeLegAction = input("Any legendary actions used? Y/N").lower()
            #     while executeLegAction != "y" and executeLegAction != "n":
            #         print("Please input only Y/N")
            #         executeLegAction = input("Any legendary actions used? Y/N").lower()
            #     if executeLegAction == "y":
            #         enemyTurn(initiative, encounter)
        else:
            encounter.setComplete(True)
            continueEncounter = False
            exportEncounterResultsToExcel(encounter)
        if not encounter.isComplete() and not endOfEncounter(initiative):
            endProgram = input("Continue Encounter? Y/N").lower()
            while endProgram != "y" and endProgram != "n":
                print("Please input only Y/N")
                endProgram = input("Continue Encounter? Y/N").lower()
            if endProgram == "n":
                continueEncounter = False
            else:
                inIdx = 0
                for turn in encounter.getInitiative():
                    turn["currentTurn"] = initiative[inIdx]["currentTurn"]
                    inIdx += 1
                saveEncounter(encounter)
        else:
            encounter.setComplete(True)
            continueEncounter = False

    #END OF PROGRAM - SAVE ENCOUNTER
    inIdx = 0
    for turn in encounter.getInitiative():
        turn["currentTurn"] = initiative[inIdx]["currentTurn"]
        inIdx += 1
    saveEncounter(encounter)
def main():
    # try:
    #     print("Welcome to Encounter Simulator!")
    #
    #     if TEST_MODE:
    #         print("[TEST MODE ENABLED]")
    #
    #     print("\nType 1 to load a previous Encounter")
    #     print("Type 2 to create a new Encounter")
    #     encChoice = input().strip()
    #     while encChoice != "1" and encChoice != "2":
    #         print("Type 1 to load a previous Encounter")
    #         print("Type 2 to create a new Encounter")
    #         encChoice = input().strip()
    #
    #     if encChoice == "1":
    #         encIdx = chooseEncounter()
    #         encounter = loadEncounter(encIdx)
    #     else:
    #         encounter = createEncounter()
    #         saveEncounter(encounter)
    #     runEncounter(encounter)
    # finally:
    #     # --- Clean shutdown for pytest and normal runs ---
    #     try:
    #         if hasattr(console, "file"):
    #             # Only flush, never close under pytest
    #             console.file.flush()
    #             if "pytest" not in sys.modules:
    #                 console.file.close()
    #     except Exception:
    #         pass
    #
    #     # Only perform hard exit for pytest-controlled sessions
    #     if "pytest" in sys.modules:
    #         import atexit
    #         atexit.register(lambda: os._exit(0))
    #
    #     print("[SHUTDOWN] Console and file handles flushed. Program terminated cleanly.")
    with open(ENCOUNTER_LIST_FILE, "r") as ef:
        encounters = json.load(ef)
        encounter = {}
        for enc in encounters:
            if enc["eid"] == "a4695c2":
                encounter = enc
                break
        encounter = loadEncounter(encounter)
        # console.print(monsterTurn(encounter.getMonster(0), setActiveInitiative(encounter)))
        console.print(playerTurn(encounter.getPlayer(0), setActiveInitiative(encounter)))

if __name__ == "__main__":
    main()